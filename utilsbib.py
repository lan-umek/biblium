from __future__ import annotations

# --- Standard library ---
import os
import re
import math
import datetime
from functools import reduce
from collections import Counter
from itertools import chain, combinations
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

# --- Data handling ---
import pandas as pd
import numpy as np

# --- Text processing & NLP ---
import nltk
from nltk.stem import WordNetLemmatizer
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.sentiment import SentimentIntensityAnalyzer
from rapidfuzz import fuzz
import difflib

# --- Machine Learning & Clustering ---
from sklearn.feature_extraction.text import (
    CountVectorizer,
    TfidfVectorizer,
    TfidfTransformer,
    ENGLISH_STOP_WORDS,
)
from sklearn.decomposition import (
    TruncatedSVD,
    NMF,
    LatentDirichletAllocation,
)
from sklearn.manifold import MDS, TSNE
from sklearn.feature_selection import SelectKBest, chi2, mutual_info_classif
from sklearn.metrics import (
    silhouette_score,
    davies_bouldin_score,
    r2_score,
    mean_squared_error,
    pairwise_distances,
)
from sklearn.preprocessing import (
    StandardScaler,
    OneHotEncoder,
)
from sklearn.cluster import (
    KMeans,
    AgglomerativeClustering,
    DBSCAN,
    SpectralClustering,
    SpectralCoclustering,
)
from sklearn.utils.extmath import randomized_svd

# --- Statistics ---
from scipy.stats import (
    rankdata,
    entropy,
    skew,
    kurtosis,
    normaltest,
    ks_2samp,
    f_oneway,
    kruskal,
    shapiro,
    fisher_exact,
    chi2_contingency,
    pearsonr,
)
from statsmodels.stats.multitest import multipletests
from scipy.spatial.distance import pdist
from scipy.cluster.hierarchy import (
    linkage,
    fcluster
)
import scipy.sparse.linalg as spla
from scipy.sparse import csr_matrix

# --- Excel and reporting ---
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- Network analysis ---
import networkx as nx

try:
    import igraph as ig
except ImportError:
    ig = None

try:
    import community as community_louvain
except ImportError:
    community_louvain = None

# --- Advanced data analysis ---
try:
    import prince
except ImportError:
    prince = None

# --- UMAP (dimensionality reduction) ---
try:
    import umap
except ImportError:
    umap = None

# --- spaCy (NLP) ---
try:
    import spacy
    nlp = spacy.load("en_core_web_sm", disable=["parser", "ner"])
except Exception:
    nlp = None

# --- OpenAI (if needed) ---
try:
    import openai
except ImportError:
    openai = None



# general
fd = os.path.dirname(__file__)


# folder manipulation

def make_folder(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)
        
def make_folders(folders):
    for folder in folders:
        make_folder(folder)
        
# misc string manipulation fucttions


# misc object manipulationg functions

def rename_attributes(obj, rename_dict):
    for old_attr, new_attr in rename_dict.items():
        if hasattr(obj, old_attr):
            setattr(obj, new_attr, getattr(obj, old_attr))
            delattr(obj, old_attr)

# add document labels

def add_document_labels_abbrev(df):
    """
    Adds 'Document Short Label' and 'Document Label' using 'Abbreviated Source Title' instead of 'Source title'.
    
    'Document Short Label': 'FirstAuthor, AbbreviatedSourceTitle (Year)'
    'Document Label': Adds first three words of title after the short label.

    Parameters:
        df (pd.DataFrame): Input dataframe.

    Returns:
        pd.DataFrame: DataFrame with updated label columns.
    """
    def extract_first_author(authors):
        if pd.isna(authors) or not isinstance(authors, str):
            return ""
        return authors.split(";")[0].strip()

    def extract_first_three_words(title):
        if pd.isna(title) or not isinstance(title, str):
            return ""
        return " ".join(title.strip().split()[:3])

    df = df.copy()

    first_authors = df["Authors"].fillna("").apply(extract_first_author)
    source_titles = df["Abbreviated Source Title"].fillna("")
    years = df["Year"].fillna("").astype(str).replace("nan", "")
    titles = df["Title"].fillna("").apply(extract_first_three_words)

    df["Document Short Label"] = first_authors + ", " + source_titles + " (" + years + ")"
    df["Document Label"] = df["Document Short Label"] + ": " + titles

    return df

# misc manipulations with dataframes


def merge_on_key(df1, df2, key_column):
    """
    Merge two pandas DataFrames on a common key column.

    The resulting DataFrame will:
    - Contain only rows where the key column values exist in both df1 and df2 (intersection).
    - Include all columns from both DataFrames (column union).
    - Preserve the column order and values from df1 for overlapping columns.

    Parameters:
        df1 (pd.DataFrame): The first DataFrame (typically larger, with fewer columns).
        df2 (pd.DataFrame): The second DataFrame (subset of rows, with additional columns).
        key_column (str): The name of the column to align and merge on.

    Returns:
        pd.DataFrame: A merged DataFrame with intersected rows and unioned columns.
    """
    # Step 1: Find intersection of keys
    common_keys = df1[key_column].isin(df2[key_column])
    df1_common = df1[common_keys]
    df2_common = df2[df2[key_column].isin(df1[key_column])]

    # Step 2: Avoid duplicate columns (except key)
    df2_extra_cols = [col for col in df2_common.columns if col not in df1.columns or col == key_column]

    # Step 3: Merge on the key column
    merged = pd.merge(
        df1_common,
        df2_common[df2_extra_cols],
        on=key_column,
        how="inner",
        suffixes=("", "_df2")  # Prevent name clash but keep df1 priority
    )

    return merged

def combine_item_dataframes(df_list, df_names=None):
    """
    Combines a list of dataframes with "Item", "Number of documents",
    "Fractional number of documents", and "Proportion of documents"
    into a single dataframe where each row corresponds to one input dataframe.
    Columns represent item-metric pairs. Missing values in populated rows
    are filled with 0. Completely empty dataframes result in rows filled with NaN.

    Parameters:
        df_list (list of pd.DataFrame): List of input dataframes.
        df_names (list of str, optional): Row index names for the result.

    Returns:
        pd.DataFrame: Combined dataframe with one row per input dataframe.
    """
    combined_rows = []
    all_columns = set()

    # First pass: process non-empty frames to collect all possible column names
    temp_rows = []
    for df in df_list:
        if not df.empty:
            df = df[["Item", "Number of documents", "Fractional number of documents", "Proportion of documents"]].copy()
            df_row = (
                df.set_index("Item")
                  .stack()
                  .rename_axis(["Item", "Metric"])
                  .reset_index(name="Value")
                  .assign(Col=lambda d: d["Item"] + " [" + d["Metric"] + "]")
                  .set_index("Col")["Value"]
            )
            all_columns.update(df_row.index)
            temp_rows.append(df_row)
        else:
            temp_rows.append(None)

    # Second pass: build aligned rows
    all_columns = sorted(all_columns)
    for df_row in temp_rows:
        if df_row is None:
            combined_rows.append(pd.Series(index=all_columns, dtype=float))  # All NaNs
        else:
            combined_rows.append(df_row.reindex(all_columns, fill_value=0))

    result = pd.DataFrame(combined_rows)

    if df_names:
        result.index = df_names

    return result


# missings

def check_missing_values(df, columns=None):
    """
    Returns the number and proportion of missing values for each column in a given list of columns in a pandas DataFrame.
    
    Parameters:
        df (pandas.DataFrame): The pandas DataFrame to check for missing values.
        columns (list): A list of column names to check for missing values.
    
    Returns:
        pandas.DataFrame: A DataFrame containing the number and proportion of missing values for each column in the given list of columns, as well as a new column indicating the quality of each column based on the number of missing values.
    """
    if columns is None:
        columns = df.columns
    else:
        columns = [c for c in columns if c in df.columns]
    
    missing_values_list = []
    missing_d = {}
    
    for column in columns:
        missing_count = df[column].isna().sum()
        proportion_missing = missing_count / len(df)
        missing_d[column] = df[column].isna()
        
        # Set the missing value quality based on the proportion of missing values
        if proportion_missing == 0:
            quality = "Excellent (0%)"
        elif proportion_missing < 0.1:
            quality = "Good (<10%)"
        elif proportion_missing < 0.5:
            quality = "Fair (10-50%)"
        elif proportion_missing < 0.9:
            quality = "Poor (50-90%)"
        else:
            quality = "Bad (>90%)"
        
        missing_values_list.append({
            "Column": column,
            "Missing Values": missing_count,
            "Proportion": proportion_missing,
            "Missing Value Quality": quality
        })
    
    # Convert the list of dictionaries to a DataFrame
    missing_values_df = pd.DataFrame(missing_values_list)
    
    # Sort the DataFrame by the proportion of missing values in ascending order
    missing_values_df = missing_values_df.sort_values("Proportion")
    
    return missing_values_df, missing_d


# filtering




def filter_dataframe(
    df: pd.DataFrame,
    filters: dict = {},
    bradford_filter: str = "all"
) -> pd.DataFrame:
    """
    Filters a DataFrame based on multiple criteria per column, with optional Bradford's Law zone filtering.

    Supported filter keys per column:
    - "regex_include": list of regex patterns (OR-matched)
    - "regex_exclude": list of regex patterns (OR-matched)
    - "include": list of exact values to include
    - "exclude": list of exact values to exclude
    - "min": minimum value (for numeric/date columns)
    - "max": maximum value (for numeric/date columns)

    Parameters:
    - df (pd.DataFrame): The input DataFrame to filter.
    - filters (dict): Dictionary of column-wise filtering rules. Default is {} (no filtering).
    - bradford_filter (str): One of {"all", "core", "core+zone2"} to control filtering by Bradford zones.

    Returns:
    - pd.DataFrame: Filtered DataFrame with index reset.

    Example:
    --------
    To filter only Articles from core sources published after 2000 with at least 1 citation:

    >>> filters = {
    ...     "Document Type": {"include": ["Article"]},
    ...     "Year": {"min": 2000},
    ...     "Cited by": {"min": 1}
    ... }
    >>> filtered_df = filter_dataframe(df, filters, bradford_filter="core")
    """
    mask = pd.Series(True, index=df.index)

    for col, criteria in filters.items():
        if col not in df.columns:
            continue

        col_data = df[col].astype(str) if df[col].dtype == object else df[col]

        if "regex_include" in criteria:
            pattern = "|".join(criteria["regex_include"])
            mask &= col_data.str.contains(pattern, na=False, regex=True)

        if "regex_exclude" in criteria:
            pattern = "|".join(criteria["regex_exclude"])
            mask &= ~col_data.str.contains(pattern, na=False, regex=True)

        if "include" in criteria:
            mask &= col_data.isin(criteria["include"])

        if "exclude" in criteria:
            mask &= ~col_data.isin(criteria["exclude"])

        if "min" in criteria:
            mask &= pd.to_numeric(col_data, errors="coerce") >= criteria["min"]

        if "max" in criteria:
            mask &= pd.to_numeric(col_data, errors="coerce") <= criteria["max"]

    filtered_df = df[mask].copy()

    # Apply Bradford filtering if needed
    if bradford_filter in {"core", "core+zone2"}:
        if "Source title" not in filtered_df.columns:
            raise ValueError('"Source title" column is required for Bradford filtering.')

        source_counts = filtered_df["Source title"].value_counts()
        total_sources = len(source_counts)
        third = math.ceil(total_sources / 3)

        core_sources = source_counts.index[:third]
        zone2_sources = source_counts.index[third:2 * third]

        if bradford_filter == "core":
            allowed_sources = set(core_sources)
        elif bradford_filter == "core+zone2":
            allowed_sources = set(core_sources).union(set(zone2_sources))

        filtered_df = filtered_df[filtered_df["Source title"].isin(allowed_sources)]

    return filtered_df.reset_index(drop=True)

# misc functions - citations

def compute_average_citations_per_year(df, year_col="Year", citations_col="Cited by"):
    """
    Compute the average number of citations per document for each year.

    Parameters:
        df (pd.DataFrame): DataFrame containing bibliometric records.
        year_col (str): Name of the column indicating the publication year.
        citations_col (str): Name of the column indicating the number of citations.

    Returns:
        pd.DataFrame: A DataFrame with "Year", "Number of Documents", 
                      "Total Citations", and "Average Citations per Document".
    """
    grouped = df.groupby(year_col).agg(
        {"{}".format(citations_col): ["count", "sum"]}
    )
    grouped.columns = ["Number of Documents", "Total Citations"]
    grouped = grouped.reset_index()

    grouped["Average Citations per Document"] = (
        grouped["Total Citations"] / grouped["Number of Documents"]
    )

    return grouped

# Authors manipulation

def extract_author_mappings(df, column):
    """
    Extracts ID-to-author and author-to-ID mappings from a dataframe column.

    Parameters:
        df (pd.DataFrame): The dataframe containing the author strings.
        column (str): The column name where author strings are stored.

    Returns:
        tuple: Two dictionaries (id_to_author, author_to_id).
    """
    id_to_author = {}
    author_to_id = {}

    for author_str in df[column].dropna():
        for entry in author_str.split("; "):
            if not entry.strip():
                continue
            try:
                name, id_with_parens = entry.rsplit(" (", 1)
                author_id = id_with_parens.rstrip(")")
                id_to_author[author_id] = name
                author_to_id[name] = author_id
            except ValueError:
                continue

    return id_to_author, author_to_id

def split_author_id(df):
    """
    Splits a column named 'Author ID' in the DataFrame into two new columns:
    'Author' and 'ID'.

    The function assumes that each entry in the 'Author ID' column is a string
    in the format 'Author Name (ID)'.

    Parameters:
        df (pd.DataFrame): A pandas DataFrame with a column named 'Author ID'.

    Returns:
        pd.DataFrame: The same DataFrame with two additional columns:
                      - 'Author': the name of the author.
                      - 'ID': the ID extracted from parentheses as a string.
    """
    df["Author"] = df["Author ID"].map(lambda x: x.split(" (")[0])
    df["ID"] = df["Author ID"].map(lambda x: x.split(" (")[1][:-1])
    return df

# Collaboration Index



def collaboration_index(
    df: pd.DataFrame,
    author_col: str = "Author(s) ID",
    sep: str = ";"
) -> float:
    """
    Compute the Collaboration Index (CI) for a set of articles where authors are stored
    as a single string per row, separated by a delimiter.

    The Collaboration Index is defined as:
        CI = (Total number of author-instances in multi-authored articles)
             / (Total number of multi-authored articles)

    Only articles with more than one author are considered.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing article records, with one column listing authors as a string.
    author_col : str, default "authors"
        Name of the column in `df` where each row is a delimiter-separated string of author names.
    sep : str, default ";"
        Delimiter used to separate author names in the string.

    Returns
    -------
    float
        The Collaboration Index (average co-authors per multi-authored article).
        Returns 0.0 if there are no multi-authored articles.
    """
    # Parse the author strings into lists
    parsed = df[author_col].astype(str).map(lambda s: [a.strip() for a in s.split(sep) if a.strip()])
    
    # Filter to multi-authored articles
    multi = parsed[parsed.map(len) > 1]
    num_multi_articles = len(multi)
    if num_multi_articles == 0:
        return 0.0

    # Sum total authors across those articles
    total_authors = multi.map(len).sum()

    # Compute Collaboration Index
    return total_authors / num_multi_articles




# merging sciences

def enrich_bibliometric_data(biblio_df, asjc_map_df, asjc_meta_df):
    """
    Enrich a bibliometric dataframe with ASJC codes and their metadata (Field, Area, Science).

    Parameters:
        biblio_df (pd.DataFrame): Bibliometric data with column "Source title".
        asjc_map_df (pd.DataFrame): Journal-to-code mapping with "Source title" and "All Science Journal Classification Codes (ASJC)".
        asjc_meta_df (pd.DataFrame): ASJC metadata with "code", "Field", "Area", and "Science".

    Returns:
        pd.DataFrame: Enriched dataframe with additional columns for ASJC code, Field, Area, and Science.
    """
    if "Science" in biblio_df.columns:
        print("Sciences already in the dataset")
        return biblio_df
    
    # Rename first column of asjc_map_df to "Source title"
    asjc_map_df.columns = ["Source title"] + list(asjc_map_df.columns[1:])

    biblio_df["source title orig"] = biblio_df["Source title"]
    # Normalize titles for matching
    biblio_df["Source title"] = biblio_df["Source title"].astype(str).str.strip().str.lower()
    asjc_map_df["Source title"] = asjc_map_df["Source title"].astype(str).str.strip().str.lower()

    # Merge on source title only
    merged = biblio_df.merge(asjc_map_df, how="left", on="Source title")

    # Clean and split codes
    merged["All Science Journal Classification Codes (ASJC)"] = merged["All Science Journal Classification Codes (ASJC)"].fillna("").str.strip(";")
    exploded = merged.copy()
    exploded = exploded.assign(code=exploded["All Science Journal Classification Codes (ASJC)"].str.split(";")).explode("code")
    exploded["code"] = exploded["code"].str.strip()

    # Ensure consistent dtypes for merging
    asjc_meta_df["code"] = asjc_meta_df["code"].astype(str)
    exploded["code"] = exploded["code"].astype(str)

    # Merge with ASJC metadata
    enriched = exploded.merge(asjc_meta_df, on="code", how="left")
       
    # Optionally re-aggregate codes and metadata into lists or semicolon-separated strings
    agg = enriched.groupby("Source title", as_index=False).agg({
        "code": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Field": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Area": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Science": lambda x: "; ".join(sorted(set(x.dropna())))
    })

    # Merge aggregated results back into original dataframe
    result = biblio_df.merge(agg, on="Source title", how="left")

    result["Source title"] = biblio_df["source title orig"]
    result = result.drop(columns=["source title orig"])
    biblio_df = biblio_df.drop(columns=["source title orig"])

    return result


# countries

fd = os.path.dirname(__file__)        
df_countries = pd.read_excel(fd + "\\additional files\\countries.xlsx")
domain_dct = df_countries.set_index("Internet domain").to_dict()["Name"]
c_off_dct = df_countries.set_index("Official name").to_dict()["Name"]
code_dct = df_countries.set_index("Name").to_dict()["Code"]
country_iso3_dct =  df_countries.set_index("Name").to_dict()["ISO-3"]
continent_dct = df_countries.set_index("Name").to_dict()["Continent"]

df_countries_un_iso = df_countries.drop_duplicates(subset="ISO-3")
code_to_coords = df_countries_un_iso[["ISO-3", "latitude", "longitude"]].set_index("ISO-3")[["latitude", "longitude"]].to_dict(orient="index")

       
l_countries = list(df_countries["Name"])
eu_countries = list(df_countries[df_countries["EU"] == 1]["Name"])

def correct_country_name(s):
    """
    Return the corrected country name based on known lists and mappings.

    Parameters:
        s (str): Input country name.

    Returns:
        str: Corrected country name if recognized, empty string otherwise.
    """
    if not isinstance(s, str):
        return ""
    if s in l_countries:
        return s
    return c_off_dct.get(s, "")

def split_ca(s):
    """
    Split a Scopus corresponding author string into name, affiliation, and country.

    Parameters:
        s (str): Raw Scopus corresponding author string.

    Returns:
        tuple: (corresponding author, affiliation, country) or (np.nan, np.nan, np.nan) if parsing fails.
    """
    try:
        ca, long_aff = s.split("; ", 1)
        parts = long_aff.split(", ")
        return ca, parts[0], parts[-1]
    except Exception:
        return np.nan, np.nan, np.nan

def parse_mail(s):
    """
    Attempt to extract the country based on the email domain.

    Parameters:
        s (str): Full string that may contain an email.

    Returns:
        str or np.nan: Country inferred from email domain or np.nan if not found.
    """
    if "@" in s:
        domain = s.split("@")[1].split(" ")[0].split(".")[-1]
        return domain_dct.get(domain, np.nan)
    return np.nan

def get_ca_country_scopus(s, l_countries=l_countries):
    """
    Extract the country of the corresponding author from a Scopus entry.

    Parameters:
        s (str): Scopus corresponding author string.
        l_countries (list): List of recognized country names.

    Returns:
        str or np.nan: Extracted or inferred country name.
    """
    ca, aff, country = split_ca(s)

    if country not in l_countries:
        if isinstance(country, str):
            matches = [c for c in l_countries if c in country]
            if len(matches) == 1:
                country = matches[0]
            else:
                country = parse_mail(s)
        elif isinstance(s, str):
            matches = [c for c in l_countries if c in s]
            if len(matches) == 1:
                country = matches[0]
            else:
                country = parse_mail(s)

    return country
        
def get_ca_country_wos(s, l_countries=l_countries):
    """
    Extract the corresponding author's country from a WoS entry string.

    Parameters:
        s (str): Raw string from WoS corresponding author field.
        l_countries (list): List of valid country names.

    Returns:
        str or np.nan: Extracted country name or np.nan if not recognized.
    """
    if not isinstance(s, str):
        return np.nan

    if "USA" in s:
        return "United States"
    
    uk_terms = ["England", "Scotland", "Wales", "Northern Ireland", "Great Britain"]
    if any(term in s for term in uk_terms):
        return "United Kingdom"

    country = s.split(", ")[-1].replace(".", "")
    if country not in l_countries:
        try:
            country = c_off_dct.get(country, np.nan)
        except Exception:
            print(f"Unrecognized country: {country}")
            return np.nan
        if country not in l_countries:
            return np.nan

    return country

def get_ca_country(s, db, l_countries=l_countries):
    """
    Determine the country of the corresponding author based on the source database.

    Parameters:
        s (str): Raw corresponding author string.
        db (str): Name of the database ('scopus' or 'wos').
        l_countries (list): List of valid country names.

    Returns:
        str or np.nan: Extracted or inferred country name.
    """
    db = db.lower()
    if db == "scopus":
        return get_ca_country_scopus(s, l_countries=l_countries)
    elif db == "wos":
        return get_ca_country_wos(s, l_countries=l_countries)
    return np.nan

def add_ca_country_df(df, db):
    """
    Add a 'CA Country' column to a DataFrame based on corresponding author information.

    Parameters:
        df (pd.DataFrame): DataFrame with corresponding author information.
        db (str): Database name ('scopus' supported).

    Returns:
        pd.DataFrame: DataFrame with added 'CA Country' column if applicable.
    """
    if db.lower() == "scopus" and "Correspondence Address" in df.columns:
        df["CA Country"] = df["Correspondence Address"].map(get_ca_country_scopus)
    else:
        print("Not supported yet")
    return df        


# get all countries

def extract_countries_from_affiliations(df, aff_column="Affiliations", return_matrix=True):
    """
    Extracts valid countries from the affiliations column and computes collaboration metrics.

    This function processes the specified affiliations column to:
    - Extract country names from the last comma-separated segment of each affiliation.
    - Validate and normalize each country using the globally defined correct_country_name() function.
    - Add three new columns:
        - "Countries of Authors Multiple": all valid country names found (can repeat), joined by "; ".
        - "Countries of Authors": unique list of valid country names, joined by "; ".
        - "Countries Count": number of unique valid countries per record.
    - Compute a symmetric country collaboration matrix (co-authorships across countries).

    Parameters:
    df (pd.DataFrame): Input DataFrame containing the affiliation data.
    aff_column (str): Name of the column with affiliation strings (default: "Affiliations").
    return_matrix (bool): Whether to compute and return the country collaboration matrix (default: True).

    Returns:
    tuple: (Updated DataFrame, Collaboration matrix as a symmetric DataFrame or empty DataFrame if not computed)
    """


    multiple_countries = []
    unique_countries = []
    country_counts = []

    for affil in df[aff_column].fillna(""):
        entries = [entry.strip() for entry in affil.split(";")]
        countries_raw = [entry.split(",")[-1].strip() for entry in entries if "," in entry]
        countries_checked = [correct_country_name(c) for c in countries_raw]
        valid_countries = [c for c in countries_checked if c]
        unique_set = sorted(set(valid_countries))

        multiple_countries.append("; ".join(valid_countries))
        unique_countries.append("; ".join(unique_set))
        country_counts.append(len(unique_set))

    df = df.copy()
    df["Countries of Authors Multiple"] = multiple_countries
    df["Countries of Authors"] = unique_countries
    df["Countries Count"] = country_counts

    matrix_df = pd.DataFrame()

    if return_matrix:
        matrix_counter = Counter()
        for country_str in unique_countries:
            countries = [c.strip() for c in country_str.split(";") if c.strip()]
            if len(countries) > 1:
                for pair in combinations(sorted(set(countries)), 2):
                    matrix_counter[pair] += 1

        if matrix_counter:
            all_countries = sorted(set(c for pair in matrix_counter for c in pair))
            matrix_df = pd.DataFrame(0, index=all_countries, columns=all_countries)
            for (c1, c2), count in matrix_counter.items():
                matrix_df.loc[c1, c2] = count
                matrix_df.loc[c2, c1] = count

    return df, matrix_df

# links

def build_links_from_matrix(matrix_df, min_weight=1):
    """
    Constructs a DataFrame of collaboration links from a symmetric matrix.

    Parameters:
    matrix_df (pd.DataFrame): Symmetric collaboration matrix.
    min_weight (int): Minimum weight to include a link (default = 1).

    Returns:
    pd.DataFrame: DataFrame with columns: "source", "target", "weight".
    """
    if matrix_df.empty:
        return pd.DataFrame(columns=["source", "target", "weight"])

    links = []
    for i, source in enumerate(matrix_df.index):
        for j, target in enumerate(matrix_df.columns):
            if j <= i:
                continue  # Only upper triangle to avoid duplicates
            weight = matrix_df.iloc[i, j]
            if weight >= min_weight:
                links.append((source, target, weight))

    return pd.DataFrame(links, columns=["source", "target", "weight"])



# top cited


def select_global_top_cited_documents(df, top_n=10, cols=None, filters=None, cite_col="Cited by"):
    """
    Select globally top-cited documents from a DataFrame.

    Parameters:
    - df: input DataFrame
    - top_n: number of top entries (default: 10)
    - cols: list of columns to return (default: key bibliographic fields)
    - filters: dict of column: condition (e.g. {"Year": lambda x: x >= 2015})
    - cite_col: citation count column (default: "Cited by")
    
    Returns:
    - DataFrame with top cited documents
    """
    if cols is None:
        cols = ["Authors", "Title", "Source title", "Year", "Document Type"]
    if filters:
        for k, cond in filters.items():
            df = df[df[k].apply(cond)]
    df_sorted = df.sort_values(by=cite_col, ascending=False)
    cutoff = df_sorted[cite_col].iloc[top_n - 1] if top_n < len(df_sorted) else -1
    df_top = df_sorted[df_sorted[cite_col] >= cutoff]
    out_cols = list(dict.fromkeys(cols + [cite_col]))
    return df_top[out_cols].reset_index(drop=True)

def select_local_top_cited_documents(df, top_n=10, cols=None, filters=None,
                                     title_col="Title", ref_col="References", cite_col="Cited by"):
    """
    Select locally top-cited documents based on how often their title appears in other documents' References.

    Parameters:
    - df: input DataFrame
    - top_n: number of top entries (default: 10)
    - cols: columns to return (default: ["Authors", "Title", "Source title", "Year", "Document Type"])
    - filters: dict of column: condition (e.g. {"Year": lambda x: x >= 2015})
    - title_col: name of the column containing document titles
    - ref_col: name of the column containing references
    - cite_col: name of the column with global citations (renamed to "Global citations")

    Returns:
    - DataFrame with top locally cited documents, including local and global citation counts
    """
    if cols is None:
        cols = ["Authors", "Title", "Source title", "Year", "Document Type"]

    if filters:
        for k, cond in filters.items():
            df = df[df[k].apply(cond)]

    titles = df[title_col].dropna().unique()
    title_counts = dict.fromkeys(titles, 0)

    # Count how many times each title appears in the References column
    for ref in df[ref_col].dropna():
        for t in titles:
            if t in ref:
                title_counts[t] += 1

    # Create a local citation column
    df = df.copy()
    df["Local citations"] = df[title_col].map(title_counts).fillna(0).astype(int)
    df["Global citations"] = df[cite_col]

    # Sort by local citations and handle ties
    df_sorted = df.sort_values(by="Local citations", ascending=False)
    cutoff = df_sorted["Local citations"].iloc[top_n - 1] if top_n < len(df_sorted) else -1
    df_top = df_sorted[df_sorted["Local citations"] >= cutoff]

    out_cols = list(dict.fromkeys(cols + ["Local citations", "Global citations"]))
    return df_top[out_cols].reset_index(drop=True)


        
        
# language translation

lang_dict_df = pd.read_excel(fd + "\\additional files\\language dictionary.xlsx")

def ldf(x, lang_dict_df=lang_dict_df, l="en"):
    lang_dict = lang_dict_df.set_index("term").to_dict()[l]
    return lang_dict[x] if x in lang_dict else x


# descriptive statistics

def compute_descriptives(df, column, col_type, stopwords=None):
    """
    Compute descriptive statistics for a given DataFrame column based on its type.

    Parameters:
        df (pd.DataFrame): Input DataFrame.
        column (str): Name of the column to analyze.
        col_type (str): One of {"numeric", "categorical", "text", "list"}.
        stopwords (set or list): Custom stopwords for text analysis (used for "text" type).

    Returns:
        dict: Summary of descriptive statistics.
    """
    if column not in df.columns:
        return {"Variable not found": np.nan}

    series = df[column]
    result = {}

    non_missing = series.dropna()
    result["Number of documents"] = len(non_missing)
    result["Missing values"] = {
        "count": series.isna().sum(),
        "percent": round(series.isna().mean() * 100, 2)
    }

    if col_type == "numeric":
        values = non_missing.astype(float)
        result.update({
            "Mean": values.mean(),
            "Median": values.median(),
            "Mode": values.mode().iloc[0] if not values.mode().empty else None,
            "Min": values.min(),
            "Max": values.max(),
            "Range": values.max() - values.min(),
            "Standard deviation": values.std(),
            "Percentiles": {
                "10%": values.quantile(0.10),
                "25%": values.quantile(0.25),
                "50%": values.quantile(0.50),
                "75%": values.quantile(0.75),
                "90%": values.quantile(0.90),
            },
            "Skewness": skew(values),
            "Kurtosis": kurtosis(values),
        })
        stat, p = normaltest(values)
        result["Normality test (D’Agostino–Pearson)"] = {
            "statistic": stat,
            "p-value": p,
            "normal": p >= 0.05
        }

    elif col_type == "categorical":
        freqs = non_missing.value_counts()
        total = len(non_missing)
        top_items = freqs.head(10)
        result.update({
            "Number of unique values": non_missing.nunique(),
            "Top 10 categories": [
                {"value": idx, "count": count, "percent": round(100 * count / total, 2)}
                for idx, count in top_items.items()
            ]
        })

    elif col_type == "text":
        stopwords = set(stopwords) if stopwords is not None else ENGLISH_STOP_WORDS
        word_lengths = non_missing.apply(lambda x: len(re.findall(r"\w+", x)))
        all_words = [
            word.lower()
            for text in non_missing
            for word in re.findall(r"\w+", text)
            if word.lower() not in stopwords
        ]
        word_counts = Counter(all_words).most_common(10)
        total_words = len(all_words)
        result.update({
            "Length (in words)": {
                "mean": word_lengths.mean(),
                "median": word_lengths.median(),
                "min": word_lengths.min(),
                "max": word_lengths.max(),
                "std": word_lengths.std()
            },
            "Total unique words": len(set(all_words)),
            "Top 10 words": [
                {"word": word, "count": count, "percent": round(100 * count / total_words, 2)}
                for word, count in word_counts
            ]
        })

    elif col_type == "list":
        parsed = non_missing.apply(lambda x: [i.strip().lower() for i in str(x).split(";") if i.strip()])
        lengths = parsed.apply(len)
        all_items = [item for sublist in parsed for item in sublist]
        total_items = len(all_items)
        item_counts = Counter(all_items).most_common(10)
        result.update({
            "List length": {
                "mean": lengths.mean(),
                "median": lengths.median(),
                "min": lengths.min(),
                "max": lengths.max(),
                "std": lengths.std()
            },
            "Total unique elements": len(set(all_items)),
            "Top 10 items": [
                {"item": item, "count": count, "percent": round(100 * count / total_items, 2)}
                for item, count in item_counts
            ]
        })

    else:
        result["Invalid column type"] = col_type

    return result


def flatten_descriptives(name, summary):
    """
    Flatten a nested dictionary into a list of (Variable, Indicator, Value) tuples.

    - Lists of dicts (e.g., Top 10 items) are collapsed into a single string.
    - Items are separated by ';\n' for readability in Excel cells.
    """
    rows = []
    for key, value in summary.items():
        if isinstance(value, dict):
            for subkey, subval in value.items():
                label = f"{key} - {subkey}"
                rows.append((name, label, subval))
        elif isinstance(value, list):
            if all(isinstance(item, dict) and "count" in item for item in value):
                collapsed = ";\n".join(
                    f"{item.get('value') or item.get('word') or item.get('item')} ({item['count']})"
                    for item in value
                )
                rows.append((name, key, collapsed))
            else:
                rows.append((name, key, "\n".join(str(x) for x in value)))
        else:
            rows.append((name, key, value))
    return rows


def save_descriptives_to_excel(dataframes_with_sheets, excel_path, freeze_top_row=False):
    """
    Save one or more descriptive summary DataFrames to a styled Excel file.

    Parameters:
        dataframes_with_sheets (list): List of (DataFrame, sheet_name) pairs.
        excel_path (str): Path to save the Excel file.
        freeze_top_row (bool): Whether to freeze the top row on each sheet.
    """
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for df, sheet_name in dataframes_with_sheets:
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
            sheet = writer.sheets[sheet_name]

            # Bold and center-align header row
            for col in range(1, df.shape[1] + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Merge and center-align cells in "Variable" column (if it exists)
            if "Variable" in df.columns:
                current_row = 2
                for var in df["Variable"].unique():
                    count = (df["Variable"] == var).sum()
                    if count > 1:
                        sheet.merge_cells(start_row=current_row, start_column=1,
                                          end_row=current_row + count - 1, end_column=1)
                    cell = sheet.cell(row=current_row, column=1)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    current_row += count

            # Freeze header row if requested
            if freeze_top_row:
                sheet.freeze_panes = "A2"

            # Autofit column widths
            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                adjusted_width = min(max_length + 2, 100)
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Wrap text in all cells and auto-adjust height
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

def compute_descriptive_statistics(df, columns_with_types, stopwords=None):
    """
    Compute descriptive bibliometric statistics for multiple columns.

    Parameters:
        df (pd.DataFrame): Input DataFrame.
        columns_with_types (list): List of (column_name, column_type) pairs.
        stopwords (set or list): Custom stopwords for text analysis.

    Returns:
        pd.DataFrame: Summary table with columns: Variable, Indicator, Value.
    """
    all_rows = []

    for column, col_type in columns_with_types:
        summary = compute_descriptives(df, column, col_type, stopwords)
        flat = flatten_descriptives(column, summary)
        all_rows.extend(flat)

    return pd.DataFrame(all_rows, columns=["Variable", "Indicator", "Value"])



# counting


def count_occurrences(df, column_name, count_type="single", ngram_range=(1, 1), 
                      item_column_name="Item", rename_dict=None, translated_column_name="Translated Item",
                      sep="; "):
    """
    Processes a DataFrame column and returns a DataFrame with counts and proportions.
    
    Parameters:
        df (pd.DataFrame): Input DataFrame.
        column_name (str): Name of the column to process.
        count_type (str): Type of processing - "single", "list", or "text".
        ngram_range (tuple): N-gram range for text processing.
        item_column_name (str): Custom name for the "Item" column (default is "Item").
        rename_dict (dict): Dictionary for renaming items; if provided, a new column is added.
        translated_column_name (str): Custom name for the translated column (default is "Translated Item").
    
    Returns:
        pd.DataFrame: Processed counts sorted in descending order.
    """
    total_rows = len(df)
    data = df[column_name].dropna().astype(str).str.strip()  # Remove NaNs and trim whitespace
    data = data[data != ""]  # Remove empty values

    if count_type == "single":
        counts = Counter(data)
        
    elif count_type == "list":
        split_data = data.str.split("; ")
        split_data = split_data.dropna().tolist()
        flattened = list(chain.from_iterable(split_data))
        counts = Counter(flattened)
        
        fractional_counts = Counter()
        for items in split_data:
            unique_items = set(items)
            weight = 1 / len(unique_items) if unique_items else 0
            for item in unique_items:
                fractional_counts[item] += weight
        
    elif count_type == "text":
        if data.empty:
            columns = [item_column_name, "Number of documents", "Proportion of documents", 
                       "Percentage of documents", "Number of occurrences"]
            if rename_dict:
                columns.insert(1, translated_column_name)  # Insert translated column
            return pd.DataFrame(columns=columns)
        
        vectorizer = CountVectorizer(ngram_range=ngram_range)
        term_matrix = vectorizer.fit_transform(data)
        terms = vectorizer.get_feature_names_out()
        
        doc_counts = (term_matrix > 0).sum(axis=0).A1  # Number of documents containing the term
        total_counts = term_matrix.sum(axis=0).A1  # Total occurrences in all documents
        
        result_df = pd.DataFrame({
            item_column_name: terms,
            "Number of documents": doc_counts,
            "Proportion of documents": doc_counts / total_rows,
            "Percentage of documents": (doc_counts / total_rows) * 100,
            "Number of occurrences": total_counts
        })
        
        # Insert translated column if rename_dict is provided
        if rename_dict:
            result_df.insert(1, translated_column_name, result_df[item_column_name].map(rename_dict).fillna(""))

        return result_df.sort_values(by="Number of documents", ascending=False)
    
    else:
        raise ValueError('Invalid count_type. Choose from "single", "list", or "text".')
    
    # Create result DataFrame for "single" and "list" cases
    result_data = {
        item_column_name: list(counts.keys()),
        "Number of documents": list(counts.values()),
        "Proportion of documents": [v / total_rows for v in counts.values()],
        "Percentage of documents": [(v / total_rows) * 100 for v in counts.values()],
    }
    
    # Add fractional counts for "list" case
    if count_type == "list":
        result_data["Fractional number of documents"] = [fractional_counts[item] for item in counts.keys()]
    
    result_df = pd.DataFrame(result_data)

    # Insert translated column if rename_dict is provided
    if rename_dict:
        result_df.insert(1, translated_column_name, result_df[item_column_name].map(rename_dict).fillna(""))

    return result_df.sort_values(by="Number of documents", ascending=False)



# Scientific production


def get_scientific_production(df, relative_counts=True, cumulative=True, predict_last_year=True, percent_change=True):
    """
    Computes the annual scientific production statistics from a dataset containing publication years and citation counts.
    
    Parameters:
    df (pd.DataFrame): DataFrame containing at least "Year" and "Cited by" columns.
    relative_counts (bool, optional): Whether to compute relative proportions and percentages of documents. Default is True.
    cumulative (bool, optional): Whether to compute cumulative document and citation counts. Default is True.
    predict_last_year (bool, optional): Whether to predict the current year's document and citation counts. Default is True.
    percent_change (bool, optional): Whether to compute year-over-year percentage change in documents and citations. Default is True.
    
    Returns:
    pd.DataFrame: A DataFrame containing yearly statistics on document counts, citations, and optionally, predictions for the current year.
    """
    # Ensure "Year" is treated as an integer
    df["Year"] = df["Year"].astype(int)
    
    # Create a complete range of years
    all_years = pd.Series(range(df["Year"].min(), df["Year"].max() + 1), name="Year")
    
    # Aggregate counts and citations
    production = df.groupby("Year").agg(
        Documents=("Year", "count"),
        Total_Citations=("Cited by", "sum")
    ).reset_index()
    
    # Merge with the full range of years and fill missing values with 0
    production = all_years.to_frame().merge(production, on="Year", how="left").fillna(0)
    
    # Convert back to integer where applicable
    production["Documents"] = production["Documents"].astype(int)
    production["Total_Citations"] = production["Total_Citations"].astype(int)
    
    # Compute relative counts if needed
    if relative_counts:
        total_docs = production["Documents"].sum()
        production["Proportion Documents"] = production["Documents"] / total_docs
        production["Percentage Documents"] = production["Proportion Documents"] * 100
    
    # Compute cumulative values if needed
    if cumulative:
        production["Cumulative Documents"] = production["Documents"].cumsum()
        production["Cumulative Citations"] = production["Total_Citations"].cumsum()
        
        if relative_counts:
            production["Cumulative Proportion Documents"] = production["Cumulative Documents"] / total_docs
            production["Cumulative Percentage Documents"] = production["Cumulative Proportion Documents"] * 100

    # Compute percentage change if needed
    if percent_change:
        production["Percentage Change Documents"] = production["Documents"].pct_change() * 100
        production["Percentage Change Citations"] = production["Total_Citations"].pct_change() * 100

    # Predict last year if applicable
    if predict_last_year:
        current_year = datetime.datetime.now().year
        if production["Year"].max() == current_year:
            previous_years = production[production["Year"] < current_year]
            avg_growth_docs = previous_years["Documents"].pct_change().mean()
            avg_growth_citations = previous_years["Total_Citations"].pct_change().mean()
            
            if np.isfinite(avg_growth_docs):
                production.loc[production["Year"] == current_year, "Predicted Documents"] = production["Documents"].iloc[-2] * (1 + avg_growth_docs)
            else:
                production.loc[production["Year"] == current_year, "Predicted Documents"] = production["Documents"].iloc[-2]
            
            if np.isfinite(avg_growth_citations):
                production.loc[production["Year"] == current_year, "Predicted Citations"] = production["Total_Citations"].iloc[-2] * (1 + avg_growth_citations)
            else:
                production.loc[production["Year"] == current_year, "Predicted Citations"] = production["Total_Citations"].iloc[-2]
            
            production["Predicted Documents"] = production["Predicted Documents"].fillna(production["Documents"]).astype(int)
            production["Predicted Citations"] = production["Predicted Citations"].fillna(production["Total_Citations"]).astype(int)
    
    production = production.rename(columns={
        "Documents": "Number of Documents",
        "Total_Citations": "Total Citations",
        "Predicted Documents": "Predicted Number of Documents",
        "Predicted Citations": "Predicted Total Citations"
    })

    return production


def summarize_publication_timeseries(production_df):
    """
    Compute and format summary statistics from a time series production_df.

    Parameters:
        production_df (pd.DataFrame): Must contain:
            'Year', 'Number of Documents', 'Total Citations',
            'Percentage Change Documents'

    Returns:
        pd.DataFrame: Formatted DataFrame with columns: Variable, Indicator, Value
    """
    df = production_df.copy()
    df = df.sort_values("Year").reset_index(drop=True)
    n_years = len(df)

    # Timespan
    timespan = f"{df['Year'].min()}–{df['Year'].max()}"

    # Most productive year
    max_docs_row = df.loc[df["Number of Documents"].idxmax()]
    most_productive = f"{int(max_docs_row['Year'])} ({int(max_docs_row['Number of Documents'])} documents)"

    # Highest / lowest growth (excluding inf/-inf)
    valid_growth = df[~df["Percentage Change Documents"].replace([np.inf, -np.inf], np.nan).isna()]
    max_growth_row = valid_growth.loc[valid_growth["Percentage Change Documents"].idxmax()]
    min_growth_row = valid_growth.loc[valid_growth["Percentage Change Documents"].idxmin()]
    highest_growth = f"{int(max_growth_row['Year'])} ({round(max_growth_row['Percentage Change Documents'], 2)}%)"
    lowest_growth = f"{int(min_growth_row['Year'])} ({round(min_growth_row['Percentage Change Documents'], 2)}%)"

    # Geometric mean helper
    def geometric_mean_growth(series):
        rates = 1 + series.dropna() / 100
        if len(rates) == 0:
            return None
        gmean = np.prod(rates) ** (1 / len(rates)) - 1
        return round(gmean * 100, 2)

    # Average growth strings
    avg_growth_all = f"{geometric_mean_growth(df['Percentage Change Documents'])}%"
    avg_growth_3 = f"{geometric_mean_growth(df['Percentage Change Documents'].tail(3))}%" if n_years >= 3 else None
    avg_growth_5 = f"{geometric_mean_growth(df['Percentage Change Documents'].tail(5))}%" if n_years >= 5 else None
    avg_growth_10 = f"{geometric_mean_growth(df['Percentage Change Documents'].tail(10))}%" if n_years >= 10 else None

    # Most influential years
    df["Citations per Document"] = df["Total Citations"] / df["Number of Documents"]

    def influential(df_slice):
        if len(df_slice) == 0:
            return None
        row = df_slice.loc[df_slice["Citations per Document"].idxmax()]
        return f"{int(row['Year'])} ({round(row['Citations per Document'], 2)} citations/doc)"

    influential_all = influential(df)
    influential_3 = influential(df.tail(3)) if n_years >= 3 else None
    influential_5 = influential(df.tail(5)) if n_years >= 5 else None
    influential_10 = influential(df.tail(10)) if n_years >= 10 else None

    # Collect rows
    rows = [
        ("Time series analysis", "Timespan", timespan),
        ("Time series analysis", "Most Productive Year", most_productive),
        ("Time series analysis", "Highest Growth", highest_growth),
        ("Time series analysis", "Lowest Growth", lowest_growth),
        ("Time series analysis", "Average Growth (All Years)", avg_growth_all),
    ]
    if avg_growth_3: rows.append(("Time series analysis", "Average Growth (Last 3 Years)", avg_growth_3))
    if avg_growth_5: rows.append(("Time series analysis", "Average Growth (Last 5 Years)", avg_growth_5))
    if avg_growth_10: rows.append(("Time series analysis", "Average Growth (Last 10 Years)", avg_growth_10))

    rows.append(("Time series analysis", "Most Influential Year", influential_all))
    if influential_3: rows.append(("Time series analysis", "Most Influential (Last 3 Years)", influential_3))
    if influential_5: rows.append(("Time series analysis", "Most Influential (Last 5 Years)", influential_5))
    if influential_10: rows.append(("Time series analysis", "Most Influential (Last 10 Years)", influential_10))

    return pd.DataFrame(rows, columns=["Variable", "Indicator", "Value"])




# Keywords processing


def merge_keywords_columns(
    df: pd.DataFrame,
    author_col: str = "Author Keywords",
    index_col: str = "Index Keywords"
) -> pd.Series:
    """
    Merges two keyword columns (default: "Author Keywords" and "Index Keywords") in a DataFrame.
    Removes duplicates and returns a new Series with combined keywords separated by "; ".

    Parameters:
        df (pd.DataFrame): The DataFrame containing the keyword columns.
        author_col (str): Name of the column with author keywords. Default is "Author Keywords".
        index_col (str): Name of the column with index keywords. Default is "Index Keywords".

    Returns:
        pd.Series: A new Series with merged and deduplicated keywords.
    """
    def merge_keywords(row):
        ak = row[author_col].split("; ") if pd.notnull(row[author_col]) else []
        ik = row[index_col].split("; ") if pd.notnull(row[index_col]) else []
        merged = sorted(set(ak + ik))
        return "; ".join(merged)

    return df.apply(merge_keywords, axis=1)


def merge_text_columns(
    df: pd.DataFrame,
    title_col: str = "Title",
    abstract_col: str = "Abstract",
    author_col: str = "Author Keywords",
    index_col: str = "Index Keywords",
    combined_col: str = "Combined Text"
) -> pd.DataFrame:
    """
    Builds a single text field by concatenating Title, Abstract, and merged keywords,
    removes punctuation and non-alphanumeric characters, and collapses repeated runs of the same character.

    Parameters:
        df (pd.DataFrame): Input DataFrame.
        title_col (str): Column name for titles. Default is "Title".
        abstract_col (str): Column name for abstracts. Default is "Abstract".
        author_col (str): Column name for author keywords. Default is "Author Keywords".
        index_col (str): Column name for index keywords. Default is "Index Keywords".
        combined_col (str): Name of the new combined-text column. Default is "Combined Text".

    Returns:
        pd.DataFrame: DataFrame with a new column `combined_col` containing the cleaned, merged text.
    """
    # 1) Merge keywords into a Series
    keywords_series = merge_keywords_columns(df, author_col=author_col, index_col=index_col)

    # 2) Helper to merge fields and clean text
    def merge_fields(row, keywords):
        title = row[title_col] if pd.notnull(row[title_col]) else ""
        abstract = row[abstract_col] if pd.notnull(row[abstract_col]) else ""
        kw = keywords[row.name]
        parts = [title.strip(), abstract.strip(), kw]
        raw = " ".join(part for part in parts if part)
        # remove any non-word, non-space characters
        cleaned = re.sub(r"[^\w\s]", "", raw)
        # collapse any character repeated more than twice to a single instance
        cleaned = re.sub(r"(\w)\1{2,}", r"\1", cleaned)
        return cleaned

    # 3) Apply and assign
    df[combined_col] = df.apply(lambda r: merge_fields(r, keywords_series), axis=1)
    return df



# Download necessary nltk data
nltk.download('wordnet')

def preprocess_keywords(df, column, exclude_list=None, synonyms=None, lemmatize=False, sep="; "):
    """
    Preprocess keywords in a given DataFrame column.
    
    Steps:
    1. Convert all keywords to lowercase.
    2. (Optional) Remove unwanted keywords.
    3. (Optional) Replace synonyms (DataFrame or dictionary).
    4. (Optional) Lemmatize to singular form.
    5. Uses a custom separator (default: "; ").
    
    Args:
        df (pd.DataFrame): Input DataFrame.
        column (str): Name of the column containing keywords.
        exclude_list (list or pd.DataFrame, optional): Either a list of keywords to exclude, 
                                                       or a DataFrame where the first column 
                                                       contains words to exclude. Default is None.
        synonyms (pd.DataFrame or dict, optional): If a DataFrame, the first column contains words 
                                                   to replace, and the second contains replacement words.
                                                   If a dictionary, keys are replacement words, 
                                                   and values are lists of words to replace.
        lemmatize (bool): Whether to apply lemmatization (default: False).
        sep (str): Separator used in the keyword column (default: "; ").
    
    Returns:
        pd.DataFrame: Modified DataFrame with an additional column containing processed keywords.
    
    Raises:
        ValueError: If the specified column does not exist in the DataFrame.
    """
    # Check if column exists in DataFrame
    if column not in df.columns:
        print(f"Column '{column}' not found in the DataFrame.")
        return df

    lemmatizer = WordNetLemmatizer()

    # Convert synonyms to a dictionary
    synonym_dict = {}
    if isinstance(synonyms, pd.DataFrame) and not synonyms.empty:
        old_keywords = synonyms.iloc[:, 0].str.lower()  # First column = old keywords (to be replaced)
        new_keywords = synonyms.iloc[:, 1].str.lower()  # Second column = new keywords (replacement)
        synonym_dict = dict(zip(old_keywords, new_keywords))
    elif isinstance(synonyms, dict):
        # Reverse mapping: multiple old words to one replacement word
        for replacement_word, words_to_replace in synonyms.items():
            for word in words_to_replace:
                synonym_dict[word.lower()] = replacement_word.lower()

    # Convert exclude_list to a set (supports both list and DataFrame input)
    if isinstance(exclude_list, pd.DataFrame) and not exclude_list.empty:
        exclude_set = set(exclude_list.iloc[:, 0].str.lower())  # First column contains words to exclude
    elif isinstance(exclude_list, list):
        exclude_set = set(word.lower() for word in exclude_list)
    else:
        exclude_set = set()  # No exclusion if exclude_list is None

    def process_row(keyword_str):
        if pd.isna(keyword_str):
            return ""
        
        # Convert keywords to lowercase and split using the specified separator
        keywords = [k.lower() for k in keyword_str.split(sep)]
        
        processed_keywords = []
        for keyword in keywords:
            # Replace with synonym if found
            keyword = synonym_dict.get(keyword, keyword)
            
            # (Optional) Lemmatize to singular
            if lemmatize:
                keyword = lemmatizer.lemmatize(keyword)
            
            # (Optional) Exclude unwanted keywords
            if keyword not in exclude_set:
                processed_keywords.append(keyword)
        
        return sep.join(sorted(set(processed_keywords)))  # Sort and remove duplicates

    # Define new column name: "Processed " + original column name
    new_column = "Processed " + column
    
    # Add a new column with processed keywords
    df[new_column] = df[column].apply(process_row)
    
    return df



# abstract processing

stopwords_file = fd + "\\additional files\\language dictionary.xlsx"

def process_text_column(df, column_name, stopwords_file=None, lang="english", remove_numbers=True, remove_two_letter_words=True):
    """
    Process a text column in a DataFrame by removing stopwords, lemmatizing the text, optionally removing numbers,
    and optionally removing words with just two letters.
    
    Parameters:
    df (pd.DataFrame): The DataFrame containing the text column.
    column_name (str): The name of the column to process.
    stopwords_file (str, optional): Path to the Excel file containing stopwords. If not provided, uses NLTK default stopwords.
    lang (str): Language of the stopwords (default is "english").
    remove_numbers (bool): Whether to remove numbers from the text (default is True).
    remove_two_letter_words (bool): Whether to remove words with exactly two letters (default is True).
    
    Returns:
    pd.DataFrame: DataFrame with an additional column for processed text.
    """
    # Download required nltk resources if not already available
    nltk.download("punkt")
    nltk.download("wordnet")
    nltk.download("stopwords")
    
    # Load stopwords from the Excel file if provided, otherwise use NLTK default
    if stopwords_file:
        stopwords_df = pd.read_excel(stopwords_file)
        stopwords_set = set(stopwords_df[lang].dropna().tolist())  # Ensure no NaN values
    else:
        stopwords_set = set(stopwords.words("english"))
    
    lemmatizer = WordNetLemmatizer()
    
    def clean_text(text):
        if pd.isna(text):  # Handle missing values
            return None
        
        tokens = word_tokenize(text)
        processed_tokens = []
        
        for token in tokens:
            lower_token = token.lower()
            if lower_token in stopwords_set:
                continue
            if remove_numbers and lower_token.isdigit():  # Remove numbers if the option is enabled
                continue
            if remove_two_letter_words and len(lower_token) == 2:  # Remove two-letter words if enabled
                continue
            if lower_token.isalnum():
                processed_tokens.append(lemmatizer.lemmatize(lower_token))
        
        return " ".join(processed_tokens)
    
    # Apply the function to the specified column
    df[f"Processed {column_name}"] = df[column_name].apply(clean_text)
    
    return df


# topic modelling


def determine_optimal_topics(texts, max_topics=10, language="english"):
    """Determine the optimal number of topics using perplexity scores.
    
    Args:
        texts (list of str): Preprocessed text documents.
        max_topics (int): Maximum number of topics to evaluate.
        language (str): Language stop words to use (default: English).
    
    Returns:
        int: Optimal number of topics.
    """
    vectorizer = CountVectorizer(stop_words=language)
    doc_term_matrix = vectorizer.fit_transform(texts)
    perplexities = []
    topic_range = range(2, max_topics + 1)
    for n_topics in topic_range:
        lda = LatentDirichletAllocation(n_components=n_topics, random_state=42)
        lda.fit(doc_term_matrix)
        perplexities.append(lda.perplexity(doc_term_matrix))
    optimal_topics = topic_range[np.argmin(perplexities)]
    return optimal_topics

def topic_modeling(df, text_column, model_type="LDA", n_topics=None, max_topics=10, 
                    max_features=5000, stop_words="english"):
    """Perform topic modeling using LDA, NMF, or LSA.
    
    Args:
        df (pd.DataFrame): DataFrame containing the text data.
        text_column (str): Column name containing text data.
        model_type (str): Topic modeling algorithm ('LDA', 'NMF', 'LSA').
        n_topics (int, optional): Number of topics (default: None, auto-detects optimal topics).
        max_topics (int): Maximum topics to consider if auto-detecting.
        max_features (int): Maximum number of features for vectorization.
        stop_words (str): Language stop words to use.
    
    Returns:
        tuple: DataFrame with topic assignments, DataFrame with topic keywords and weights.
    """
    vectorizer = TfidfVectorizer(max_features=max_features, stop_words=stop_words)
    if model_type in ["LDA", "LSA"]:
        vectorizer = CountVectorizer(max_features=max_features, stop_words=stop_words)
    
    doc_term_matrix = vectorizer.fit_transform(df[text_column])
    
    if n_topics is None:
        n_topics = determine_optimal_topics(df[text_column], max_topics, stop_words)
    
    if model_type == "LDA":
        model = LatentDirichletAllocation(n_components=n_topics, random_state=42)
    elif model_type == "NMF":
        model = NMF(n_components=n_topics, random_state=42)
    elif model_type == "LSA":
        model = TruncatedSVD(n_components=n_topics, random_state=42)
    else:
        raise ValueError("Invalid model type. Choose 'LDA', 'NMF', or 'LSA'.")
    
    model.fit(doc_term_matrix)
    
    feature_names = vectorizer.get_feature_names_out()
    topic_data = []
    for i, topic in enumerate(model.components_):
        for idx in topic.argsort()[:-11:-1]:
            topic_data.append({"Topic": f"Topic {i+1}", "Word": feature_names[idx], "Weight": topic[idx]})
    
    topic_df = pd.DataFrame(topic_data)
    topic_assignments = model.transform(doc_term_matrix).argmax(axis=1)
    df["Topic"] = topic_assignments + 1  # Making topics 1-based index
    
    return df, topic_df


# semantic interdisciplinarity analysis



# Optional import with fallback
try:
    from sentence_transformers import SentenceTransformer
    _model_available = True
    _model = SentenceTransformer("all-MiniLM-L6-v2")
except ImportError:
    _model_available = False
    _model = None

def semantic_interdisciplinarity(text: str, mode: str = "keywords") -> float:
    """
    Compute semantic interdisciplinarity based on average pairwise cosine distance 
    using sentence-transformers embeddings.

    Parameters:
        text (str): Keywords separated by "; ", or free-form text.
        mode (str): Either 'keywords' or 'text'.

    Returns:
        float: Mean pairwise cosine distance between embedded concepts.
               Returns np.nan if input is missing or insufficient.
    
    Raises:
        ImportError: If sentence-transformers is not installed.
    """
    if not _model_available:
        raise ImportError(
            "The 'sentence-transformers' package is required for semantic_interdisciplinarity().\n"
            "Install it with: pip install sentence-transformers"
        )

    if not isinstance(text, str) or not text.strip():
        return np.nan  # Handle None, NaN, or empty string

    if mode == "keywords":
        tokens = [t.strip() for t in text.split("; ") if t.strip()]
    elif mode == "text":
        tokens = text.split()
    else:
        raise ValueError("mode must be 'keywords' or 'text'")

    if len(tokens) < 2:
        return np.nan  # Not enough concepts to compare

    embeddings = _model.encode(tokens, convert_to_numpy=True)
    normed = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)

    distances = [1 - np.dot(normed[i], normed[j])
                 for i, j in combinations(range(len(normed)), 2)]

    return float(np.mean(distances)) if distances else np.nan


# sentiment analysis

def analyze_sentiment(df, text_column, sentiment_threshold=0.05, top_words=10):
    """
    Performs sentiment analysis on a given DataFrame.

    Parameters:
    - df (pd.DataFrame): Input DataFrame with text data.
    - text_column (str): Name of the column containing the text to analyze.
    - sentiment_threshold (float): Threshold for sentiment classification (default 0.05).
    - top_words (int): Number of most common words to return for each sentiment category.

    Returns:
    - df (pd.DataFrame): Updated DataFrame with sentiment scores and categories.
    - stats_df (pd.DataFrame): DataFrame containing sentiment statistics and common words.
    """

    # Download lexicon and initialize sentiment analyzer
    nltk.download("vader_lexicon")
    sia = SentimentIntensityAnalyzer()
    
    # Compute sentiment scores
    df["Sentiment Score"] = df[text_column].apply(lambda x: sia.polarity_scores(str(x))["compound"])
    
    # Assign sentiment categories
    df["Sentiment Category"] = df["Sentiment Score"].apply(
        lambda x: "Positive" if x > sentiment_threshold else "Negative" if x < -sentiment_threshold else "Neutral"
    )
    
    # Compute statistics
    stats = {
        "Mean Sentiment Score": df["Sentiment Score"].mean(),
        "Median Sentiment Score": df["Sentiment Score"].median(),
        "Standard Deviation": df["Sentiment Score"].std()
    }
    
    sentiment_distribution = df["Sentiment Category"].value_counts(normalize=True) * 100
    stats.update(sentiment_distribution.to_dict())

    # Word Frequency Analysis for each sentiment category
    word_stats = {}
    for category in ["Positive", "Neutral", "Negative"]:
        words = " ".join(df[df["Sentiment Category"] == category][text_column].astype(str)).lower().split()
        common_words = Counter(words).most_common(top_words)
        word_stats[f"Top Words ({category} Documents)"] = "; ".join([word for word, _ in common_words])

    # Convert stats and word frequencies to DataFrame
    stats_dict = {**stats, **word_stats}
    stats_df = pd.DataFrame(list(stats_dict.items()), columns=["Metric", "Value"])
    
    return df, stats_df

# Performance indicators




# --- Performance Indicators: List-based functions ---

def h_index(citations, alpha=1):
    """
    Compute the generalized h-index with scaling factor alpha.
    """
    return sum(c >= alpha * (i + 1) for i, c in enumerate(sorted(citations, reverse=True)))

def g_index(citations):
    """
    Compute the g-index based on citation distribution.
    """
    citations = sorted(citations, reverse=True)
    cumulative = np.cumsum(citations)
    thresholds = np.arange(1, len(citations) + 1) ** 2
    differences = cumulative - thresholds
    failing = [i for i, x in enumerate(differences) if x < 0]
    return failing[0] if failing else len(citations)

def hg_index(citations):
    """
    Compute the HG-index as the geometric mean of h and g indices.
    """
    return np.sqrt(h_index(citations) * g_index(citations))

def c_index(citations, thresholds=[5, 10, 20, 50, 100]):
    """
    Compute the C-index for a set of citation thresholds.
    Returns a dictionary of counts.
    """
    return {c: sum(np.array(citations) >= c) for c in thresholds}

def tapered_h_index(citations):
    """
    Compute the tapered h-index (a more nuanced variant of h-index).
    """
    citations = list(map(int, citations))
    h = 0
    for i, cit in enumerate(citations):
        k = min(i + 1, cit)
        h += k / (2 * i + 1)
        h += sum(1 / (2 * i + 1) for i in range(i + 1, cit))
    return h

def chi_index(citations):
    """
    Compute the chi-index, a robust citation indicator.
    """
    citations = sorted(citations, reverse=True)
    if len(citations) == 0:
        return np.nan
    return np.sqrt(max((i + 1) * c for i, c in enumerate(citations)))

# tele preveri, ker so narejeni z AI
def a_index(citations):
    """
    Compute the a-index: the average number of citations in the h-core.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        float: a-index value.
    """
    h = h_index(citations)
    return np.mean(sorted(citations, reverse=True)[:h]) if h > 0 else 0


def r_index(citations):
    """
    Compute the r-index: square root of total citations in the h-core.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        float: r-index value.
    """
    h = h_index(citations)
    return np.sqrt(sum(sorted(citations, reverse=True)[:h])) if h > 0 else 0


def h2_index(citations):
    """
    Compute the h(2)-index: largest number h2 such that h2 papers have at least (h2)^2 citations.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        int: h(2)-index value.
    """
    citations = sorted(citations, reverse=True)
    return sum(c >= (i + 1) ** 2 for i, c in enumerate(citations))


def w_index(citations):
    """
    Compute the w-index: number of papers with at least 10 × rank citations.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        int: w-index value.
    """
    citations = sorted(citations, reverse=True)
    return sum(c >= 10 * (i + 1) for i, c in enumerate(citations))


def t_index(citations):
    """
    Compute the t-index: square root of the sum of square roots of citation counts.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        float: t-index value.
    """
    return np.sqrt(sum(np.sqrt(c) for c in citations if c >= 0))


def pi_index(citations):
    """
    Compute the pi-index: total citations for the top √N publications.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        int: pi-index value.
    """
    n = len(citations)
    top_k = int(np.sqrt(n))
    return sum(sorted(citations, reverse=True)[:top_k])


def gini_index(citations):
    """
    Compute the Gini coefficient: measures inequality in the citation distribution.

    Parameters:
        citations (list or array): Citation counts per publication.

    Returns:
        float: Gini coefficient (0 = equality, 1 = inequality).
    """
    citations = np.array(sorted(citations))
    n = len(citations)
    if n == 0:
        return 0.0
    cum_cit = np.cumsum(citations)
    return (n + 1 - 2 * np.sum(cum_cit) / cum_cit[-1]) / n if cum_cit[-1] > 0 else 0.0



# --- Performance Indicators: DataFrame-based functions ---

def total_citations(df):
    """
    Return total citations from the 'Cited by' column.
    """
    return df["Cited by"].sum() if "Cited by" in df.columns else None

def count_cited_documents(df):
    """
    Return the number of documents that have been cited (non-zero entries in 'Cited by' column).
    """
    if "Cited by" not in df.columns:
        return None
    return df["Cited by"].gt(0).sum()

def average_year(df):
    """
    Return average publication year.
    """
    return df["Year"].mean() if "Year" in df.columns else None

def percentile_year(df, quantile=0.5):
    """
    Return year at given percentile (e.g., median with quantile=0.5).
    """
    return df["Year"].quantile(q=quantile) if "Year" in df.columns else None

def first_publication_year(df):
    """
    Return the earliest publication year.
    """
    return df["Year"].min() if "Year" in df.columns else None

def last_publication_year(df):
    """
    Return the latest publication year.
    """
    return df["Year"].max() if "Year" in df.columns else None

def h_index_df(df, alpha=1):
    """
    Compute h-index from DataFrame.
    """
    return h_index(df["Cited by"].tolist(), alpha) if "Cited by" in df.columns else None

def g_index_df(df):
    """
    Compute g-index from DataFrame.
    """
    return g_index(df["Cited by"].tolist()) if "Cited by" in df.columns else None

def hg_index_df(df):
    """
    Compute HG-index from DataFrame.
    """
    return hg_index(df["Cited by"].tolist()) if "Cited by" in df.columns else None

def c_index_df(df, thresholds=[5, 10, 20, 50, 100]):
    """
    Compute C-index from DataFrame.
    """
    if "Cited by" in df.columns:
        return c_index(df["Cited by"].tolist(), thresholds)
    return {c: None for c in thresholds}

def tapered_h_index_df(df):
    """
    Compute tapered h-index from DataFrame.
    """
    return tapered_h_index(df["Cited by"].tolist()) if "Cited by" in df.columns else None

def chi_index_df(df):
    """
    Compute chi-index from DataFrame.
    """
    return chi_index(df["Cited by"].tolist()) if "Cited by" in df.columns else None


# --- Aggregator Functions ---

def get_performance_indicators(df, name=None, mode="core", name_col="Name", diversity_measure=entropy):
    """
    Compute a set of bibliometric performance indicators from a DataFrame.

    Parameters:
        df (pd.DataFrame): Input data containing at least 'Cited by' and optionally 'Year' and 'Cited <field>'.
        name (str, optional): Optional identifier (e.g., author name) to include in results.
        mode (str): One of "core", "extended", or "full" to determine indicator depth.
        name_col (str): Name of the column representing the identifier.
        diversity_measure (function): Function used to compute interdisciplinarity.

    Returns:
        list of tuples: List of (indicator_name, value) pairs.
    """
    indicators = [(name_col, name)] if name else []

    # Core indicators
    indicators += [
        ("Number of documents", len(df)),
        ("Total citations", total_citations(df)),
        ("H-index", h_index_df(df)),
        ("Average year", average_year(df))
    ]

    # Extended indicators
    if mode in ["extended", "full"]:
        g_idx = g_index_df(df)
        c_idx = c_index_df(df)

        indicators += [
            ("G-index", g_idx),
            *[(f"C{c}", c_idx[c]) for c in sorted(c_idx)],
            ("First year", first_publication_year(df)),
            ("Q1 year", percentile_year(df, 0.25)),
            ("Median year", percentile_year(df, 0.5)),
            ("Q3 year", percentile_year(df, 0.75)),
            ("Last year", last_publication_year(df))
        ]

    # Full mode includes advanced indicators
    if mode == "full" and "Cited by" in df.columns:
        citations = df["Cited by"].tolist()

        indicators += [
            ("Number of cited documents", count_cited_documents(df)),
            ("A-index", a_index(citations)),
            ("R-index", r_index(citations)),
            ("H(2)-index", h2_index(citations)),
            ("W-index", w_index(citations)),
            ("T-index", t_index(citations)),
            ("Pi-index", pi_index(citations)),
            ("Gini index", gini_index(citations)),
            ("HG-index", hg_index(citations)),
            ("Chi-index", chi_index(citations)),
            ("Tapered H-index", tapered_h_index(citations))
        ]

        # Interdisciplinarity
        fields = [c for c in df.columns if "Cited" in c and c != "Cited by"]
        if fields:
            cited_fields = df[fields].sum()
            diversity = diversity_measure(cited_fields)
            if diversity_measure == entropy:
                diversity /= np.log(len(fields))
            indicators += [("Interdisciplinarity", diversity)] + list(zip(cited_fields.index, cited_fields))

    ci = None    
    if "Author(s) ID" in df.columns:
        ci = collaboration_index(df, "Author(s) ID")
    elif "Author full names" in df.columns:
        ci = collaboration_index(df, "Author(s) ID")
    elif "Authors" in df.columns:
        try:
            ci = collaboration_index(df, "Authors")
        except:
            pass
    if ci is not None:
        indicators += [("Collaboration index", ci)]

    return indicators


def get_specific_indicators(df, name=None, mode="core", name_col="Name", **kwargs):
    indicators = [(name_col, name)] if name else []
    try:
        text_column = next((col for col in ["Processed Abstract", "Abstract"] if col in df.columns), None)
        df, stats_df = analyze_sentiment(df, text_column, sentiment_threshold=0.05, top_words=10)
        
        indicators += [("Mean sentiment score", np.mean(df["Sentiment Score"])),
        ("Stdev of sentiment score", np.std(df["Sentiment Score"])),
        ("Highest sentiment score", np.max(df["Sentiment Score"])),
        ("Lowest sentiment score", np.min(df["Sentiment Score"])),
         ]
    except:
        pass
    
    
    
    try:
        top5 = df.nlargest(5, "Sentiment Score")["Title"]
        bottom5 = df.nsmallest(5, "Sentiment Score")["Title"]

        # Join into newline-separated strings
        top5_titles = "\n".join(top5.tolist())
        bottom5_titles = "\n".join(bottom5.tolist())
        
        indicators += [("5 titles of documents with highest sentiment score\n(based on abstracts)", top5_titles),
                       ("5 titles of documents with lowest sentiment score\n(based on abstracts)", bottom5_titles)]
    except:
        pass
    

    return indicators

# selection of dataframe for further performance analysis

def match_items_and_compute_binary_indicators(
    df,
    col,
    items_of_interest,
    value_type="string",
    separator="; ",
    indicators=True,
    missing_as_zero=True
):
    """
    Match items of interest in a specified dataframe column and optionally compute binary indicators.

    Notes:
        For 'text' matching, values are converted to lowercase to enable substring matching.
        Full-column lowercase transformation is avoided to prevent issues (e.g. with journal names).

    Parameters:
        df (pd.DataFrame): Input dataframe.
        col (str): Column in which to search for matches.
        items_of_interest (list): Items to search for.
        value_type (str): Type of values in column: 'string', 'list', or 'text'.
        separator (str): Separator for splitting list-type entries (used if value_type is 'list').
        indicators (bool): Whether to compute binary indicator columns.
        missing_as_zero (bool): If True, missing indicator values are replaced with 0.

    Returns:
        match_indices (dict): Dictionary mapping each item to a list of matched row indices.
        indicators_dict (dict): Dictionary of indicator DataFrames (empty if indicators=False).
    """

    match_indices = {item: [] for item in items_of_interest}

    for idx, val in df[col].items():
        if pd.isna(val):
            continue
        val_str = str(val).lower()
        if value_type == "string":
            if val in items_of_interest:
                match_indices[val].append(idx)
        elif value_type == "list":
            parts = [v.strip() for v in val.split(separator)]
            for item in items_of_interest:
                if item in parts:
                    match_indices[item].append(idx)
        elif value_type == "text":
            for item in items_of_interest:
                if item.lower() in val_str:
                    match_indices[item].append(idx)

    indicators_dict = {}
    if not indicators:
        return match_indices, indicators_dict

    indicator_01 = pd.DataFrame(index=df.index, columns=items_of_interest, dtype="float")
    # Identify rows with missing values in col
    missing_rows = df[col].isna()

    for item in items_of_interest:
        # Initialize mask: 1.0 for match, 0.0 for no match
        mask = df.index.isin(match_indices[item]).astype(float)
        # Set to NaN for rows with missing values in the relevant column
        mask[missing_rows] = np.nan
        indicator_01[item] = mask
        if missing_as_zero:
            indicator_01[item] = indicator_01[item].fillna(0)
    indicators_dict["binary"] = indicator_01

    return match_indices, indicators_dict




def select_documents(df, col, items_of_interest=None, exclude_items=None,
                     top_items_df=None, top_items_col=None, top_items_criterion="Number of documents", top_n=20,
                     regex_include=None, regex_exclude=None, indicators=False, missing_as_zero=False,
                     separator="; ", value_type="string", text_norm="tfidf"):
    if value_type not in {"string", "list", "text"}:
        raise ValueError('value_type must be one of "string", "list", or "text"')
    if text_norm not in {"tfidf", "df-icf", "mtf-idf", None}:
        raise ValueError('text_norm must be one of "tfidf", "df-icf", "mtf-idf", or None')

    items_of_interest = items_of_interest or []
    exclude_items = exclude_items or []

    if not items_of_interest:
        if top_items_df is None or top_items_col is None:
            raise ValueError("top_items_df and top_items_col are required when items_of_interest is not provided")

        filtered_df = top_items_df.dropna(subset=[top_items_col])

        if regex_include:
            filtered_df = filtered_df[filtered_df[top_items_col].astype(str).str.contains(regex_include, regex=True, na=False)]
        if regex_exclude:
            filtered_df = filtered_df[~filtered_df[top_items_col].astype(str).str.contains(regex_exclude, regex=True, na=False)]

        if regex_include or regex_exclude:
            items_of_interest = filtered_df[top_items_col].astype(str).tolist()
        else:
            df_sorted = filtered_df.sort_values(by=top_items_criterion, ascending=False)
            if len(df_sorted) <= top_n:
                items_of_interest = df_sorted[top_items_col].astype(str).tolist()
            else:
                cutoff = df_sorted.iloc[top_n - 1][top_items_criterion]
                selected = df_sorted[df_sorted[top_items_criterion] >= cutoff]
                items_of_interest = selected[top_items_col].astype(str).tolist()

    items_of_interest = list(set(items_of_interest) - set(exclude_items))
        
    match_indices, indicators_dict = match_items_and_compute_binary_indicators(df,
                                                                               col,
                                                                               items_of_interest,
                                                                               value_type=value_type,
                                                                               separator=separator,
                                                                               indicators=indicators,
                                                                               missing_as_zero=missing_as_zero)

    #df[col] = df[col].astype(str).str.lower() # tole naredi probleme z revijami, je pa treba pogledati, da ne bodo težave s ključnim besedami

    for idx, val in df[col].items():
        if pd.isna(val):
            continue
        val_str = str(val).lower()
        if value_type == "string":
            if val in items_of_interest:
                match_indices[val].append(idx)
        elif value_type == "list":
            parts = [v.strip() for v in val.split(separator)]
            for item in items_of_interest:
                if item in parts:
                    match_indices[item].append(idx)
        elif value_type == "text":
            for item in items_of_interest:
                if item.lower() in val_str:
                    match_indices[item].append(idx)

    indicators_dict = {}
    if not indicators:
        return match_indices, indicators_dict

    indicator_01 = pd.DataFrame(index=df.index, columns=items_of_interest, dtype="float")
    for item in items_of_interest:
        indicator_01[item] = df.index.isin(match_indices[item]).astype(float)
        if missing_as_zero:
            indicator_01[item] = indicator_01[item].fillna(0)
    indicators_dict["binary"] = indicator_01

    if value_type == "list":
        indicator_frac = pd.DataFrame(0, index=df.index, columns=items_of_interest, dtype="float")
        for idx, val in df[col].items():
            if pd.isna(val):
                if missing_as_zero:
                    indicator_frac.loc[idx] = 0
                continue
            parts = [v.strip() for v in str(val).split(separator)]
            total = len(parts)
            for p in parts:
                if p in items_of_interest:
                    indicator_frac.at[idx, p] += 1 / total
        indicators_dict["fractional"] = indicator_frac

    if value_type == "text":
        count_df = pd.DataFrame(0, index=df.index, columns=items_of_interest, dtype="float")
        for idx, val in df[col].items():
            if pd.isna(val):
                if missing_as_zero:
                    count_df.loc[idx] = 0
                continue
            val_str = str(val).lower()
            for item in items_of_interest:
                count_df.at[idx, item] = val_str.count(str(item).lower())
        indicators_dict["count"] = count_df

        if text_norm == "tfidf":
            tfidf = TfidfTransformer()
            tfidf_data = tfidf.fit_transform(count_df.fillna(0))
            tfidf_df = pd.DataFrame(tfidf_data.toarray(), index=df.index, columns=items_of_interest)
            indicators_dict["tfidf"] = tfidf_df
        elif text_norm == "df-icf":
            df_vec = (count_df > 0).sum(axis=0)
            icf = np.log((1 + len(count_df)) / (1 + df_vec)).values
            df_icf_df = count_df.multiply(icf, axis=1)
            indicators_dict["df-icf"] = df_icf_df
        elif text_norm == "mtf-idf":
            mtf = count_df.div(count_df.max(axis=1).replace(0, np.nan), axis=0)
            df_vec = (count_df > 0).sum(axis=0)
            idf = np.log((1 + len(count_df)) / (1 + df_vec)).values
            mtf_idf_df = mtf.multiply(idf, axis=1)
            indicators_dict["mtf-idf"] = mtf_idf_df

    return match_indices, indicators_dict

# Redefine get_entity_stats
def get_entity_stats(df, entity_col, entity_label,
                     items_of_interest=None, exclude_items=None, top_n=20,
                     counts_df=None, count_method=None,
                     regex_include=None, regex_exclude=None,
                     value_type="string", indicators=False,
                     missing_as_zero=False, mode="full"):

    if items_of_interest is None:
        if counts_df is None:
            if count_method is None:
                raise ValueError("Either counts_df or count_method must be provided.")
            counts_df = count_method()

        counts_df = counts_df.sort_values("Number of documents", ascending=False)
        first_col = counts_df.columns[0]

        filtered_df = counts_df.copy()
        if regex_include:
            filtered_df = filtered_df[filtered_df[first_col].astype(str).str.contains(regex_include, regex=True, na=False)]
        if regex_exclude:
            filtered_df = filtered_df[~filtered_df[first_col].astype(str).str.contains(regex_exclude, regex=True, na=False)]
        if (regex_include is None) and  (regex_exclude is None):
            items_of_interest = filtered_df[first_col].astype(str).tolist()[:top_n]
        else:
            items_of_interest = filtered_df[first_col].astype(str).tolist()

    selected_entities, indicator_df = select_documents(
        df, entity_col,
        value_type=value_type,
        items_of_interest=items_of_interest,
        exclude_items=exclude_items,
        top_items_df=None,
        top_items_col=None,
        top_n=top_n,
        regex_include=None,
        regex_exclude=None,
        indicators=indicators,
        missing_as_zero=missing_as_zero
    )

    stats_list = []
    for name in selected_entities:
        entity_df = df.iloc[selected_entities[name]]
        metrics = get_performance_indicators(entity_df, name=name, mode=mode, name_col=entity_label)
        stats_list.append(dict(metrics))

    stats_df = pd.DataFrame(stats_list)

    if items_of_interest is None and counts_df is not None:
        stats_df = merge_on_key(counts_df, stats_df, entity_label)
        stats_df = stats_df.sort_values("Number of documents", ascending=False)

    return stats_df, indicator_df if indicators else None



def get_all_performances(df, name_col, items, search_mode="exact", mode="core"):
    """
    Get performance indicators for multiple items in a DataFrame.

    Parameters:
        df (pd.DataFrame): Full dataset.
        name_col (str): Column used to identify individuals/items.
        items (list): List of names/items to retrieve.
        search_mode (str): 'exact' or 'substring'.
        mode (str): One of 'core', 'extended', 'full'.

    Returns:
        pd.DataFrame: Performance summary.
    """
    results = []
    for item in items:
        if search_mode == "exact":
            subset = df[df[name_col] == item]
        elif search_mode == "substring":
            subset = df[df[name_col].astype(str).str.contains(item)]
        else:
            raise ValueError("search_mode must be 'exact' or 'substring'")
        perf = get_performance_indicators(subset, item, mode=mode, name_col=name_col)
        results.append(perf)
    return pd.DataFrame([dict(p) for p in results])



# Excel saving

def to_excel_fancy(data, f_name="styled_output.xlsx", sheet_names=None, top_n=3, bottom_n=3, top_color="99FF99", bottom_color="FF9999", autofit=True, conditional_formatting=True):
    """
    Save one or multiple DataFrames to an Excel file with optional formatting:
    - Accepts a single DataFrame or a list of DataFrames.
    - Saves each DataFrame to a separate sheet if a list is provided.
    - Allows user-defined sheet names; defaults to Sheet1, Sheet2, etc., if None or empty list.
    - Highlights top_n highest values (not unique) with a user-defined color (default: green).
    - Highlights bottom_n lowest values (not unique) with a user-defined color (default: red).
    - Uses ranking to determine which values should be highlighted.
    - Autofits column width (optional, default=True).
    - Applies conditional formatting (optional, default=True).
    """
    
    # Ensure data is a list of DataFrames
    if isinstance(data, pd.DataFrame):
        data = [data]
    
    # Default sheet names if none provided
    if sheet_names is None or not sheet_names:
        sheet_names = [f"Sheet{i+1}" for i in range(len(data))]
    
    # Ensure correct sheet name length
    if len(sheet_names) != len(data):
        raise ValueError("Number of sheet names must match number of DataFrames.")
    
    # Define colors
    top_fill = PatternFill(start_color=top_color, end_color=top_color, fill_type="solid")  
    bottom_fill = PatternFill(start_color=bottom_color, end_color=bottom_color, fill_type="solid")  
    
    with pd.ExcelWriter(f_name, engine="openpyxl") as writer:
        for df, sheet_name in zip(data, sheet_names):
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet = writer.sheets[sheet_name]
            
            # Autofit column width if enabled
            if autofit:
                for col in sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter  # Get column letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    sheet.column_dimensions[col_letter].width = max_length + 2
            
            # Apply conditional formatting if enabled
            if conditional_formatting:
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if pd.api.types.is_numeric_dtype(df[col_name]):
                        values = [cell for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2) for cell in row]
                        
                        if values:
                            numeric_values = [(i, float(cell.value)) for i, cell in enumerate(values) if isinstance(cell.value, (int, float))]
                            val_list = [val for _, val in numeric_values]
                            
                            # Get rankings for bottom and top values
                            ranks_min = rankdata(val_list, method='min')  # Ranking with ties assigned the lowest rank
                            ranks_max = rankdata(val_list, method='max')  # Ranking with ties assigned the highest rank
                            
                            min_rank_threshold = bottom_n  # Get bottom_n ranks
                            max_rank_threshold = len(ranks_max) - top_n + 1  # Get top_n ranks
                            
                            # Apply formatting based on ranking
                            for (i, val), rank_min, rank_max in zip(numeric_values, ranks_min, ranks_max):
                                if rank_min <= min_rank_threshold:
                                    values[i].fill = bottom_fill
                                if rank_max >= max_rank_threshold:
                                    values[i].fill = top_fill
    
    print(f"Saved to {f_name}")
    
    
# References management


def parse_reference(ref, excluded_sources=None):
    """
    Parse a bibliographic reference string into a DataFrame with structured fields.
    
    Parameters:
        ref (str): The reference string to parse.
        excluded_sources (set, optional): A set of known non-source terms. Defaults to empty set.
    
    Returns:
        pd.DataFrame: A single-row DataFrame with columns:
                      ['Authors', 'Title', 'Source', 'Volume', 'Pages', 'Year']
    """
    if excluded_sources is None:
        excluded_sources = set()

    # Extract year
    year_match = re.search(r"\((\d{4})\)", ref)
    year = year_match.group(1) if year_match else None
    ref_wo_year = re.sub(r"\(\d{4}\)", "", ref).strip()

    # Extract authors from start
    author_match = re.match(r"^((?:[^,]+?\.,\s?)+)", ref_wo_year)
    authors = author_match.group(1).strip().rstrip(",") if author_match else None

    # Get remaining parts
    remaining = ref_wo_year[len(authors):].lstrip(", ").strip() if authors else ref_wo_year
    parts = [p.strip() for p in remaining.split(",") if p.strip()]

    pages = volume = source = title = None

    # Right-to-left: extract pages and volume
    while parts:
        part = parts[-1]
        if re.search(r"(pp\.\s*)?\d{1,5}-\d{1,5}", part):
            pages = re.search(r"\d{1,5}-\d{1,5}", part).group(0)
            parts.pop()
        elif re.match(r"^\d{1,4}$", part):
            volume = parts.pop()
        else:
            break

    # Source: last valid non-excluded part
    for i in reversed(range(len(parts))):
        candidate = parts[i]
        if candidate not in excluded_sources:
            source = candidate
            parts = parts[:i]
            break

    # Title: whatever remains
    if parts:
        title = ", ".join(parts)

    return pd.DataFrame([{
        "Authors": authors,
        "Title": title,
        "Source": source,
        "Volume": volume,
        "Pages": pages,
        "Year": year
    }])

def parse_references(ref_blob, excluded_sources=None):
    """
    Parse a semicolon-separated string of bibliographic references.

    Parameters:
        ref_blob (str): A long reference string with multiple references separated by semicolons.
        excluded_sources (set, optional): Terms to exclude from source detection.

    Returns:
        pd.DataFrame: Combined DataFrame with parsed information for each reference.
    """
    refs = [r.strip() for r in ref_blob.split(";") if r.strip()]
    parsed_dfs = [parse_reference(ref, excluded_sources) for ref in refs]
    return pd.concat(parsed_dfs, ignore_index=True)


def parse_references_dataframe(df, excluded_sources=None):
    """
    Parse a DataFrame with a 'References' column into structured bibliographic records.

    Parameters:
        df (pd.DataFrame): Input DataFrame with a 'References' column and optional 'Doc ID'.
        excluded_sources (set, optional): Set of known non-source terms to exclude. Defaults to empty set.

    Returns:
        pd.DataFrame: Parsed references with structured columns and 'Document source' as the first column.

    Raises:
        ValueError: If the 'References' column is missing.
    """
    if "References" not in df.columns:
        raise ValueError("Input DataFrame must contain a 'References' column.")

    if excluded_sources is None:
        excluded_sources = set()

    parsed_records = []

    for idx, row in df.iterrows():
        doc_id = row["Doc ID"] if "Doc ID" in df.columns else f"Row_{idx}"
        references = row["References"]

        if pd.isna(references):
            continue

        for ref in references.split(";"):
            ref = ref.strip()
            if ref:
                parsed = parse_references(ref, excluded_sources)
                parsed.insert(0, "Document source", doc_id)  # Make it the first column
                parsed_records.append(parsed)

    if parsed_records:
        return pd.concat(parsed_records, ignore_index=True)
    else:
        return pd.DataFrame(columns=["Document source", "Authors", "Title", "Source", "Volume", "Pages", "Year"])
    
    
def summarize_parsed_references(df_references):
    """
    Summarizes a structured references DataFrame into a compact descriptive format.

    Parameters:
        df_references (pd.DataFrame): Input parsed references with columns:
            - 'Document source'
            - 'Authors'
            - 'Source'
            - 'Year'

    Returns:
        pd.DataFrame: Summary with columns ['Variable', 'Item', 'Value']
    """
    records = []
    var_name = "references stats"

    # Clean and cast Year
    year_series = pd.to_numeric(df_references["Year"], errors="coerce").dropna().astype(int)

    # Distinct sources
    records.append((var_name, "Number of distinct sources", df_references["Source"].dropna().nunique()))

    # Top 10 sources
    top_sources = df_references["Source"].dropna().value_counts().head(10)
    sources_str = "\n".join([f"{s} ({c})" for s, c in top_sources.items()])
    records.append((var_name, "Top 10 cited sources", sources_str))

    # Year stats
    if not year_series.empty:
        records.append((var_name, "Average year of references", round(year_series.mean(), 2)))
        records.append((var_name, "Median year of references", year_series.median()))
        records.append((var_name, "Q1 (25%)", year_series.quantile(0.25)))
        records.append((var_name, "Q3 (75%)", year_series.quantile(0.75)))
        records.append((var_name, "Time span of references", f"{year_series.min()}–{year_series.max()}"))
    else:
        records.extend([
            (var_name, "Average year of references", None),
            (var_name, "Median year of references", None),
            (var_name, "Q1 (25%)", None),
            (var_name, "Q3 (75%)", None),
            (var_name, "Time span of references", None),
        ])

    # Authors - clean before splitting
    cleaned_authors = (
        df_references["Authors"]
        .dropna()
        .str.replace(r"\bet al\.?\b", "", case=False, regex=True)
        .str.replace(r"\s+", " ", regex=True)
    )

    author_list = (
        cleaned_authors
        .str.split(",")
        .explode()
        .str.strip()
        .replace(r"^\.?$", pd.NA, regex=True)
        .dropna()
    )

    records.append((var_name, "Number of distinct cited authors", author_list.nunique()))
    top_authors = author_list.value_counts().head(10)
    authors_str = "\n".join([f"{a} ({c})" for a, c in top_authors.items()])
    records.append((var_name, "Top 10 cited authors", authors_str))

    # References per document
    refs_per_doc = df_references["Document source"].value_counts()
    records.append((var_name, "Average number of references per document", round(refs_per_doc.mean(), 2)))

    return pd.DataFrame(records, columns=["Variable", "Item", "Value"])


def extract_cited_sciences(df, asjc_map_df, asjc_meta_df):
    """
    Processes the 'References' column of a DataFrame to extract, enrich, and aggregate science field occurrences
    based on parsed source titles. Returns a DataFrame with counts of cited science fields.

    Parameters:
        df (pd.DataFrame): Input DataFrame with a 'References' column containing reference strings.
        asjc_map_df (pd.DataFrame): Mapping table for ASJC codes.
        asjc_meta_df (pd.DataFrame): Metadata for ASJC codes.

    Returns:
        pd.DataFrame: Aggregated DataFrame with cited science field counts, column names prefixed with 'Cited '. 
    """
    all_counts = []

    for _, row in df.iterrows():
        references = row.get("References")
        
        if pd.notna(references):
            parsed_refs = parse_references(references)
            parsed_refs = parsed_refs.rename(columns={"Source": "Source title"})

            enriched_refs = enrich_bibliometric_data(parsed_refs, asjc_map_df, asjc_meta_df)
            science_counts = count_occurrences(enriched_refs, "Science", count_type="list", sep="; ")
        else:
            science_counts = pd.DataFrame()

        all_counts.append(science_counts)

    aggregated_counts = combine_item_dataframes(all_counts)
    aggregated_counts = aggregated_counts.rename(columns=lambda col: f"Cited {col}")

    return aggregated_counts




def compute_interdisciplinarity_entropy(df: pd.DataFrame, counting_types: list[str]) -> pd.DataFrame:
    """
    Computes interdisciplinarity (Shannon entropy) for each row based on column subsets
    defined by different counting types. Adds one column per type with entropy scores.

    Parameters:
    -----------
    df : pd.DataFrame
        The input dataframe where column names contain counting type in square brackets.
        Example: 'Physics [Number of documents]', 'Biology [Proportion of documents]', etc.

    counting_types : list of str
        A list of counting types to compute entropy for. Each string should match the
        text inside brackets exactly, e.g. "Number of documents".

    Returns:
    --------
    pd.DataFrame
        The input dataframe with additional columns:
        'Interdisciplinarity {counting_type}' for each type in counting_types.
        Rows with no data for a given type will have NaN in the corresponding column.
    """

    for ctype in counting_types:
        # Select columns that match the counting type
        pattern = re.compile(fr'\[{re.escape(ctype)}\]')
        selected_cols = [col for col in df.columns if pattern.search(col)]

        if not selected_cols:
            continue

        # Subset and normalize row-wise to get probability distributions
        data_subset = df[selected_cols].fillna(0)
        row_sums = data_subset.sum(axis=1)

        # Normalize only non-zero rows
        prob_dist = data_subset.div(row_sums, axis=0).where(row_sums != 0)

        # Compute entropy: rows with all zero will remain NaN
        entropy_values = prob_dist.apply(lambda row: entropy(row.dropna(), base=2) if row.notna().any() else np.nan, axis=1)

        df[f'Interdisciplinarity {ctype}'] = entropy_values

    df["Interdisciplinarity"] = df["Interdisciplinarity Number of documents"]
    return df



# Relations between concepts


def compute_relation_matrix(
    df1: pd.DataFrame,
    df2: pd.DataFrame = None,
    normalization: bool = False,
    pmi: bool = False,
    tfidf: bool = False,
    network: bool = False,
    eps: float = 1e-9
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], 'Union[nx.Graph, None]', pd.DataFrame, pd.DataFrame]:
    """
    Compute a co-occurrence or relation matrix from one or two document-item matrices,
    with optional TF-IDF weighting, PMI transformation, several normalization metrics,
    Fisher's exact test p-values, and optional network output.

    Parameters
    ----------
    df1 : pd.DataFrame
        Document-item matrix 1 (documents as rows, items as columns).
    df2 : pd.DataFrame, optional
        Document-item matrix 2. If None, computes co-occurrence within df1.
    normalization : bool, optional
        If True, compute all supported normalizations (see below).
    pmi : bool, optional
        If True, apply Pointwise Mutual Information (only when df2 is None or df1).
    tfidf : bool, optional
        If True, apply TF-IDF weighting to df1 (and df2 if provided).
    network : bool, optional
        If True, return the result as a NetworkX graph in addition to the matrix.
    eps : float, optional
        Small constant to avoid division or log of zero.

    Returns
    -------
    relation : pd.DataFrame
        Raw relation matrix (counts or weighted counts).
    normalized_matrices : dict[str, pd.DataFrame]
        Dictionary of all computed normalizations (each as DataFrame).
    G : nx.Graph | None
        Optional NetworkX graph if network=True, else None.
    all_measures_df : pd.DataFrame
        Wide-format DataFrame with all normalizations, counts, and conditional proportions.
        Indexed by (row item, column item).
    all_measures_df_T : pd.DataFrame
        Wide-format DataFrame with all normalizations, counts, and conditional proportions.
        Indexed by (col item, row item).

    Notes
    -----
    - All normalizations ('association', 'inclusion', 'salton', 'jaccard', 'equivalence', 'yule_q', 'fisher_p')
      are computed for both square and rectangular matrices. For rectangular (non-square) matrices, Jaccard and Equivalence measure co-occurrence ratio rather than strict similarity.
    - Yule's Q and Fisher's p-value assume binary input; results for non-binary matrices may not be interpretable.
    - The all_measures_df combines all normalizations, raw counts, and conditional proportions (row/col-based) for easy comparison and analysis.
    """


    if tfidf:
        transformer = TfidfTransformer()
        df1 = pd.DataFrame(
            transformer.fit_transform(csr_matrix(df1.values)).toarray(),
            index=df1.index,
            columns=df1.columns
        )
        if df2 is not None:
            df2 = pd.DataFrame(
                transformer.fit_transform(csr_matrix(df2.values)).toarray(),
                index=df2.index,
                columns=df2.columns
            )

    if df2 is None:
        df2 = df1

    relation = df1.T @ df2

    if pmi:
        if not df1.equals(df2):
            raise ValueError("PMI can only be applied to co-occurrence matrices (df2 must be None or df1).")
        total = relation.values.sum()
        Pi = np.diag(relation).astype(float) / total
        Pij = relation / total
        Pi = np.maximum(Pi, eps)
        Pij = np.maximum(Pij, eps)
        PMI = np.log(Pij / (Pi[:, None] * Pi[None, :]))
        PMI[PMI < 0] = 0
        relation = pd.DataFrame(PMI, index=df1.columns, columns=df2.columns)

    normalized_matrices = {}

    if normalization:
        row_sums = np.array(df1.sum(axis=0), dtype=float)
        col_sums = np.array(df2.sum(axis=0), dtype=float)
        R = relation.values

        def safe_df(matrix: np.ndarray) -> pd.DataFrame:
            return pd.DataFrame(matrix, index=df1.columns, columns=df2.columns)

        try:
            normalized_matrices["association"] = safe_df(R / (row_sums[:, None] * col_sums[None, :] + eps))
        except Exception:
            pass
        try:
            normalized_matrices["inclusion"] = safe_df(R / (np.minimum(row_sums[:, None], col_sums[None, :]) + eps))
        except Exception:
            pass
        try:
            normalized_matrices["salton"] = safe_df(R / (np.sqrt(row_sums[:, None] * col_sums[None, :]) + eps))
        except Exception:
            pass
        try:
            n_docs = df1.shape[0]
            a = R
            b = row_sums[:, None] - R
            c = col_sums[None, :] - R
            d = n_docs - (a + b + c)
            denom = (a * d + b * c + eps)
            yule_q = (a * d - b * c) / denom
            normalized_matrices["yule_q"] = safe_df(yule_q)
        except Exception:
            pass
        try:
            fisher_p = np.zeros_like(R, dtype=float)
            for i in range(R.shape[0]):
                for j in range(R.shape[1]):
                    table = np.array([
                        [a[i, j], b[i, j]],
                        [c[i, j], d[i, j]]
                    ])
                    try:
                        _, p = fisher_exact(table, alternative='two-sided')
                    except Exception:
                        p = 1.0
                    fisher_p[i, j] = p
            normalized_matrices["fisher_p"] = safe_df(fisher_p)
        except Exception:
            pass
        try:
            denom = row_sums[:, None] + col_sums[None, :] - R
            normalized_matrices["jaccard"] = safe_df(R / (denom + eps))
        except Exception:
            pass
        try:
            normalized_matrices["equivalence"] = safe_df((R ** 2) / (row_sums[:, None] * col_sums[None, :] + eps))
        except Exception:
            pass
        try:
            count_df = safe_df(R)
            prop_given_row_df = safe_df(R / (row_sums[:, None] + eps))
            prop_given_col_df = safe_df(R / (col_sums[None, :] + eps))
            stacked_dict = normalized_matrices.copy()
            stacked_dict["count"] = count_df
            stacked_dict["prop_given_row"] = prop_given_row_df
            stacked_dict["prop_given_col"] = prop_given_col_df
            measures_list = []
            for measure, df in stacked_dict.items():
                df_ = df.stack().rename(measure)
                measures_list.append(df_)
            all_measures_df = pd.concat(measures_list, axis=1)
            all_measures_df = all_measures_df.reset_index()
            all_measures_df = all_measures_df.set_index(list(all_measures_df.columns[:2]))
        except Exception:
            all_measures_df = None
    else:
        try:
            R = relation.values
            row_sums = np.array(df1.sum(axis=0), dtype=float)
            col_sums = np.array(df2.sum(axis=0), dtype=float)
            def safe_df(matrix: np.ndarray) -> pd.DataFrame:
                return pd.DataFrame(matrix, index=df1.columns, columns=df2.columns)
            count_df = safe_df(R)
            prop_given_row_df = safe_df(R / (row_sums[:, None] + eps))
            prop_given_col_df = safe_df(R / (col_sums[None, :] + eps))
            stacked_dict = {"count": count_df, "prop_given_row": prop_given_row_df, "prop_given_col": prop_given_col_df}
            measures_list = []
            for measure, df in stacked_dict.items():
                df_ = df.stack().rename(measure)
                measures_list.append(df_)
            all_measures_df = pd.concat(measures_list, axis=1)
            all_measures_df = all_measures_df.reset_index()
            all_measures_df = all_measures_df.set_index(list(all_measures_df.columns[:2]))
        except Exception:
            all_measures_df = None

    G = None
    if network:
        G = nx.from_pandas_adjacency(relation) if df1.equals(df2) else nx.from_pandas_adjacency(relation, create_using=nx.DiGraph)

    try:
        if all_measures_df is not None:
            transposed_dict = {}
            for key, df in stacked_dict.items():
                transposed_dict[key] = df.T
            measures_list_T = []
            for measure, df in transposed_dict.items():
                df_ = df.stack().rename(measure)
                measures_list_T.append(df_)
            all_measures_df_T = pd.concat(measures_list_T, axis=1)
            all_measures_df_T = all_measures_df_T.reset_index()
            all_measures_df_T = all_measures_df_T.set_index(list(all_measures_df_T.columns[:2]))
        else:
            all_measures_df_T = None
    except Exception:
        all_measures_df_T = None
    except Exception:
        all_measures_df_T = None

    return relation, normalized_matrices, G, all_measures_df, all_measures_df_T

# Analysis of these relations

def remove_zero_margins(df_relation):
    """
    Removes rows and columns from the contingency table that have zero marginal sums.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.

    Returns:
        pd.DataFrame: Cleaned table with only non-zero-sum rows and columns.
    """
    df_clean = df_relation.loc[df_relation.sum(axis=1) > 0, df_relation.sum(axis=0) > 0]
    return df_clean



def compute_diversity_metrics(df_relation: pd.DataFrame, clean_zeros: bool = True) -> dict:
    """
    Compute a variety of diversity metrics for a count-based matrix relating two concepts,
    applied both row-wise (axis=1) and column-wise (axis=0).

    Parameters:
        df_relation (pd.DataFrame): A matrix of counts (e.g., Authors × Sources or similar)
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        dict: A dictionary with keys "row_metrics" and "column_metrics",
              each containing a DataFrame of diversity scores.
    """

    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    def shannon_entropy(counts: np.ndarray) -> float:
        probs = counts / counts.sum() if counts.sum() > 0 else np.zeros_like(counts)
        probs = probs[probs > 0]
        return -np.sum(probs * np.log2(probs)) if len(probs) > 0 else 0.0

    def normalized_shannon(counts: np.ndarray) -> float:
        if counts.sum() == 0:
            return 0.0
        H = shannon_entropy(counts)
        return H / np.log2(len(counts)) if len(counts) > 1 else 0.0

    def gini_coefficient(counts: np.ndarray) -> float:
        if counts.sum() == 0:
            return 0.0
        sorted_counts = np.sort(counts)
        n = len(counts)
        cum_x = np.cumsum(sorted_counts)
        return (n + 1 - 2 * np.sum(cum_x) / cum_x[-1]) / n

    def herfindahl_index(counts: np.ndarray) -> float:
        if counts.sum() == 0:
            return 0.0
        proportions = counts / counts.sum()
        return np.sum(proportions**2)

    def simpson_index(counts: np.ndarray) -> float:
        if counts.sum() == 0:
            return 0.0
        proportions = counts / counts.sum()
        return 1 - np.sum(proportions**2)

    def richness(counts: np.ndarray) -> int:
        return np.count_nonzero(counts)

    metric_functions = {
        "Shannon": shannon_entropy,
        "Normalized Shannon": normalized_shannon,
        "Gini": gini_coefficient,
        "HHI": herfindahl_index,
        "Simpson": simpson_index,
        "Richness": richness
    }

    row_metrics = pd.DataFrame(index=df_relation.index)
    column_metrics = pd.DataFrame(index=df_relation.columns)

    for name, func in metric_functions.items():
        row_metrics[name] = df_relation.apply(func, axis=1)
        column_metrics[name] = df_relation.apply(func, axis=0)

    return {
        "row_metrics": row_metrics,
        "column_metrics": column_metrics
    }



def analyze_bipartite_relation(df_relation: pd.DataFrame, stats: list = None, clean_zeros: bool = True) -> dict:
    """
    Perform bipartite network analysis on a relation matrix, including projections and metrics.

    Parameters:
        df_relation (pd.DataFrame): Count matrix with rows and columns as distinct concepts.
        stats (list): List of node-level stats to compute on projections.
                      Supported: "degree", "strength", "betweenness", "closeness",
                                 "eigenvector", "pagerank", "clustering", "triangle_count".
                      If None, all are computed.
        clean_zeros (bool): Whether to remove rows and columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "bipartite_graph": B,
            "row_projection": G_row,
            "column_projection": G_col,
            "row_stats": pd.DataFrame,
            "column_stats": pd.DataFrame,
            "bipartite_global": dict,
            "row_global": dict,
            "column_global": dict
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    B = nx.Graph()
    row_nodes = df_relation.index.tolist()
    col_nodes = df_relation.columns.tolist()

    # Add nodes
    B.add_nodes_from(row_nodes, bipartite=0)
    B.add_nodes_from(col_nodes, bipartite=1)

    # Add weighted edges
    for row in row_nodes:
        for col, weight in df_relation.loc[row][df_relation.loc[row] > 0].items():
            B.add_edge(row, col, weight=weight)

    # Projections
    G_row = nx.bipartite.weighted_projected_graph(B, row_nodes)
    G_col = nx.bipartite.weighted_projected_graph(B, col_nodes)

    if stats is None:
        stats = [
            "degree", "strength", "betweenness", "closeness",
            "eigenvector", "pagerank", "clustering", "triangle_count"
        ]

    def compute_node_stats(G):
        data = {}
        if "degree" in stats:
            data["Degree"] = dict(G.degree())
        if "strength" in stats:
            data["Strength"] = dict(G.degree(weight="weight"))
        if "betweenness" in stats:
            data["Betweenness"] = nx.betweenness_centrality(G, weight="weight")
        if "closeness" in stats:
            data["Closeness"] = nx.closeness_centrality(G)
        if "eigenvector" in stats:
            try:
                data["Eigenvector"] = nx.eigenvector_centrality(G, weight="weight", max_iter=500)
            except nx.PowerIterationFailedConvergence:
                data["Eigenvector"] = {node: np.nan for node in G.nodes()}
        if "pagerank" in stats:
            data["PageRank"] = nx.pagerank(G, weight="weight")
        if "clustering" in stats:
            data["Clustering"] = nx.clustering(G, weight="weight")
        if "triangle_count" in stats:
            data["Triangles"] = nx.triangles(G)
        return pd.DataFrame(data)

    def compute_global_stats(G, weighted: bool = False, clustering: bool = False):
        degrees = dict(G.degree())
        strengths = dict(G.degree(weight="weight")) if weighted else None
        largest_cc = max(nx.connected_components(G), key=len) if not nx.is_connected(G) else G.nodes
        return {
            "Nodes": G.number_of_nodes(),
            "Edges": G.number_of_edges(),
            "Density": nx.density(G),
            "AvgDegree": np.mean(list(degrees.values())) if degrees else 0,
            "AvgStrength": np.mean(list(strengths.values())) if strengths else 0,
            "Components": nx.number_connected_components(G),
            "LargestComponentSize": len(largest_cc),
            "AvgClustering": nx.average_clustering(G, weight="weight") if clustering else None
        }

    return {
        "bipartite_graph": B,
        "row_projection": G_row,
        "column_projection": G_col,
        "row_stats": compute_node_stats(G_row),
        "column_stats": compute_node_stats(G_col),
        "bipartite_global": compute_global_stats(B),
        "row_global": compute_global_stats(G_row, weighted=True, clustering=True),
        "column_global": compute_global_stats(G_col, weighted=True, clustering=True)
    }

# clustering on relationship matrix



def cluster_relation_matrix(
    df_relation: pd.DataFrame,
    method: str = "kmeans",
    axis: int = 0,
    scale: bool = True,
    k_range: tuple = (2, 10),
    n_clusters: int = None,
    linkage_method: str = "ward",
    return_scores: bool = False,
    clean_zeros: bool = True
) -> dict:
    """
    Cluster rows or columns of a relation matrix using various clustering methods.

    Parameters:
        df_relation (pd.DataFrame): Matrix of counts between two concepts.
        method (str): Clustering method: "kmeans", "hierarchical", "spectral".
        axis (int): 0 for columns, 1 for rows.
        scale (bool): Whether to standardize data (recommended for KMeans/Spectral).
        k_range (tuple): Range for automatic KMeans cluster selection (if n_clusters not given).
        n_clusters (int): Number of clusters to use (if known; overrides k_range).
        linkage_method (str): Linkage method for hierarchical clustering.
        return_scores (bool): Whether to return silhouette scores (only for KMeans).
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "clusters": pd.Series of cluster labels,
            "n_clusters": number of clusters,
            "silhouette_scores": dict of silhouette scores (if applicable)
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    if axis == 1:
        X = df_relation.values
        labels = df_relation.index
    else:
        X = df_relation.T.values
        labels = df_relation.columns

    if scale and method in ["kmeans", "spectral"]:
        X = StandardScaler().fit_transform(X)

    scores = {}
    clusters = None
    best_k = None

    if method == "kmeans":
        if n_clusters is None:
            for k in range(k_range[0], k_range[1] + 1):
                model = KMeans(n_clusters=k, random_state=42)
                labels_k = model.fit_predict(X)
                if len(np.unique(labels_k)) > 1:
                    scores[k] = silhouette_score(X, labels_k)
            best_k = max(scores, key=scores.get)
            model = KMeans(n_clusters=best_k, random_state=42)
        else:
            best_k = n_clusters
            model = KMeans(n_clusters=best_k, random_state=42)
        final_labels = model.fit_predict(X)
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    elif method == "hierarchical":
        dist = pdist(X)
        Z = linkage(dist, method=linkage_method)
        best_k = n_clusters or 5
        final_labels = fcluster(Z, best_k, criterion="maxclust")
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    elif method == "spectral":
        best_k = n_clusters or 5
        model = SpectralClustering(n_clusters=best_k, affinity="nearest_neighbors", random_state=42)
        final_labels = model.fit_predict(X)
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    else:
        raise ValueError(f"Unsupported method: {method}")

    result = {
        "clusters": clusters,
        "n_clusters": best_k
    }

    if method == "kmeans" and return_scores:
        result["silhouette_scores"] = scores

    return result

def bicluster_relation_matrix(
    df_relation: pd.DataFrame,
    n_clusters: int = 5,
    scale: bool = True,
    clean_zeros: bool = True
) -> dict:
    """
    Perform biclustering on a relation matrix using Spectral Co-clustering.

    Parameters:
        df_relation (pd.DataFrame): Matrix of counts between two concepts.
        n_clusters (int): Number of biclusters to form.
        scale (bool): Whether to standardize the matrix before clustering.
        clean_zeros (bool): Whether to remove rows and columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "model": fitted SpectralCoclustering model,
            "row_clusters": pd.Series with cluster labels for rows,
            "column_clusters": pd.Series with cluster labels for columns
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    X = df_relation.values
    if scale:
        X = StandardScaler().fit_transform(X)

    model = SpectralCoclustering(n_clusters=n_clusters, random_state=42)
    model.fit(X)

    row_clusters = pd.Series(model.row_labels_, index=df_relation.index, name="RowCluster")
    col_clusters = pd.Series(model.column_labels_, index=df_relation.columns, name="ColumnCluster")

    return {
        "model": model,
        "row_clusters": row_clusters,
        "column_clusters": col_clusters
    }



def compute_correspondence_analysis(df_relation, n_components=2, clean_zeros=True):
    """
    Performs correspondence analysis with explicit SVD to compute explained inertia.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        n_components (int): Number of CA components.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        row_coords (pd.DataFrame): Coordinates of rows.
        col_coords (pd.DataFrame): Coordinates of columns.
        explained_inertia (list): Proportion of inertia explained by each component.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    # CA for coordinates
    ca = prince.CA(n_components=n_components, n_iter=10, engine='sklearn', random_state=42)
    ca = ca.fit(df_relation)
    row_coords = ca.row_coordinates(df_relation)
    col_coords = ca.column_coordinates(df_relation)

    # Step 1: Compute correspondence matrix
    N = df_relation.values
    n_total = N.sum()
    P = N / n_total  # relative frequency matrix

    # Step 2: Compute row and column marginal frequencies
    r = P.sum(axis=1, keepdims=True)
    c = P.sum(axis=0, keepdims=True)

    # Step 3: Compute matrix of standardized residuals (centered and scaled)
    E = r @ c  # expected frequency matrix under independence
    S = (P - E) / np.sqrt(r @ c)

    # Step 4: Truncated SVD to get singular values
    U, Sigma, VT = randomized_svd(S, n_components=n_components, random_state=42)
    inertia = Sigma**2
    explained_inertia = (inertia / inertia.sum()).tolist()

    return row_coords, col_coords, explained_inertia


def extract_sorted_residual_pairs(df_relation, clean_zeros: bool = True):
    """
    Computes Pearson residuals and returns sorted (row, column, residual) triplets
    by descending absolute residual value. Optionally removes zero-marginal rows/columns.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums.

    Returns:
        sorted_residuals (pd.DataFrame): Columns: ['Row', 'Column', 'Residual'] sorted by abs.
        expected_df (pd.DataFrame): Expected counts under independence.
        chi2_stat (float): Total chi-squared statistic.
        dof (int): Degrees of freedom.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    # Compute expected and residuals
    observed = df_relation.values
    row_totals = observed.sum(axis=1, keepdims=True)
    col_totals = observed.sum(axis=0, keepdims=True)
    grand_total = observed.sum()

    expected = row_totals @ col_totals / grand_total
    residuals = (observed - expected) / np.sqrt(expected)

    residuals_df = pd.DataFrame(residuals, index=df_relation.index, columns=df_relation.columns)

    expected_df = pd.DataFrame(expected, index=df_relation.index, columns=df_relation.columns)

    residuals_flat = [
        (row_label, col_label, residuals[i, j])
        for i, row_label in enumerate(df_relation.index)
        for j, col_label in enumerate(df_relation.columns)
    ]

    sorted_residuals = pd.DataFrame(residuals_flat, columns=['Row', 'Column', 'Residual'])
    sorted_residuals['AbsResidual'] = sorted_residuals['Residual'].abs()
    sorted_residuals = sorted_residuals.sort_values(by='AbsResidual', ascending=False).drop(columns='AbsResidual')

    chi2_stat = np.sum((observed - expected) ** 2 / expected)
    dof = (df_relation.shape[0] - 1) * (df_relation.shape[1] - 1)

    return residuals_df, sorted_residuals, expected_df, chi2_stat, dof


def compute_svd_statistics(df_relation, n_components=2, clean_zeros=True):
    """
    Applies Truncated SVD to a normalized contingency table.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        n_components (int): Number of SVD components to retain.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        row_projection (pd.DataFrame): Row projections onto components.
        singular_values (np.ndarray): Singular values of components.
        explained_variance (np.ndarray): Explained variance ratio per component.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    # Normalize (relative frequency)
    normed = df_relation / df_relation.values.sum()

    # Apply SVD
    svd_model = TruncatedSVD(n_components=n_components, random_state=42)
    row_proj = svd_model.fit_transform(normed)

    # Format output
    row_projection = pd.DataFrame(row_proj, index=df_relation.index,
                                  columns=[f'Comp {i+1}' for i in range(n_components)])
    singular_values = svd_model.singular_values_
    explained_variance = svd_model.explained_variance_ratio_

    return row_projection, singular_values, explained_variance


def compute_log_ratio(df_relation, clean_zeros: bool = True):
    """
    Computes log(observed / expected) values from a contingency table.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums.

    Returns:
        log_ratio_df (pd.DataFrame): Matrix of log(observed / expected) values.
        expected_df (pd.DataFrame): Expected frequency matrix under independence.
        sorted_log_ratios (pd.DataFrame): Flattened DataFrame with ['Row', 'Column', 'LogRatio'],
                                          sorted by LogRatio descending.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    observed = df_relation.values
    row_totals = observed.sum(axis=1, keepdims=True)
    col_totals = observed.sum(axis=0, keepdims=True)
    total = observed.sum()
    expected = row_totals @ col_totals / total

    log_ratio = np.log((observed + 1e-6) / (expected + 1e-6))
    log_ratio_df = pd.DataFrame(log_ratio, index=df_relation.index, columns=df_relation.columns)
    expected_df = pd.DataFrame(expected, index=df_relation.index, columns=df_relation.columns)

    # Flatten and sort the log-ratio values
    flattened = [
        (row_label, col_label, log_ratio[i, j])
        for i, row_label in enumerate(df_relation.index)
        for j, col_label in enumerate(df_relation.columns)
    ]
    sorted_log_ratios = pd.DataFrame(flattened, columns=['Row', 'Column', 'LogRatio'])
    sorted_log_ratios = sorted_log_ratios.sort_values(by='LogRatio', ascending=False).reset_index(drop=True)

    return log_ratio_df, expected_df, sorted_log_ratios


# Netwrok analysis

def export_graph_formats(G: nx.Graph, filename_base: str, output_dir: str = ".") -> None:
    """
    Export a NetworkX graph to GraphML, GEXF, and Pajek NET formats.

    Parameters:
        G (nx.Graph): The graph to export.
        filename_base (str): Base filename without extension.
        output_dir (str): Directory to save files (default is current directory).

    Returns:
        None
    """
    os.makedirs(output_dir, exist_ok=True)

    path_graphml = os.path.join(output_dir, f"{filename_base}.graphml")
    path_gexf = os.path.join(output_dir, f"{filename_base}.gexf")
    path_net = os.path.join(output_dir, f"{filename_base}.net")

    nx.write_graphml(G, path_graphml)
    nx.write_gexf(G, path_gexf)
    nx.write_pajek(G, path_net)

    print(f"Graphs exported to:\n- {path_graphml}\n- {path_gexf}\n- {path_net}")
    
    
    
# functions for group analysis



def generate_group_matrix(
    df: pd.DataFrame,
    group_desc: Union[str, pd.DataFrame, Dict[str, str], Dict[str, List[str]]],
    force_type: Optional[str] = None,
    sep: str = "; ",
    text_column: Optional[str] = None,
    regex_flags: int = re.IGNORECASE,
    top_n: Optional[int] = None,
    binary_as_int: bool = False,
    include_items: Optional[List[str]] = None,
    exclude_items: Optional[List[str]] = None,
    year_column: Optional[str] = None,
    year_range: Optional[Tuple[int, int]] = None,
    invert_matrix: bool = False,
    cutpoints: Optional[List[int]] = None,
    n_periods: Optional[int] = None,
    cut_labels: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    Generate a binary (document × group) matrix from different group descriptions.

    Parameters
    ----------
    df : pd.DataFrame
        Main dataframe with documents as rows.
    group_desc : Union[str, pd.DataFrame, Dict[str, str], Dict[str, List[str]]]
        Describes group definitions.
    force_type : Optional[str], default=None
        Options: "column", "multiitem", "regex", "binary", "year".
    sep : str, default="; "
        Separator for multi-item group column.
    text_column : Optional[str]
        Required for regex or dict-based matching.
    regex_flags : int, default=re.IGNORECASE
        Flags passed to regex matching.
    top_n : Optional[int]
        Limit to most frequent N items (multi-item only).
    binary_as_int : bool
        Convert boolean matrix to 0/1 integers.
    include_items : Optional[List[str]]
        Keep only these items (multi-item only).
    exclude_items : Optional[List[str]]
        Remove these items (multi-item only).
    year_column : Optional[str]
        Column for time-based row filtering.
    year_range : Optional[Tuple[int, int]]
        Range of years to keep in analysis.
    invert_matrix : bool
        Invert True/False values in the output matrix.
    cutpoints : Optional[List[int]]
        Year cutpoints for defining time intervals.
    n_periods : Optional[int]
        Automatically divide time into this many bins.
    cut_labels : Optional[List[str]]
        Labels for time bins (must match number of bins if provided).

    Returns
    -------
    pd.DataFrame
        Binary dataframe with group membership per document.
    """

    def make_non_capturing(pattern: str) -> str:
        return re.sub(r"\((?!\?)", "(?:", pattern)

    def limit_top_n(binary_df: pd.DataFrame) -> pd.DataFrame:
        if top_n is not None and top_n < binary_df.shape[1]:
            top_cols = binary_df.sum().sort_values(ascending=False).head(top_n).index
            binary_df = binary_df[top_cols]
        return binary_df

    def filter_items(binary_df: pd.DataFrame) -> pd.DataFrame:
        cols = set(binary_df.columns)
        if include_items is not None:
            cols &= set(include_items)
        if exclude_items is not None:
            cols -= set(exclude_items)
        return binary_df[list(cols)] if cols else pd.DataFrame(index=binary_df.index)

    def finalize(binary_df: pd.DataFrame) -> pd.DataFrame:
        binary_df = binary_df.astype(int) if binary_as_int else binary_df.astype(bool)
        return ~binary_df if invert_matrix else binary_df

    # Apply year-based row filtering
    if year_column is not None and year_range is not None:
        start, end = year_range
        df = df.copy()
        year_vals = pd.to_numeric(df[year_column], errors="coerce")
        df = df[year_vals.between(start, end)]

    if df.empty:
        return pd.DataFrame(columns=[], index=df.index)

    # Forced types
    if force_type == "binary":
        assert isinstance(group_desc, pd.DataFrame), "Expected binary dataframe."
        return finalize(group_desc)

    if force_type == "column":
        assert isinstance(group_desc, str), "Expected column name."
        filtered = df[group_desc].dropna()
        binary_df = pd.get_dummies(filtered)
        return finalize(binary_df.reindex(df.index, fill_value=0))

    if force_type == "multiitem":
        assert isinstance(group_desc, str), "Expected column name."
        exploded = df[group_desc].dropna().str.split(sep).explode().str.strip()
        exploded = exploded[exploded != ""]
        binary_df = pd.get_dummies(exploded).groupby(level=0).max()
        binary_df = limit_top_n(binary_df)
        binary_df = filter_items(binary_df)
        return finalize(binary_df.reindex(df.index, fill_value=0))

    if force_type == "regex":
        assert isinstance(group_desc, dict), "Expected dict for regex type."
        assert text_column is not None, "text_column must be specified."
        binary = pd.DataFrame(False, index=df.index, columns=list(group_desc.keys()))
        valid_text = df[text_column].fillna("")
        for group, pattern in group_desc.items():
            safe_pattern = make_non_capturing(pattern)
            binary[group] = valid_text.str.contains(safe_pattern, flags=regex_flags, regex=True)
        return finalize(binary)

    # Auto detection
    if isinstance(group_desc, pd.DataFrame):
        return finalize(group_desc)

    if isinstance(group_desc, dict):
        if text_column is None:
            raise ValueError("text_column must be specified for dict group_desc.")
        binary = pd.DataFrame(False, index=df.index, columns=list(group_desc.keys()))
        valid_text = df[text_column].fillna("")
        for group, matcher in group_desc.items():
            pattern = "|".join([re.escape(m) for m in matcher]) if isinstance(matcher, list) else make_non_capturing(matcher)
            binary[group] = valid_text.str.contains(pattern, flags=regex_flags, regex=True)
        return finalize(binary)

    if isinstance(group_desc, str):
        series = df[group_desc].dropna().astype(str)

        # Multiitem?
        if series.str.contains(sep).any():
            exploded = series.str.split(sep).explode().str.strip()
            exploded = exploded[exploded != ""]
            binary_df = pd.get_dummies(exploded).groupby(level=0).max()
            binary_df = limit_top_n(binary_df)
            binary_df = filter_items(binary_df)
            return finalize(binary_df.reindex(df.index, fill_value=0))

        # Looks like a year column?
        numeric = pd.to_numeric(series, errors="coerce").dropna()
        is_year = numeric.between(1900, 2100).mean() > 0.8

        if is_year and (cutpoints is not None or n_periods is not None):
            valid_years = numeric.astype(int)
            if cutpoints is not None:
                bins = [-np.inf] + cutpoints + [np.inf]
            else:
                min_y, max_y = valid_years.min(), valid_years.max()
                bins = np.linspace(min_y, max_y + 1, n_periods + 1).astype(int)
            labels = cut_labels if cut_labels else [f"{bins[i]}–{bins[i+1]-1}" for i in range(len(bins) - 1)]
            binned = pd.cut(valid_years, bins=bins, labels=labels, right=False)
            binary_df = pd.get_dummies(binned)
            return finalize(binary_df.reindex(df.index, fill_value=0))

        if is_year:
            return finalize(pd.get_dummies(numeric.astype(int)).reindex(df.index, fill_value=0))

        binary_df = pd.get_dummies(series)
        return finalize(binary_df.reindex(df.index, fill_value=0))

    raise ValueError("Unable to interpret group_desc. Consider specifying force_type.")
    
    
def merge_group_performances(group_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Merge performance dataframes from different groups by aligning on 'Variable' and 'Indicator' columns.

    Args:
        group_dfs (dict[str, pd.DataFrame]): Dictionary where keys are group names and values are
                                             corresponding performance dataframes with columns
                                             ['Variable', 'Indicator', 'Value'].

    Returns:
        pd.DataFrame: Merged dataframe with 'Variable' and 'Indicator' as index columns and one column
                      for each group's performance values.
    """
    merged_df = None
    for group_name, df in group_dfs.items():
        temp_df = df.copy()
        temp_df = temp_df.rename(columns={'Value': group_name})
        if merged_df is None:
            merged_df = temp_df
        else:
            merged_df = pd.merge(merged_df, temp_df, on=['Variable', 'Indicator'], how='outer')
    return merged_df





def count_occurrences_across_groups(groups, group_matrix, count_func_name, merge_type="all items", **kwargs):
    """
    Count item occurrences (e.g., keywords, authors, sources) across multiple groups
    and merge the results into a single DataFrame with renamed columns indicating the group.

    Args:
        groups (dict): Mapping of group names to group objects.
        group_matrix (pd.DataFrame): DataFrame with group names as columns.
        count_func_name (str): Name of the counting method to call on each group.
        merge_type (str): Either "all items" (outer join) or "shared items" (inner join).
        **kwargs: Additional keyword arguments passed to the counting function.

    Returns:
        pd.DataFrame: Merged DataFrame of counts with group-labeled columns, NaNs replaced by 0.
    """
    how_map = {"all items": "outer", "shared items": "inner"}
    how = how_map.get(merge_type)
    if how is None:
        raise ValueError('merge_type must be "all items" or "shared items"')

    dfs = [
        (df := getattr(groups[g], count_func_name)(**kwargs)).rename(columns={c: f"{c} ({g})" for c in df.columns[1:]})
        for g in group_matrix.columns
    ]
    merged = reduce(lambda l, r: pd.merge(l, r, on=dfs[0].columns[0], how=how), dfs).fillna(0)

    for col in merged.columns[1:]:
        merged[col] = pd.to_numeric(merged[col], errors="ignore")

    return merged


def compute_group_intersections(group_matrix: pd.DataFrame, include_ids: bool = False, id_column: pd.Series = None) -> pd.DataFrame:
    """
    Computes all unique intersections from a binary group membership matrix.

    Parameters:
    - group_matrix: pd.DataFrame
        Binary DataFrame with rows as items (e.g., documents) and columns as group names (0/1).
    - include_ids: bool
        If True, includes a column listing the item IDs from `id_column` (Series aligned with group_matrix) or the index if not provided.
    - id_column: pd.Series or None
        Optional Series providing IDs for the items (same index as group_matrix). If None, uses the index.

    Returns:
    - pd.DataFrame with columns:
        - 'Groups': tuple of intersecting groups
        - 'Size': number of items in the intersection
        - 'ID' (optional): list of item IDs in the intersection
    """
    group_cols = group_matrix.columns.tolist()
    results = []

    for mask, df_subset in group_matrix.groupby(group_cols):
        active_groups = tuple(col for col, flag in zip(group_cols, mask) if flag == 1)
        if not active_groups:
            continue
        if include_ids:
            ids = id_column.loc[df_subset.index].tolist() if id_column is not None else df_subset.index.tolist()
            row = {"Groups": active_groups, "Size": len(df_subset), "ID": ids}
        else:
            row = {"Groups": active_groups, "Size": len(df_subset)}
        results.append(row)

    return pd.DataFrame(results).sort_values(by="Size", ascending=False).reset_index(drop=True)

def compute_group_similarity_matrices(group_matrix: pd.DataFrame, methods: list = ["jaccard"]) -> dict:
    """
    Computes group × group similarity matrices for given methods.

    Parameters:
    - group_matrix: pd.DataFrame
        Binary matrix (0/1), rows = items, columns = group names.
    - methods: list of str
        List of methods to compute. Supported: 'jaccard', 'count', 'sokal-michener', 'simple-matching', 'rogers-tanimoto'.

    Returns:
    - dict: keys are method names, values are DataFrames (group × group similarity matrices)
    """



    supported_methods = {
        "jaccard": "Jaccard Index",
        "count": "Shared Items",
        "sokal-michener": "Sokal-Michener",
        "simple-matching": "Simple Matching",
        "rogers-tanimoto": "Rogers-Tanimoto"
    }

    binary = group_matrix.astype(bool).to_numpy()
    groups = group_matrix.columns.tolist()
    matrices = {}

    for method in methods:
        if method not in supported_methods:
            print(f"Warning: Unsupported method '{method}', skipping.")
            continue

        if method == "count":
            mat = pd.DataFrame(index=groups, columns=groups, dtype=float)
            for i, g1 in enumerate(groups):
                set1 = set(group_matrix.index[group_matrix[g1] == 1])
                for j, g2 in enumerate(groups):
                    set2 = set(group_matrix.index[group_matrix[g2] == 1])
                    mat.loc[g1, g2] = len(set1 & set2)
        else:
            dist = pairwise_distances(binary.T, metric=method)
            sim = 1 - dist
            mat = pd.DataFrame(sim, index=groups, columns=groups)

        matrices[method] = mat

    return matrices

def compare_continuous_by_binary_groups(df, numerical_cols, group_matrix, output_format="long"):
    """
    Compare continuous variables across groups defined by a binary group matrix using both parametric and 
    non-parametric statistical tests, while computing descriptive statistics and handling missing values.

    Parameters:
    df (pd.DataFrame): Original dataframe containing continuous numerical data.
    numerical_cols (list of str): List of column names in df that contain continuous numerical variables.
    group_matrix (pd.DataFrame): Binary matrix (same number of rows as df) where each column defines a group 
                                  (1 = in group, 0 = not in group).
    output_format (str): "long" (default) for detailed per-group stats; "wide" for matrix-style summary of stats and p-values.

    Returns:
    pd.DataFrame: DataFrame containing descriptive statistics and p-values for each variable and group.
    """
    results = []

    for col in numerical_cols:
        group_data = []
        group_names = []
        group_stats = {}

        for group in group_matrix.columns:
            valid_mask = group_matrix[group] == 1
            data = df.loc[valid_mask, col]
            data = data[np.isfinite(data)]

            if len(data) >= 3:
                group_data.append(data)
                group_names.append(group)

                group_stats[group] = {
                    "Mean": data.mean(),
                    "SD": data.std(),
                    "Median": data.median(),
                    "IQR": data.quantile(0.75) - data.quantile(0.25),
                    "Shapiro p": shapiro(data)[1] if len(data) <= 5000 else np.nan,
                    "N": len(data)
                }

        if len(group_data) >= 2:
            try:
                param_p = f_oneway(*group_data)[1]
            except ValueError:
                param_p = np.nan

            try:
                nonparam_p = kruskal(*group_data)[1]
            except ValueError:
                nonparam_p = np.nan

            for group in group_names:
                stats = group_stats[group]
                results.append({
                    "Variable": col,
                    "Group": group,
                    "Mean": stats["Mean"],
                    "SD": stats["SD"],
                    "Median": stats["Median"],
                    "IQR": stats["IQR"],
                    "Shapiro p": stats["Shapiro p"],
                    "N": stats["N"],
                    "Parametric p (ANOVA)": param_p,
                    "Non-parametric p (Kruskal)": nonparam_p
                })

    long_df = pd.DataFrame(results)

    if output_format == "wide":
        # Pivot group-level descriptives into wide format
        descriptives = long_df.pivot(index="Variable", columns="Group", values=["Mean", "SD", "Median", "IQR", "Shapiro p", "N"])
        descriptives.columns = [f"{stat} ({grp})" for stat, grp in descriptives.columns]

        # Add p-values to wide format
        p_values = long_df[["Variable", "Parametric p (ANOVA)", "Non-parametric p (Kruskal)"]].drop_duplicates().set_index("Variable")
        wide_df = pd.concat([descriptives, p_values], axis=1).reset_index()
        return wide_df
    else:
        return long_df





# Time series analysis

def aggregate_bibliometrics_by_group_and_year(df, binary_df, 
                                              group_columns=None, 
                                              year_column="Year", 
                                              metrics=["Cited by"],
                                              additional_binary_dfs=None,
                                              normalize=True,
                                              include_cumulative=False,
                                              include_percentage=False,
                                              group_selection_top_n=None,
                                              group_selection_include_regex=None,
                                              group_selection_exclude_regex=None,
                                              aggfunc="sum",
                                              return_format="wide"):
    """
    Aggregate bibliometric indicators by group and year using one or more binary dataframes.

    Parameters
    ----------
    df : pd.DataFrame
        The main dataframe with bibliographic data, including a "Year" column.
    binary_df : pd.DataFrame
        A binary dataframe where each column represents a group of interest.
    group_columns : list of str, optional
        Subset of binary_df columns to include in the analysis. If None, use all (after optional selection filters).
    year_column : str, default "Year"
        Name of the column in df indicating the year of publication.
    metrics : list of str
        Column names in df or additional binary dataframes to aggregate.
    additional_binary_dfs : dict of {str: pd.DataFrame}, optional
        Dictionary of binary dataframes to be used for metric calculation, where key is a metric name.
    normalize : bool, default True
        If True, include both raw and normalized metrics by document count.
    include_cumulative : bool, default False
        If True, include cumulative values.
    include_percentage : bool, default False
        If True, include percentage of yearly totals for each metric.
    group_selection_top_n : int, optional
        If specified, use only the top N columns in binary_df with the most total counts.
    group_selection_include_regex : str, optional
        A regex pattern to include only matching columns.
    group_selection_exclude_regex : str, optional
        A regex pattern to exclude matching columns.
    aggfunc : str or callable, default "sum"
        Aggregation function to use for metrics (e.g., "sum", "mean", etc.).
    return_format : str, default "wide"
        Format of the result: "long" for tidy format, "wide" for pivoted format, or "both" to return a dictionary.

    Returns
    -------
    pd.DataFrame or dict of DataFrames
        A dataframe with aggregated results by group and year or both formats.
    """

    all_columns = binary_df.columns

    if group_selection_include_regex:
        all_columns = [col for col in all_columns if re.search(group_selection_include_regex, col)]

    if group_selection_exclude_regex:
        all_columns = [col for col in all_columns if not re.search(group_selection_exclude_regex, col)]

    if group_selection_top_n is not None:
        top_cols = binary_df[all_columns].sum().sort_values(ascending=False).head(group_selection_top_n).index.tolist()
        all_columns = top_cols

    if group_columns is None:
        group_columns = all_columns

    results_wide = []
    results_long = []

    full_year_range = pd.Index(sorted(df[year_column].dropna().unique()))

    for group in group_columns:
        mask = binary_df[group] == 1
        sub_df = df.loc[mask].copy()
        sub_df["__group_marker__"] = 1

        base = sub_df[[year_column, "__group_marker__"]].copy()
        base = base.groupby(year_column).count().rename(columns={"__group_marker__": "Number of documents"})
        base = base.reindex(full_year_range, fill_value=0)

        metric_data = base.copy()

        for metric in metrics:
            is_additional = additional_binary_dfs and metric in additional_binary_dfs

            if metric in df.columns:
                values = sub_df.groupby(year_column)[metric].agg(aggfunc).reindex(full_year_range, fill_value=0)
            elif is_additional:
                bin_df = additional_binary_dfs[metric]
                values = bin_df.loc[mask].groupby(df[year_column]).sum().sum(axis=1).reindex(full_year_range, fill_value=0)
            else:
                continue

            metric_data[metric] = values

            if normalize and metric in metric_data.columns:
                norm_col = f"{metric} (Per Document)"
                metric_data[norm_col] = metric_data[metric] / metric_data["Number of documents"].replace(0, pd.NA)

                if is_additional:
                    percent_col = f"{metric} (Percentage of Group)"
                    metric_data[percent_col] = metric_data[metric] / metric_data["Number of documents"].replace(0, pd.NA) * 100

        if include_cumulative:
            for metric in metric_data.columns:
                metric_data[f"{metric} (Cumulative)"] = metric_data[metric].cumsum()

        if include_percentage:
            total_per_year = metric_data.sum(axis=0, numeric_only=True)
            for metric in metrics:
                if metric in metric_data.columns:
                    metric_data[f"{metric} (Percentage)"] = metric_data[metric] / total_per_year[metric] * 100

        metric_data["Group"] = group
        metric_data = metric_data.reset_index().rename(columns={"index": year_column})

        wide_df = metric_data.copy()
        results_wide.append(wide_df)

        long_df = wide_df.melt(id_vars=[year_column, "Group"], var_name="Metric", value_name="Value")
        results_long.append(long_df)

    wide_result = pd.concat(results_wide, ignore_index=True)
    long_result = pd.concat(results_long, ignore_index=True)

    if return_format == "long":
        return long_result
    elif return_format == "wide":
        return wide_result
    elif return_format == "both":
        return {"wide": wide_result, "long": long_result}



# Group analysis

def get_scientific_production_by_group0(df, group_matrix, relative_counts=True, cumulative=True, predict_last_year=True, percent_change=True, output_format="both", rename_wide_columns=True):
    """
    Computes the annual scientific production statistics separately for each group and returns the results in long, wide, or both formats.

    Parameters:
    df (pd.DataFrame): DataFrame containing at least "Year" and "Cited by" columns.
    group_matrix (pd.DataFrame): Binary DataFrame where columns are group names and rows align with df rows.
    relative_counts (bool, optional): Whether to compute relative proportions and percentages of documents. Default is True.
    cumulative (bool, optional): Whether to compute cumulative document and citation counts. Default is True.
    predict_last_year (bool, optional): Whether to predict the current year's document and citation counts. Default is True.
    percent_change (bool, optional): Whether to compute year-over-year percentage change in documents and citations. Default is True.
    output_format (str, optional): Format of the result: "long", "wide", or "both". Default is "both".
    rename_wide_columns (bool, optional): Whether to rename wide-format columns to include group names. Default is True.

    Returns:
    pd.DataFrame or dict: A single DataFrame in long or wide format, or a dictionary with both if output_format="both".
    """
    group_dfs = []
    all_years = pd.Series(range(df["Year"].min(), df["Year"].max() + 1), name="Year")

    for group in group_matrix.columns:
        group_rows = group_matrix[group] == 1
        if group_rows.any():
            group_df = df[group_rows].copy()
            production = get_scientific_production(
                group_df,
                relative_counts=relative_counts,
                cumulative=cumulative,
                predict_last_year=predict_last_year,
                percent_change=percent_change
            )
            production = all_years.to_frame().merge(production, on="Year", how="left").fillna(0)
            production["Group"] = group
            group_dfs.append(production)

    if not group_dfs:
        return pd.DataFrame()

    long_df = pd.concat(group_dfs, ignore_index=True)

    if output_format == "long":
        return long_df

    elif output_format == "wide":
        id_vars = ["Year", "Group"]
        value_vars = [col for col in long_df.columns if col not in id_vars]
        wide_df = long_df.pivot(index="Year", columns="Group", values=value_vars)
        wide_df = wide_df.fillna(0)

        if rename_wide_columns:
            wide_df.columns = [f"{val} (" + f"{grp})" for val, grp in wide_df.columns]
            wide_df.columns.name = None

        return wide_df

    elif output_format == "both":
        id_vars = ["Year", "Group"]
        value_vars = [col for col in long_df.columns if col not in id_vars]
        wide_df = long_df.pivot(index="Year", columns="Group", values=value_vars)
        wide_df = wide_df.fillna(0)

        if rename_wide_columns:
            wide_df.columns = [f"{val} (" + f"{grp})" for val, grp in wide_df.columns]
            wide_df.columns.name = None

        return {"long": long_df, "wide": wide_df}

    else:
        raise ValueError("output_format must be one of: \"long\", \"wide\", \"both\"")
        
        
# Bibliographic laws

def compute_lotka_distribution(df, author_col="Authors", separator="; "):
    """
    Compute author productivity distribution and expected values under Lotka's law.

    Parameters:
        df (pd.DataFrame): DataFrame containing author information.
        author_col (str): Column name where author names are listed.
        separator (str): Separator between multiple authors in a single cell.

    Returns:
        pd.DataFrame: A DataFrame with observed and expected author productivity.
    """
    # Flatten list of authors across all papers
    all_authors = df[author_col].dropna().str.split(separator).explode()

    # Count number of publications per author
    author_counts = Counter(all_authors)

    # Count how many authors published n papers
    productivity = Counter(author_counts.values())

    # Convert to DataFrame
    lotka_df = pd.DataFrame(sorted(productivity.items()), columns=["n_pubs", "n_authors"])

    # Normalize and apply Lotka's expected law: expected ~ C / n^2
    C = lotka_df["n_authors"].iloc[0]  # authors with 1 publication
    lotka_df["expected_n_authors"] = C / (lotka_df["n_pubs"] ** 2)

    return lotka_df

def evaluate_lotka_fit(lotka_df):
    """
    Compute fit statistics comparing observed and expected values under Lotka's Law.
    
    Parameters:
        lotka_df (pd.DataFrame): Output of compute_lotka_distribution.
        
    Returns:
        dict: Dictionary with R2, RMSE, and KS statistic and p-value.
    """
    observed = lotka_df["n_authors"]
    expected = lotka_df["expected_n_authors"]

    r2 = r2_score(observed, expected)
    rmse = np.sqrt(mean_squared_error(observed, expected))
    ks_stat, ks_pvalue = ks_2samp(observed, expected)
    
    metrics = {
        "R2": r2,
        "RMSE": rmse,
        "KS_statistic": ks_stat,
        "KS_pvalue": ks_pvalue
    }
    return pd.DataFrame(list(metrics.items()),
                      columns=["Measure", "Value"])

def compute_bradford_distribution(df, source_col="Source title", zone_count=3, lowercase=False):
    """
    Compute Bradford's Law distribution by dividing sources into zones.
    """
    sources = df[source_col].dropna()
    if lowercase:
        sources = sources.str.lower()

    source_counts = sources.value_counts().reset_index()
    source_counts.columns = ["Source", "Document_Count"]

    source_counts["Cumulative_Documents"] = source_counts["Document_Count"].cumsum()
    source_counts["Cumulative_Percentage"] = source_counts["Cumulative_Documents"] / source_counts["Document_Count"].sum() * 100

    total_documents = source_counts["Document_Count"].sum()
    documents_per_zone = total_documents / zone_count
    zones = []
    current_zone = 1
    documents_in_current_zone = 0

    for count in source_counts["Document_Count"]:
        documents_in_current_zone += count
        zones.append(current_zone)
        if documents_in_current_zone >= documents_per_zone and current_zone < zone_count:
            current_zone += 1
            documents_in_current_zone = 0

    source_counts["Zone"] = zones

    return source_counts


def evaluate_bradford_fit(source_counts, zone_count=3):
    zone_stats = source_counts.groupby("Zone")["Document_Count"].sum()
    total_documents = source_counts["Document_Count"].sum()
    expected_per_zone = total_documents / zone_count
    deviations = (zone_stats - expected_per_zone).abs() / expected_per_zone

    return pd.DataFrame({"Documents per Zone": zone_stats.to_dict(),
        "Expected Documents per Zone": expected_per_zone,
        "Deviation per Zone": deviations.to_dict(),
        "Mean Deviation": deviations.mean()})

def compute_zipf_distribution_from_counts(df, word_col=0, count_col=1):
    """
    Compute Zipf's Law distribution given a DataFrame with word/item counts.
    """
    if isinstance(word_col, int):
        word_col = df.columns[word_col]
    if isinstance(count_col, int):
        count_col = df.columns[count_col]

    zipf_df = df[[word_col, count_col]].copy()
    zipf_df.columns = ["Word", "Frequency"]
    zipf_df = zipf_df.sort_values(by="Frequency", ascending=False).reset_index(drop=True)
    zipf_df["Rank"] = np.arange(1, len(zipf_df) + 1)
    return zipf_df

def evaluate_zipf_fit(zipf_df):
    """
    Evaluate Zipf's Law fit statistics.
    """
    log_rank = np.log(zipf_df["Rank"])
    log_freq = np.log(zipf_df["Frequency"])

    r2 = r2_score(log_freq, -log_rank)
    rmse = np.sqrt(mean_squared_error(log_freq, -log_rank))
    ks_stat, ks_pvalue = ks_2samp(log_freq, -log_rank)

    return {
        "R2": r2,
        "RMSE": rmse,
        "KS_statistic": ks_stat,
        "KS_pvalue": ks_pvalue
    }

def evaluate_prices_law(author_counts):
    """
    Evaluate Price's Law: check if square root of authors produces 50% of the documents.

    Parameters:
        author_counts (pd.Series): Series with authors as index and number of documents as values.

    Returns:
        dict: Core size, actual proportion produced by core, ideal square root size.
    """
    sorted_counts = author_counts.sort_values(ascending=False)
    total_docs = sorted_counts.sum()
    total_authors = len(sorted_counts)
    ideal_core_size = int(np.sqrt(total_authors))
    actual_core_docs = sorted_counts.iloc[:ideal_core_size].sum()
    actual_proportion = actual_core_docs / total_docs

    return {
        "Total Authors": total_authors,
        "Ideal Core Size (sqrt N)": ideal_core_size,
        "Core Documents": actual_core_docs,
        "Proportion from Core": actual_proportion
    }

def evaluate_pareto_principle(counts, top_percentage=20, outcome_percentage=80):
    """
    Evaluate Pareto Principle (default 80/20 rule).

    Parameters:
        counts (pd.Series): Series of items and their counts.
        top_percentage (float): Percentage of top items (default 20).
        outcome_percentage (float): Expected percentage of outcomes (default 80).

    Returns:
        dict: Top items needed, actual outcome proportion, comparison with expected.
    """
    sorted_counts = counts.sort_values(ascending=False)
    total = sorted_counts.sum()
    top_n = int(np.ceil(top_percentage / 100 * len(sorted_counts)))
    top_sum = sorted_counts.iloc[:top_n].sum()
    actual_percentage = top_sum / total * 100

    return {
        "Total Items": len(sorted_counts),
        "Top Items Needed": top_n,
        "Actual Outcome %": actual_percentage,
        "Expected Outcome %": outcome_percentage
    }

# Association group analysis



def compute_binary_associations(
    groups_df: pd.DataFrame,
    items_df: pd.DataFrame,
    association_measures: Optional[List[str]] = None,
    p_adjust_method: str = "fdr_bh",
    min_count: int = 1,
    min_jaccard: float = 0.0,
    significance_level: float = 0.05,
    treat_na_as_zero: bool = False,
    output_format: str = "long",
    filters: Optional[dict] = None
) -> pd.DataFrame:
    if association_measures is None:
        association_measures = [
            "Jaccard", "Sokal-Michener", "Dice", "Yule's Q", "Phi",
            "Odds Ratio", "Kulczynski", "Ochiai", "Cosine",
            "Conditional on Group (a/b)", "Conditional on Feature (a/c)",
            "Chi2 p", "Relative Risk", "Cramer's V"
        ]

    def calculate_measures(a, b, c, d):
        n = a + b + c + d
        measures = {}
        if "Jaccard" in association_measures:
            measures["Jaccard"] = a / (a + b + c) if (a + b + c) else np.nan
        if "Sokal-Michener" in association_measures:
            measures["Sokal-Michener"] = (a + d) / n if n else np.nan
        if "Dice" in association_measures:
            measures["Dice"] = 2 * a / (2 * a + b + c) if (2 * a + b + c) else np.nan
        if "Yule's Q" in association_measures:
            measures["Yule's Q"] = (a * d - b * c) / (a * d + b * c) if (a * d + b * c) else np.nan
        if "Phi" in association_measures:
            denom = np.sqrt((a + b) * (a + c) * (b + d) * (c + d))
            measures["Phi"] = (a * d - b * c) / denom if denom else np.nan
        if "Odds Ratio" in association_measures:
            measures["Odds Ratio"] = (a * d / (b * c)) if b * c != 0 else np.nan
        if "Kulczynski" in association_measures:
            measures["Kulczynski"] = 0.5 * (a / (a + b) + a / (a + c)) if (a + b) and (a + c) else np.nan
        if "Ochiai" in association_measures:
            measures["Ochiai"] = a / np.sqrt((a + b) * (a + c)) if (a + b) and (a + c) else np.nan
        if "Cosine" in association_measures:
            measures["Cosine"] = a / np.sqrt((a + b) * (a + c)) if (a + b) and (a + c) else np.nan
        if "Conditional on Group (a/b)" in association_measures:
            measures["Conditional on Group (a/b)"] = a / b if b else np.nan
        if "Conditional on Feature (a/c)" in association_measures:
            measures["Conditional on Feature (a/c)"] = a / c if c else np.nan
        if "Relative Risk" in association_measures:
            measures["Relative Risk"] = (a / (a + b)) / (c / (c + d)) if (a + b) and (c + d) else np.nan
        if "Cramer's V" in association_measures:
            chi2 = (a * d - b * c) ** 2 * n / ((a + b) * (c + d) * (a + c) * (b + d)) if n else np.nan
            measures["Cramer's V"] = np.sqrt(chi2 / n) if n else np.nan
        return measures

    results = []
    fisher_pvals = []

    for group in groups_df.columns:
        for item in items_df.columns:
            x = groups_df[group]
            y = items_df[item]

            pair = pd.concat([x, y], axis=1)
            if not treat_na_as_zero:
                pair = pair.dropna()
            else:
                pair = pair.fillna(0)

            x_vals = pair.iloc[:, 0].astype(int).values
            y_vals = pair.iloc[:, 1].astype(int).values

            a = np.sum((x_vals == 1) & (y_vals == 1))
            b = np.sum((x_vals == 1) & (y_vals == 0))
            c = np.sum((x_vals == 0) & (y_vals == 1))
            d = np.sum((x_vals == 0) & (y_vals == 0))

            if a < min_count:
                continue

            row = {
                "Group": group,
                "Item": item,
                "a": a, "b": b, "c": c, "d": d
            }

            row.update(calculate_measures(a, b, c, d))

            try:
                _, fisher_p = fisher_exact([[a, b], [c, d]])
            except:
                fisher_p = np.nan
            row["Fisher p"] = fisher_p
            fisher_pvals.append(fisher_p)

            try:
                _, chi2_p, _, _ = chi2_contingency([[a, b], [c, d]], correction=False)
            except:
                chi2_p = np.nan
            row["Chi2 p"] = chi2_p

            results.append(row)

    df = pd.DataFrame(results)

    if not df.empty:
        df["P-adj"] = np.nan
        mask = df["Fisher p"].notnull()
        adjusted = multipletests(df.loc[mask, "Fisher p"].values, method=p_adjust_method)
        df.loc[mask, "P-adj"] = adjusted[1]
        df["Significant"] = df["P-adj"] < significance_level

        if filters:
            for key, val in filters.items():
                if key in df.columns:
                    df = df[df[key] >= val]

        if "Jaccard" in df.columns:
            df = df[df["Jaccard"] >= min_jaccard]

        if output_format == "wide":
            df = df.pivot(index="Group", columns="Item")

    return df

# additional normalization methods for a given symetric dataframe

def normalize_symmetric_matrix(matrix_df, method="jaccard"):
    """
    Normalizes a symmetric co-occurrence matrix using the specified method.

    Supported methods:
    - "jaccard": Jaccard similarity
    - "cosine": Cosine similarity
    - "row": Row-wise proportion (normalized by row sums)
    - "none": No normalization, returns a copy

    Parameters:
    matrix_df (pd.DataFrame): Symmetric matrix (e.g. country collaboration).
    method (str): Normalization method ("jaccard", "cosine", "row", "none").

    Returns:
    pd.DataFrame: Normalized matrix with same index and columns.
    """
    if matrix_df.empty:
        return pd.DataFrame()

    if method == "none":
        return matrix_df.copy()

    if method == "row":
        row_sums = matrix_df.sum(axis=1).replace(0, 1)
        return matrix_df.div(row_sums, axis=0)

    if method == "cosine":
        
        similarity = cosine_similarity(matrix_df)
        return pd.DataFrame(similarity, index=matrix_df.index, columns=matrix_df.columns)

    if method == "jaccard":
        index = matrix_df.index
        row_sums = matrix_df.sum(axis=1)
        jaccard_df = pd.DataFrame(0.0, index=index, columns=index)

        for i in index:
            for j in index:
                if i == j:
                    jaccard_df.loc[i, j] = 1.0
                else:
                    numerator = matrix_df.loc[i, j]
                    denominator = row_sums[i] + row_sums[j] - numerator
                    jaccard_df.loc[i, j] = numerator / denominator if denominator > 0 else 0.0

        return jaccard_df


# Factor analysis




def build_document_term_matrix(
    df: pd.DataFrame,
    field: str = "Author Keywords",
    method: str = "count",
    min_df: int = 2,
    ngram_range: tuple = (1, 1),
    use_lemmatization: bool = False,
    pos_filter: list = None
) -> pd.DataFrame:
    """
    Build a document-term matrix with optional TF-IDF, n-grams, and lemmatization/POS filtering.
    """
    texts = df[field].fillna("")
    processed_texts = []
    for doc in texts:
        if use_lemmatization and nlp is not None:
            tokens = []
            for token in nlp(doc):
                if pos_filter and token.pos_ not in pos_filter:
                    continue
                tokens.append(token.lemma_)
            processed_texts.append(" ".join(tokens))
        else:
            processed_texts.append(doc)

    Vectorizer = TfidfVectorizer if method == "tfidf" else CountVectorizer
    vectorizer = Vectorizer(
        token_pattern=r"(?u)\b\w+\b",
        min_df=min_df,
        ngram_range=ngram_range
    )
    X = vectorizer.fit_transform(processed_texts)
    return pd.DataFrame(
        X.toarray(),
        index=df.index,
        columns=vectorizer.get_feature_names_out()
    )


def suggest_k(term_coords: np.ndarray, min_k: int = 2, max_k: int = 10) -> dict:
    """
    Suggest optimal number of clusters using silhouette scores for KMeans.

    Examples
    --------
    >>> result = conceptual_structure_analysis(df)
    >>> suggest_k(result['term_embeddings'], min_k=2, max_k=8)
    """
    scores = {}
    max_k = min(max_k, len(term_coords) - 1)
    for k in range(min_k, max_k + 1):
        labels = KMeans(n_clusters=k, random_state=0).fit_predict(term_coords)
        scores[k] = silhouette_score(term_coords, labels)
    return scores

def conceptual_structure_analysis(
    df: pd.DataFrame,
    field: str = "Author Keywords",
    dr_method: str = "MCA",
    cluster_method: str = "kmeans",
    n_clusters: int = 5,
    n_terms: int = 100,
    n_components: int = 2,
    dtm_method: str = "count",
    term_selection: str = "frequency",
    y: np.ndarray = None,
    min_df: int = 2,
    ngram_range: tuple = (1, 1),
    use_lemmatization: bool = False,
    pos_filter: list = None,
    include_terms: list = None,
    exclude_terms: list = None,
    term_regex: str = None,
    compute_metrics: bool = False
) -> dict:
    """
    Perform conceptual structure analysis with flexible DR, clustering, term selection, and advanced term filtering.

    Parameters
    ----------
    df : pd.DataFrame
        Source data.
    field : str, default='Author Keywords'
        Field to use for text analysis. If this field contains delimited keywords/phrases (like 'Author Keywords'),
        keywords are preserved as atomic terms (e.g., 'machine learning' is NOT split).
    dr_method : str, default='MCA'
        Dimensionality reduction method. Options: 'MCA', 'CA', 'MDS', 'PCA', 'LSA', 't-SNE', 'UMAP', 'NMF', 'LDA'.
    cluster_method : str, default='kmeans'
        Clustering method. Options: 'kmeans', 'agglomerative', 'dbscan', 'spectral', 'louvain'.
    n_clusters : int, default=5
        Number of clusters.
    n_terms : int, default=100
        Number of terms to select.
    n_components : int, default=2
        Number of dimensions for embeddings.
    dtm_method : str, default='count'
        Document-term matrix method ('count' or 'tfidf').
    term_selection : str, default='frequency'
        Term selection method ('frequency', 'chi2', 'mutual_info').
    y : np.ndarray, optional
        Target array for supervised term selection.
    min_df : int, default=2
        Minimum document frequency for term inclusion.
    ngram_range : tuple, default=(1, 1)
        N-gram range (ignored if keyword field).
    use_lemmatization : bool, default=False
        Whether to apply lemmatization (ignored if keyword field).
    pos_filter : list, optional
        POS tags to include (ignored if keyword field).
    include_terms : list of str, optional
        Specific terms to ensure are included.
    exclude_terms : list of str, optional
        Specific terms to remove from analysis.
    term_regex : str, optional
        Regex pattern to filter terms.
    compute_metrics : bool, default=False
        Whether to compute clustering metrics.

    Returns
    -------
    dict
        Dictionary with keys: 'term_embeddings', 'terms', 'term_labels', 'doc_embeddings', 'doc_labels', 'metrics', 'suggested_k'.

    Notes
    -----
    If using 'Author Keywords' or 'Index Keywords', terms are kept as atomic phrases and never split into words.
    """


    def build_document_term_matrix(df, field, method="count", min_df=2, ngram_range=(1,1), use_lemmatization=False, pos_filter=None):
        # If the field is likely to be keyword-type, split on ";" and treat as atomic
        if field.lower() in ["author keywords", "index keywords", "keywords"]:
            term_lists = df[field].fillna("").apply(lambda x: [t.strip() for t in str(x).split(";") if t.strip()])
            if method == "count":
                vec = CountVectorizer(
                    tokenizer=lambda x: x,
                    preprocessor=lambda x: x,
                    token_pattern=None,
                    min_df=min_df,
                    lowercase=False
                )
            elif method == "tfidf":
                vec = TfidfVectorizer(
                    tokenizer=lambda x: x,
                    preprocessor=lambda x: x,
                    token_pattern=None,
                    min_df=min_df,
                    lowercase=False
                )
            else:
                raise ValueError("Unknown DTM method: {}".format(method))
            X = vec.fit_transform(term_lists)
            columns = [t.strip() for t in vec.get_feature_names_out()]
            return pd.DataFrame(X.toarray(), columns=columns)
        else:
            if method == "count":
                vec = CountVectorizer(min_df=min_df, ngram_range=ngram_range)
            elif method == "tfidf":
                vec = TfidfVectorizer(min_df=min_df, ngram_range=ngram_range)
            else:
                raise ValueError("Unknown DTM method: {}".format(method))
            docs = df[field].fillna("").astype(str).values
            X = vec.fit_transform(docs)
            columns = vec.get_feature_names_out()
            return pd.DataFrame(X.toarray(), columns=columns)

    # Step 1: Build DTM
    dtm = build_document_term_matrix(
        df, field,
        method=dtm_method,
        min_df=min_df,
        ngram_range=ngram_range,
        use_lemmatization=use_lemmatization,
        pos_filter=pos_filter
    )
    # Step 2: Term selection
    if term_selection == "frequency":
        freqs = dtm.sum(axis=0).values
        top_idx = np.argsort(freqs)[::-1][:n_terms]
        selected_terms = list(dtm.columns[top_idx])
    elif term_selection in ("chi2", "mutual_info"):
        if y is None:
            raise ValueError("`y` must be provided for supervised term selection.")
        score_func = chi2 if term_selection == "chi2" else mutual_info_classif
        selector = SelectKBest(score_func=score_func, k=n_terms)
        selector.fit(dtm, y)
        selected_terms = list(dtm.columns[selector.get_support()])
    else:
        raise ValueError(f"Unknown term_selection: {term_selection}")

    # Advanced filtering
    if include_terms:
        for term in include_terms:
            if term in dtm.columns and term not in selected_terms:
                selected_terms.append(term)
    if exclude_terms:
        selected_terms = [t for t in selected_terms if t not in exclude_terms]
    if term_regex:
        pattern = re.compile(term_regex)
        selected_terms = [t for t in selected_terms if pattern.search(t)]
    selected_terms = selected_terms[:n_terms]
    dtm_top = dtm[selected_terms]

    # Step 3: Dimensionality reduction
    doc_coords = None
    metrics = {}

    if dr_method == "MCA":
        model = prince.MCA(n_components=n_components)
        model = model.fit(dtm_top)
        prince_terms = model.column_coordinates(dtm_top)
        term_coords = []
        final_terms = []
        for term in selected_terms:
            matches = [i for i in prince_terms.index if str(i).startswith(term)]
            if matches:
                row = prince_terms.loc[matches[0]].values
                term_coords.append(row)
                final_terms.append(term)
            else:
                term_coords.append(np.zeros(n_components))
                final_terms.append(term)
        term_coords = np.vstack(term_coords)
        selected_terms = final_terms
        doc_coords = model.row_coordinates(dtm_top).values
    elif dr_method == "CA":
        model = prince.CA(n_components=n_components)
        model = model.fit(dtm_top)
        prince_terms = model.column_coordinates(dtm_top)
        term_coords = []
        final_terms = []
        for term in selected_terms:
            matches = [i for i in prince_terms.index if str(i).startswith(term)]
            if matches:
                row = prince_terms.loc[matches[0]].values
                term_coords.append(row)
                final_terms.append(term)
            else:
                term_coords.append(np.zeros(n_components))
                final_terms.append(term)
        term_coords = np.vstack(term_coords)
        selected_terms = final_terms
        doc_coords = model.row_coordinates(dtm_top).values
    elif dr_method == "MDS":
        corr = np.corrcoef(dtm_top.T.values)
        dist = 1 - corr
        mds = MDS(n_components=n_components, dissimilarity="precomputed", random_state=0)
        term_coords = mds.fit_transform(dist)
        metrics['stress'] = getattr(mds, 'stress_', None)
    elif dr_method in ("PCA", "LSA"):
        svd = TruncatedSVD(n_components=n_components, random_state=0)
        doc_coords = svd.fit_transform(dtm_top)
        term_coords = svd.components_.T
    elif dr_method == "t-SNE":
        tsne = TSNE(n_components=n_components, random_state=0)
        term_coords = tsne.fit_transform(dtm_top.T.values)
        metrics['kl_divergence'] = getattr(tsne, 'kl_divergence_', None)
    elif dr_method == "UMAP":
        if umap is None:
            raise ImportError("UMAP library not installed")
        reducer = umap.UMAP(n_components=n_components, random_state=0)
        term_coords = reducer.fit_transform(dtm_top.T.values)
    elif dr_method == "NMF":
        nmf = NMF(n_components=n_components, random_state=0)
        doc_coords = nmf.fit_transform(dtm_top)
        term_coords = nmf.components_.T
    elif dr_method == "LDA":
        lda = LatentDirichletAllocation(n_components=n_components, random_state=0)
        doc_coords = lda.fit_transform(dtm_top)
        term_coords = lda.components_.T
    else:
        raise ValueError(f"Unknown DR method: {dr_method}")

    # Step 4: Clustering
    if cluster_method == "louvain":
        try:
            import community as community_louvain
        except ImportError:
            raise ImportError("python-louvain is required for louvain clustering")
        cooccur = (dtm_top.T.dot(dtm_top) > 0).astype(int)
        G = nx.from_pandas_adjacency(cooccur)
        partition = community_louvain.best_partition(G)
        term_labels = np.array([partition[i] for i in range(len(selected_terms))])
        metrics['inertia'] = None
    else:
        def get_clust(method, k):
            if method == "kmeans": return KMeans(n_clusters=k, random_state=0)
            if method == "agglomerative": return AgglomerativeClustering(n_clusters=k)
            if method == "dbscan": return DBSCAN()
            if method == "spectral": return SpectralClustering(n_clusters=k, assign_labels="discretize", random_state=0)
            raise ValueError(f"Unknown cluster method: {method}")
        clust = get_clust(cluster_method, n_clusters)
        term_labels = clust.fit_predict(term_coords)
        if hasattr(clust, 'inertia_'): metrics['inertia']=clust.inertia_

    # Step 5: Document clustering
    doc_labels = None
    if doc_coords is not None and cluster_method != "louvain":
        doc_labels = clust.fit_predict(doc_coords)

    # Step 6: Diagnostics
    if compute_metrics and len(set(term_labels)) > 1:
        metrics['silhouette'] = silhouette_score(term_coords, term_labels)
        metrics['davies_bouldin'] = davies_bouldin_score(term_coords, term_labels)

    return {
        "term_embeddings": term_coords,
        "terms": selected_terms,
        "term_labels": term_labels,
        "doc_embeddings": doc_coords,
        "doc_labels": doc_labels,
        "metrics": metrics,
        "suggested_k": suggest_k(term_coords)
    }




def words_by_cluster(
    term_embeddings: np.ndarray,
    terms: list,
    labels: np.ndarray
) -> pd.DataFrame:
    """
    Return a DataFrame with columns:
      - word: term string (original or fallback "cluster_<id>")
      - Dim1, Dim2: first two embedding components
      - cluster: cluster label

    Automatically fills missing terms if `terms` is shorter than embeddings.
    """
    emb = np.asarray(term_embeddings)
    # Ensure at least 2D
    if emb.ndim == 1:
        emb = np.vstack((emb, np.zeros_like(emb))).T

    n_pts = emb.shape[0]
    # Auto-fill missing term labels
    if len(terms) < n_pts:
        fallback = [f"cluster_{lbl}" for lbl in labels[len(terms):]]
        terms = list(terms) + fallback

    # Validate lengths
    if len(terms) != n_pts or len(labels) != n_pts:
        raise ValueError(
            f"words_by_cluster: embeddings ({n_pts}) must match len(terms) ({len(terms)}) and len(labels) ({len(labels)})"
        )

    return pd.DataFrame({
        'word':    terms,
        'Dim1':    emb[:, 0],
        'Dim2':    emb[:, 1],
        'cluster': labels
    })


def documents_per_cluster(
    df: pd.DataFrame,
    doc_embeddings: np.ndarray,
    doc_labels: np.ndarray,
    tc_field: str = 'Cited by'
) -> pd.DataFrame:
    """
    Return a DataFrame with columns:
      - Documents: from df['Title']
      - dim1, dim2: first two document embedding components
      - contrib: relative contribution (squared distance / total)
      - tc_field: values from df[tc_field] if available
      - Cluster: cluster label

    Parameters
    ----------
    tc_field : str
        Name of the column in df to use for citation counts (e.g. 'Cited by').

    Examples
    --------
    >>> result = conceptual_structure_analysis(df)
    >>> df_docs = documents_per_cluster(
    ...     df, result['doc_embeddings'], result['doc_labels'], tc_field='Cited by'
    >>> )
    """
    emb = np.asarray(doc_embeddings)
    # Ensure at least 2D
    if emb.ndim == 1:
        emb = np.vstack((emb, np.zeros_like(emb))).T
    # Compute contribution as squared distance / sum of all
    squared = emb[:, 0]**2 + emb[:, 1]**2
    contrib = squared / np.sum(squared) if np.sum(squared) != 0 else np.zeros_like(squared)
    # Lookup citation counts
    if tc_field in df.columns:
        tc_vals = df[tc_field].values
    else:
        tc_vals = np.full(len(df), np.nan)
    df_out = pd.DataFrame({
        'Documents': df['Title'].values,
        'dim1': emb[:, 0],
        'dim2': emb[:, 1],
        'contrib': contrib,
        tc_field: tc_vals,
        'Cluster': doc_labels
    }, index=df.index)
    return df_out

# clustering of the documents



def vectorize_text(
    texts: list[str],
    max_features: int = 1000,
    ngram_range: tuple[int, int] = (1, 2)
) -> csr_matrix:
    """
    Transform a list of text documents into a TF-IDF feature matrix.

    :param texts: List of text strings (e.g., titles, abstracts, keywords).
    :param max_features: Maximum number of features (vocabulary size).
    :param ngram_range: The lower and upper boundary of the n-grams to be extracted.
    :return: TF-IDF feature matrix (sparse).
    """
    vectorizer = TfidfVectorizer(max_features=max_features, ngram_range=ngram_range)
    return vectorizer.fit_transform(texts)

def find_optimal_k(
    X: csr_matrix,
    k_range: range = range(2, 11)
) -> int:
    """
    Determine the optimal number of clusters for KMeans using silhouette score.

    :param X: Feature matrix.
    :param k_range: Range of k values to search.
    :return: k value with highest silhouette score.
    """
    best_k = k_range.start
    best_score = -1
    for k in k_range:
        labels = KMeans(n_clusters=k, random_state=0).fit_predict(X)
        score = silhouette_score(X, labels)
        if score > best_score:
            best_k, best_score = k, score
    return best_k

def cluster_kmeans(
    X: csr_matrix,
    n_clusters: int
) -> np.ndarray:
    """
    Apply KMeans clustering to data.

    :param X: Feature matrix.
    :param n_clusters: Number of clusters.
    :return: Array of cluster labels.
    """
    model = KMeans(n_clusters=n_clusters, random_state=0)
    return model.fit_predict(X)

def cluster_hierarchical(
    X: csr_matrix,
    n_clusters: int,
    linkage: str = "ward"
) -> np.ndarray:
    """
    Apply agglomerative (hierarchical) clustering to data.

    :param X: Feature matrix.
    :param n_clusters: Number of clusters.
    :param linkage: Linkage criterion ("ward", "complete", "average", "single").
    :return: Array of cluster labels.
    """
    model = AgglomerativeClustering(n_clusters=n_clusters, linkage=linkage)
    return model.fit_predict(X.to_numpy())

def build_coupling_network(
    refs: list[list[str]]
) -> csr_matrix:
    """
    Build a bibliographic coupling graph and return adjacency matrix.

    :param refs: List where each element is the list of reference IDs for a document.
    :return: Sparse adjacency matrix of coupling weights.
    """
    # Create mapping from ref ID to docs
    ref_to_docs: dict[str, list[int]] = {}
    for doc_idx, doc_refs in enumerate(refs):
        for r in doc_refs:
            ref_to_docs.setdefault(r, []).append(doc_idx)
    # Accumulate coupling counts
    n = len(refs)
    row, col, data = [], [], []
    for docs in ref_to_docs.values():
        for i in range(len(docs)):
            for j in range(i + 1, len(docs)):
                row += [docs[i], docs[j]]
                col += [docs[j], docs[i]]
                data += [1, 1]
    return csr_matrix((data, (row, col)), shape=(n, n))

def cluster_by_coupling(
    coupling_mat: csr_matrix,
    n_clusters: int
) -> np.ndarray:
    """
    Cluster documents based on bibliographic coupling using KMeans on embedding of the coupling graph.

    :param coupling_mat: Adjacency matrix of coupling weights.
    :param n_clusters: Number of clusters.
    :return: Array of cluster labels.
    """

    eigenvalues, eigenvectors = spla.eigs(coupling_mat, k=n_clusters + 1, which="SR")
    L = eigenvectors.real[:, 1:]
    return KMeans(n_clusters=n_clusters, random_state=0).fit_predict(L)

def save_cluster_results(
    df: pd.DataFrame,
    labels: np.ndarray,
    prefix: str = ""
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Add cluster labels to df and create a binary membership matrix.

    :param df: Original DataFrame of documents.
    :param labels: 1D array of cluster labels per document.
    :param prefix: Optional prefix for new columns.
    :return: Tuple of (annotated_df, binary_df).
    """
    df_out = df.copy()
    col_name = f"{prefix}cluster_label"
    df_out[col_name] = labels
    # One-hot encode labels
    encoder = OneHotEncoder(sparse_output=False, categories="auto")
    onehot = encoder.fit_transform(labels.reshape(-1, 1))
    binary_df = pd.DataFrame(
        onehot,
        columns=[f"{prefix}cluster_{int(c)}" for c in encoder.categories_[0]]
    )
    return df_out, binary_df, col_name

def cluster_documents(
    df: pd.DataFrame,
    text_field: str = "Abstract",
    method: str = "kmeans",
    n_clusters: int | None = None,
    k_range: range = range(2, 11),
    coupling_fields: list[str] | str | None = None,
    **vectorize_kwargs
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    General interface to cluster documents in df using specified method.

    :param df: DataFrame containing documents.
    :param text_field: Name of column to vectorize (e.g., keywords, title, abstract).
    :param method: One of {"kmeans", "hierarchical", "coupling"}.
    :param n_clusters: Number of clusters; if None, will optimize (only for kmeans) or default for coupling.
    :param k_range: Range of k values to try when optimizing (only for kmeans).
    :param coupling_fields: Field name or list of field names for coupling clustering; each may contain list-like or semicolon-delimited strings.
    :param vectorize_kwargs: Passed to vectorize_text.
    :return: Tuple of (df with cluster labels, cluster representation matrix DataFrame).
    """
    # Helper to normalize coupling entries to a list
    def _to_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, str):
            # split on semicolon and strip whitespace
            return [i.strip() for i in x.split(';') if i.strip()]
        return []

    if method in ("kmeans", "hierarchical"):
        X = vectorize_text(df[text_field].fillna(""), **vectorize_kwargs)
        if method == "kmeans":
            if n_clusters is None:
                n_clusters = find_optimal_k(X, k_range)
            labels = cluster_kmeans(X, n_clusters)
        else:
            if n_clusters is None:
                raise ValueError("n_clusters must be specified for hierarchical clustering")
            labels = cluster_hierarchical(X, n_clusters)
    elif method == "coupling":
        # Determine which fields to use for coupling
        fields = (
            [coupling_fields]
            if isinstance(coupling_fields, str)
            else coupling_fields or ["References"]
        )
        # Validate presence of columns
        missing = [f for f in fields if f not in df.columns]
        if missing:
            raise ValueError(f"DataFrame must include {missing} column(s) for coupling")
        # Build and merge coupling networks
        networks = []
        for field in fields:
            items = df[field].apply(_to_list).tolist()
            networks.append(build_coupling_network(items))
        C = networks[0]
        for net in networks[1:]:
            C = C + net
        # Default cluster count for coupling if not provided
        if n_clusters is None:
            n_clusters = len(df) // 10 or 2
        labels = cluster_by_coupling(C, n_clusters)
    else:
        raise ValueError(f"Unknown method: {method}")

    return save_cluster_results(df, labels, prefix=f"{method}_")



# spectroscopy

def extract_cited_years(reference_text, include_pre_1900=False):
    """
    Extracts all 4-digit years from a reference string. Includes 17xx and 18xx if enabled.

    Args:
        reference_text (str): String containing reference text.
        include_pre_1900 (bool): If True, includes years before 1900.

    Returns:
        list of str: List of extracted years.
    """
    if include_pre_1900:
        pattern = r"\b(17\d{2}|18\d{2}|19\d{2}|20\d{2})\b"
    else:
        pattern = r"\b(19\d{2}|20\d{2})\b"
    return re.findall(pattern, reference_text or "")

def compute_reference_spectrogram(df, reference_column="References", include_pre_1900=False):
    """
    Computes the distribution of cited years from a DataFrame column containing references.

    Args:
        df (pd.DataFrame): DataFrame with a column of reference text.
        reference_column (str): Name of the column with references.
        include_pre_1900 (bool): Whether to include years before 1900.

    Returns:
        pd.DataFrame: DataFrame with cited years and their citation counts.
    """
    current_year = datetime.datetime.now().year
    all_years = df[reference_column].fillna("").apply(lambda x: extract_cited_years(x, include_pre_1900))
    all_years_flat = list(chain.from_iterable(all_years))
    all_years_flat = [int(y) for y in all_years_flat if int(y) <= current_year]
    year_counts = Counter(all_years_flat)
    result_df = pd.DataFrame.from_dict(year_counts, orient="index", columns=["Citations"])
    result_df.index.name = "Cited Year"
    result_df = result_df.sort_index()
    return result_df



def compute_reference_correlation(df, reference_column="References", year_column="Year", include_pre_1900=False):
    """
    Computes correlation between document year and:
    1. Mean of reference years per document
    2. Each individual reference year occurrence

    Returns:
        dict: Correlation results and associated data
    """
    current_year = datetime.datetime.now().year
    ref_years = df[reference_column].fillna("").apply(lambda x: extract_cited_years(x, include_pre_1900))
    doc_years = df[year_column].values

    mean_refs = []
    repeated_pairs = []
    for doc_year, years in zip(doc_years, ref_years):
        ref_years_filtered = [int(y) for y in years if int(y) <= current_year and int(y) <= doc_year]
        if ref_years_filtered:
            mean_ref = np.mean(ref_years_filtered)
            mean_refs.append((doc_year, mean_ref))
            repeated_pairs.extend([(doc_year, y) for y in ref_years_filtered])

    mean_refs_df = pd.DataFrame(mean_refs, columns=["Document Year", "Mean Reference Year"])
    repeated_df = pd.DataFrame(repeated_pairs, columns=["Document Year", "Reference Year"])

    mean_corr = pearsonr(mean_refs_df["Document Year"], mean_refs_df["Mean Reference Year"])
    ref_corr = pearsonr(repeated_df["Document Year"], repeated_df["Reference Year"])

    return {
        "mean_reference_df": mean_refs_df,
        "repeated_year_df": repeated_df,
        "mean_reference_corr": mean_corr,
        "reference_year_corr": ref_corr
    }


# scientific production groups

def get_scientific_production_by_group(
    df: pd.DataFrame,
    group_matrix: pd.DataFrame | np.ndarray,
    group_names: list[str] = None,
    output_format: str = "wide",
    **production_kwargs
) -> pd.DataFrame:
    """
    Compute annual scientific production stats for each group and merge into a single DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain columns "Year" and "Cited by".
    group_matrix : pd.DataFrame or 2D array of shape (n_docs, n_groups)
        If DataFrame, its index must match df.index and its columns are group names.
    group_names : list of str, optional
        If group_matrix is an array, names of groups (in column order).
    output_format : str, default "wide"
        "wide" for a pivoted DataFrame with Year as a column and metrics grouped by group,
        "long" for a tidy DataFrame with columns [Group, Year, ...metrics...].
    **production_kwargs
        Passed to get_scientific_production (relative_counts, cumulative, etc.).

    Returns
    -------
    pd.DataFrame
        Merged production statistics in the specified format.
    """
    # Make a deep copy of df to avoid SettingWithCopy warnings
    df = df.copy()
    # Use .loc to safely assign
    df.loc[:, "Year"] = df.loc[:, "Year"].astype(int)

    # Prepare group DataFrame
    if not isinstance(group_matrix, pd.DataFrame):
        if group_names is None:
            raise ValueError("Please supply group_names when group_matrix is an array")
        group_df = pd.DataFrame(group_matrix, index=df.index, columns=group_names)
    else:
        group_df = group_matrix.copy()

    # Compute per-group stats, index by Year
    results = {}
    for grp in group_df.columns:
        mask = group_df[grp] == 1
        subset = df.loc[mask].copy()
        prod = (
            get_scientific_production(subset, **production_kwargs)
            if not subset.empty
            else get_scientific_production(df.iloc[0:0], **production_kwargs)
        )
        prod = prod.set_index("Year")
        results[grp] = prod

    # Concatenate into MultiIndex: [Group, Year]
    merged = pd.concat(results, names=["Group", "Year"])

    if output_format == "wide":
        # Pivot groups into wide format
        wide = merged.unstack(level="Group")
        # Flatten column index and adjust metric names
        cols = []
        for metric, grp in wide.columns:
            m = metric.replace("Number of Documents", "Number of documents")
            cols.append(f"{m} {grp}")
        wide.columns = cols
        # fill missing years/groups with zeros
        wide = wide.fillna(0)
        # Flatten Year index into a column
        wide = wide.reset_index()
        return wide

    elif output_format == "long":
        # Reset index; missing group-year combos may be omitted
        long = merged.reset_index()
        # adjust number of documents label
        long = long.rename(columns={"Number of Documents": "Number of documents"})
        # fill numeric NaNs with zeros
        num_cols = long.select_dtypes(include=[np.number]).columns
        long.loc[:, num_cols] = long.loc[:, num_cols].fillna(0)
        return long

    else:
        raise ValueError("output_format must be 'long' or 'wide'")

# group performance

def group_entity_stats(df, group_matrix, entity_col, entity_label,
                       items_of_interest=None, exclude_items=None, top_n=20,
                       counts_df=None, count_method=None,
                       regex_include=None, regex_exclude=None,
                       value_type="string", indicators=False,
                       missing_as_zero=False, mode="full",
                       output_format="wide"):
    if output_format not in ("wide", "long", "pivot", "matrix"):  # validate
        raise ValueError(f"output_format must be 'wide', 'long', 'pivot', or 'matrix', got {output_format}")

    if not hasattr(group_matrix, "columns"):
        group_matrix = pd.DataFrame(group_matrix, index=df.index,
                                    columns=[f"group_{i}" for i in range(group_matrix.shape[1])])

    all_stats = []
    indicator_dfs = {} if indicators else None

    for grp in group_matrix.columns:
        mask = group_matrix[grp] == 1
        df_grp = df.loc[mask].reset_index(drop=True)
        stats, ind = get_entity_stats(
            df_grp, entity_col, entity_label,
            items_of_interest=items_of_interest,
            exclude_items=exclude_items,
            top_n=top_n,
            counts_df=counts_df,
            count_method=count_method,
            regex_include=regex_include,
            regex_exclude=regex_exclude,
            value_type=value_type,
            indicators=indicators,
            missing_as_zero=missing_as_zero,
            mode=mode
        )
        stats = stats.copy()
        stats['group'] = grp
        stats = stats.set_index(['group', entity_label])
        all_stats.append(stats)
        if indicators:
            indicator_dfs[grp] = ind

    stats_df = pd.concat(all_stats, axis=0)

    if output_format == "long":
        stats_df = stats_df.reset_index().melt(
            id_vars=["group", entity_label], var_name="metric", value_name="value"
        )
    elif output_format == "pivot":
        pivot = stats_df.unstack(level=0)
        pivot.columns = [f"{metric} {grp}" for metric, grp in pivot.columns]
        stats_df = pivot.reset_index()
    elif output_format == "matrix":
        matrix = stats_df.unstack(level=0)
        matrix.columns.set_names(["metric", "group"], inplace=True)
        matrix = matrix.reset_index()
        first_col = matrix.columns[0]
        if first_col == "":
            matrix = matrix.rename(columns={first_col: entity_label})
        stats_df = matrix

    return stats_df, indicator_dfs


# comparison of two frequency distributions (global, local)

def compare_counts(series_full: pd.Series, series_subset: pd.Series, top_n: int = None) -> pd.DataFrame:
    """
    Compare item counts between a reference (full) and a subset dataset.

    Parameters
    ----------
    series_full : pd.Series
        Item counts from the full dataset.
    series_subset : pd.Series
        Item counts from the subset dataset.
    top_n : int, optional
        If specified, return only the top_n items by absolute percentage point difference.

    Returns
    -------
    pd.DataFrame
        DataFrame with counts, proportions, percentage point difference, fold-change, and flag.
    """
    df = pd.DataFrame({
        "Count_Full": series_full,
        "Count_Sub": series_subset
    }).fillna(0)

    df["Prop_Full"] = df["Count_Full"] / df["Count_Full"].sum()
    df["Prop_Sub"] = df["Count_Sub"] / df["Count_Sub"].sum()
    df["PP_Diff"] = 100 * (df["Prop_Sub"] - df["Prop_Full"])

    with np.errstate(divide="ignore", invalid="ignore"):
        df["Rel_Diff"] = np.where(df["Prop_Full"] > 0, df["Prop_Sub"] / df["Prop_Full"], np.nan)

    df["Interesting"] = (df["PP_Diff"].abs() > 2) | (df["Rel_Diff"] > 2) | (df["Rel_Diff"] < 0.5)

    if top_n is not None:
        df = df.reindex(df["PP_Diff"].abs().sort_values(ascending=False).index).head(top_n)

    return df

# LLM

def invoke_llm(
    prompt: str,
    model: str = "sshleifer/distilbart-cnn-12-6",
    provider: str = "bert",
    hf_token: Optional[str] = None
) -> str:
    """
    Invoke a language model to generate a completion based on the given prompt.

    Args:
        prompt (str): The prompt to send to the model.
        model (str): Model identifier or alias (e.g., "gpt-3.5-turbo", "sshleifer/distilbart-cnn-12-6", "meta-llama/Llama-2-7b-chat-hf").
        provider (str): Which provider to use: "openai", "huggingface", "local", "bert", or "llama".
        hf_token (Optional[str]): Hugging Face API token.

    Returns:
        str: The generated text.
    """
    if provider == "openai":

        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY environment variable not set")
        client = openai.OpenAI(api_key=api_key)
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}]
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            msg = str(e)
            if 'insufficient_quota' in msg or getattr(e, 'code', None) == 'insufficient_quota':
                raise RuntimeError("OpenAI insufficient quota. Please check your plan and billing.")
            raise

    elif provider == "huggingface":
        from huggingface_hub import InferenceApi
        if hf_token is None:
            hf_token = os.getenv("HUGGINGFACE_API_TOKEN")
        api = InferenceApi(repo_id=model, token=hf_token)
        output = api(inputs=prompt)
        if isinstance(output, dict) and "generated_text" in output:
            return output["generated_text"].strip()
        if isinstance(output, str):
            return output.strip()
        return str(output)

    elif provider == "local":
        from transformers import pipeline
        nlp = pipeline("text-generation", model=model)
        result = nlp(prompt, max_length=512, num_return_sequences=1)
        return result[0]["generated_text"].strip()

    elif provider == "bert":
        # Use a Hugging Face summarization model (BART-based) by default
        from transformers import pipeline
        summarizer = pipeline("summarization", model=model)
        result = summarizer(prompt, max_length=512, min_length=30, do_sample=False)
        return result[0].get("summary_text", "").strip()

    elif provider == "llama":
        from transformers import pipeline
        generator = pipeline("text-generation", model=model, device_map="auto")
        result = generator(prompt, max_length=512, num_return_sequences=1)
        return result[0].get("generated_text", "").strip()

    else:
        raise ValueError(f"Unsupported provider: {provider}")

def llm_summarize_abstracts(
    abstracts: List[str],
    llm_fn: Callable[[str], str] = invoke_llm,
    model: str = "sshleifer/distilbart-cnn-12-6",
    provider: str = "bert",
    prompt_template: str = (
        "Please provide a concise summary (3-5 sentences) of the following abstracts:\n"
        "{abstracts}\n"
    ),
    hf_token: Optional[str] = None
) -> str:
    """
    Summarize a list of abstracts into a short, coherent summary using an LLM.

    Args:
        abstracts (List[str]): A list of abstract texts to summarize.
        llm_fn (Callable): Function to call the LLM (defaults to invoke_llm).
        model (str): Model identifier.
        provider (str): Provider to use.
        prompt_template (str): Template for the prompt.
        hf_token (Optional[str]): Hugging Face API token.

    Returns:
        str: The summary generated by the LLM.
    """
    joined = "\n---\n".join(abstracts)
    prompt = prompt_template.format(abstracts=joined)
    return llm_fn(prompt=prompt, model=model, provider=provider, hf_token=hf_token)

def llm_describe_table(
    table: Any,
    llm_fn: Callable[[str], str] = invoke_llm,
    model: str = "sshleifer/distilbart-cnn-12-6",
    provider: str = "bert",
    prompt_template: str = (
        "You are given the following dataframe."
        "Please describe its main information and performance highlights in one paragraph:\n{table_md}\n"
    ),
    hf_token: Optional[str] = None
) -> str:
    """
    Generate a description of a table's main information and performance highlights using an LLM.

    Args:
        table (Any): A table-like object (e.g., pandas.DataFrame or markdown string).
        llm_fn (Callable): Function to call the LLM (defaults to invoke_llm).
        model (str): Model identifier.
        provider (str): Provider to use.
        prompt_template (str): Template for the prompt.
        hf_token (Optional[str]): Hugging Face API token.

    Returns:
        str: The table description generated by the LLM.
    """
    try:

        if isinstance(table, pd.DataFrame):
            table_md = table.to_markdown(index=False)
        else:
            table_md = str(table)
    except ImportError:
        table_md = str(table)
    prompt = prompt_template.format(table_md=table_md)
    return llm_fn(prompt=prompt, model=model, provider=provider, hf_token=hf_token)


# Citation network of documents



def build_citation_network(
    df: pd.DataFrame,
    threshold: int = 90,
    largest_only: bool = True
) -> tuple[nx.DiGraph, dict[str, list[str]]]:
    """
    Build a citation network from document titles and reference strings.
    Normalization is internal.

    Args:
        df: DataFrame with "Doc ID", "Title", and "References" columns.
        threshold: Fuzzy-match score threshold (0-100) for linking.
        largest_only: If True, return only the largest connected component.

    Returns:
        G: Directed graph with Doc ID labels.
        unmatched: Dict mapping Doc ID to list of unmatched reference strings.
    """
    def normalize(text: str) -> str:
        text = text.lower()
        text = re.sub(r"[\W_]+", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    titles = df["Title"].tolist()
    norm_titles = [normalize(t) for t in titles]

    G = nx.DiGraph()
    G.add_nodes_from(range(len(titles)))
    unmatched: dict[int, list[str]] = {}

    for idx, refs in df["References"].items():
        if not isinstance(refs, str) or pd.isna(refs):
            continue
        ref_list = [r.strip() for r in refs.split(";") if r.strip()]
        local_unmatched: list[str] = []
        for ref in ref_list:
            norm_ref = normalize(ref)
            matched = False
            if norm_ref in norm_titles:
                tgt = norm_titles.index(norm_ref)
                G.add_edge(idx, tgt)
                matched = True
            else:
                best_score, best_tgt = 0, None
                for j, nt in enumerate(norm_titles):
                    score = fuzz.partial_ratio(norm_ref, nt)
                    if score > best_score:
                        best_score, best_tgt = score, j
                if best_score >= threshold and best_tgt is not None:
                    G.add_edge(idx, best_tgt)
                    matched = True
            if not matched:
                local_unmatched.append(ref)
        if local_unmatched:
            unmatched[idx] = local_unmatched

    # Prune self-loops and singletons
    G.remove_edges_from(nx.selfloop_edges(G))
    singletons = [n for n in G.nodes() if G.degree(n) == 0]
    G.remove_nodes_from(singletons)

    # Optionally keep largest component
    if largest_only and G.number_of_nodes() > 0:
        comps = list(nx.weakly_connected_components(G))
        largest = max(comps, key=len)
        G = G.subgraph(largest).copy()

    # Relabel nodes to Doc ID and filter unmatched
    mapping = {i: df.loc[i, "Doc ID"] for i in G.nodes()}
    G = nx.relabel_nodes(G, mapping)
    unmatched = {mapping[i]: refs for i, refs in unmatched.items() if i in mapping}

    return G, unmatched

def compute_main_path(G: nx.DiGraph) -> list[str]:
    """
    Compute the main citation path. Condenses cycles to DAG.

    Args:
        G: Directed graph with Doc ID labels.

    Returns:
        List of Doc ID labels on the critical path.
    """
    cG = nx.condensation(G)
    comp_path = nx.dag_longest_path(cG)
    main_path = []
    for comp in comp_path:
        members = sorted(cG.nodes[comp]["members"])
        main_path.append(members[0])
    return main_path

# historiograph
from rapidfuzz import fuzz
import difflib

def extract_reference_titles(ref_string): # this could be replaced by other functions
    """Extract potential titles from raw reference strings using pattern-based heuristics."""
    if not isinstance(ref_string, str):
        return []
    refs = [s.strip() for s in ref_string.split(";") if s.strip()]
    cleaned = []
    for ref in refs:
        # Remove author names and years: keep the middle (likely title) segment
        parts = [part.strip() for part in ref.split(",")]
        # Heuristic: remove authors, year, source; keep middle part
        if len(parts) >= 3:
            candidate = parts[2]  # usually the 3rd component has the title or journal name
        else:
            candidate = parts[0]
        cleaned.append(candidate)
    return cleaned

def approximate_match(ref_title, titles, cutoff=0.85):
    """Find the closest match to a reference title among known titles."""
    if not isinstance(ref_title, str):
        return None
    matches = difflib.get_close_matches(ref_title.lower(), titles, n=1, cutoff=cutoff)
    return matches[0] if matches else None


def build_historiograph(
    df,
    title_col="Title",
    year_col="Year",
    refs_col="References",
    cutoff=0.85,
    label_col="Document Short Label",
    weight_col="Cited by",
    save_path=None,
):
    """Construct a directed citation graph using approximate title matching and optional node relabeling."""
    G = nx.DiGraph()

    title_map = {title.lower(): title for title in df[title_col].dropna().unique()}

    label_map = {}
    for _, row in df.iterrows():
        title = row.get(title_col)
        label = row.get(label_col) if label_col and pd.notna(row.get(label_col)) else title
        weight = row.get(weight_col, 0) if weight_col in row else 0
        if pd.notna(title) and pd.notna(row.get(year_col)):
            label_map[title.lower()] = label
            G.add_node(label, year=row[year_col], title=label, **{weight_col: weight})

    for _, row in df.iterrows():
        citing_title = row.get(title_col)
        citing_label = label_map.get(citing_title.lower()) if citing_title else None
        if not citing_label:
            continue

        cited_titles = extract_reference_titles(row.get(refs_col))

        for ref_title in cited_titles:
            match_key = approximate_match(ref_title, list(title_map.keys()), cutoff=cutoff)
            if match_key:
                matched_title = title_map[match_key]
                cited_label = label_map.get(matched_title.lower())
                if cited_label and cited_label != citing_label:
                    G.add_edge(cited_label, citing_label)

    if save_path:
        save_network(G, save_path)

    return G

# Network analysis functions



def louvain_partition(G, resolution=1.0, randomize=False, random_state=None):
    """
    Detect communities using the Louvain method.

    Parameters
    ----------
    G : networkx.Graph
    resolution : float, optional
    randomize : bool, optional
    random_state : int or RandomState, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if community_louvain is None:
        raise ImportError("python-louvain package is required for Louvain partition")
    return community_louvain.best_partition(
        G, resolution=resolution, randomize=randomize, random_state=random_state
    )

def greedy_modularity_partition(G, weight="weight"):
    """
    Detect communities by greedy modularity maximization.

    Parameters
    ----------
    G : networkx.Graph
    weight : str or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.greedy_modularity_communities(G, weight=weight)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def label_propagation_partition(G, weight=None, seed=None):
    """
    Detect communities via asynchronous label propagation.

    Parameters
    ----------
    G : networkx.Graph
    weight : str or None, optional
    seed : int or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.asyn_lpa_communities(G, weight=weight, seed=seed)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def girvan_newman_partition(G, n_communities=2):
    """
    Detect communities using the Girvan-Newman algorithm.

    Parameters
    ----------
    G : networkx.Graph
    n_communities : int, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if n_communities < 2:
        raise ValueError("n_communities must be >= 2")
    comp_gen = nx.algorithms.community.girvan_newman(G)
    for _ in range(n_communities - 2):
        next(comp_gen)
    communities = next(comp_gen)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def k_clique_partition(G, k=3):
    """
    Detect communities using k-clique percolation.

    Parameters
    ----------
    G : networkx.Graph
    k : int, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.k_clique_communities(G, k)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def kernighan_lin_partition(G, max_iter=10):
    """
    Bisect graph using Kernighan-Lin algorithm.

    Parameters
    ----------
    G : networkx.Graph
    max_iter : int, optional

    Returns
    -------
    dict
        Node-to-partition mapping {0,1}.
    """
    sets = nx.algorithms.community.kernighan_lin_bisection(G, max_iter=max_iter)
    partition = {node: 0 for node in sets[0]}
    partition.update({node: 1 for node in sets[1]})
    return partition

def edge_betweenness_partition(G, n_communities=2):
    """
    Alias for Girvan-Newman via edge betweenness.
    """
    return girvan_newman_partition(G, n_communities)

def walktrap_partition(G, steps=4, weights=None):
    """
    Detect communities using the Walktrap algorithm via igraph.

    Parameters
    ----------
    G : networkx.Graph
    steps : int, optional
    weights : str or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    g = ig.Graph.TupleList(G.edges(data=bool(weights)), directed=False,
                            edge_attrs=[weights] if weights else [])
    wc = g.community_walktrap(weights=weights, steps=steps).as_clustering()
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def infomap_partition(G, edge_weights="weight"):
    """
    Detect communities using the Infomap algorithm.
    """
    g = ig.Graph.TupleList(G.edges(data=edge_weights), directed=False,
                            edge_attrs=[edge_weights])
    wc = g.community_infomap(edge_weights)
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def leading_eigenvector_partition(G):
    """
    Detect communities using leading eigenvector method.
    """
    g = ig.Graph.TupleList(G.edges(), directed=False)
    wc = g.community_leading_eigenvector()
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def leiden_partition(G, resolution_parameter=1.0):
    """
    Detect communities using the Leiden algorithm.
    """
    g = ig.Graph.TupleList(G.edges(), directed=False)
    wc = g.community_leiden(resolution_parameter=resolution_parameter)
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def spinglass_partition(G, weights=None, start_temp=1.0, stop_temp=0.01, cool_fact=0.99, spins=2):
    """
    Detect communities using the Spinglass algorithm via igraph.

    Parameters
    ----------
    G : networkx.Graph
        The graph to partition.
    weights : str or None, optional
        Edge attribute name for weights.
    start_temp : float, optional
        Starting temperature for the spin glass model.
    stop_temp : float, optional
        Stopping temperature for the model.
    cool_fact : float, optional
        Cooling factor between iterations.
    spins : int, optional
        Number of spins (must be >= 2).

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if spins < 2:
        spins = 2
    g = ig.Graph.TupleList(G.edges(data=bool(weights)), directed=False,
                            edge_attrs=[weights] if weights else [])
    wc = g.community_spinglass(weights=weights, start_temp=start_temp,
                                stop_temp=stop_temp, cool_fact=cool_fact,
                                spins=spins)
    return {v['name']: wc.membership[v.index] for v in g.vs}

def add_partitions(
    G,
    louvain_kwargs=None,
    greedy_kwargs=None,
    label_kwargs=None,
    girvan_kwargs=None,
    k_clique_kwargs=None,
    kernighan_kwargs=None,
    walktrap_kwargs=None,
    edge_betweenness_kwargs=None,
    infomap_kwargs=None,
    leading_kwargs=None,
    leiden_kwargs=None,
    spinglass_kwargs=None,
):
    """
    Compute and attach multiple community partitions to G.

    Returns
    -------
    dict
        Mapping method names to partition dicts.
    """
    configs = {
        "walktrap": (walktrap_partition, walktrap_kwargs or {}),
        "edge_betweenness": (edge_betweenness_partition, edge_betweenness_kwargs or {}),
        "infomap": (infomap_partition, infomap_kwargs or {}),
        "leading_eigenvector": (leading_eigenvector_partition, leading_kwargs or {}),
        "leiden": (leiden_partition, leiden_kwargs or {}),
        "spinglass": (spinglass_partition, spinglass_kwargs or {}),
        "louvain": (louvain_partition, louvain_kwargs or {}),
        "greedy_modularity": (greedy_modularity_partition, greedy_kwargs or {}),
        "label_propagation": (label_propagation_partition, label_kwargs or {}),
        "girvan_newman": (girvan_newman_partition, girvan_kwargs or {}),
        "k_clique": (k_clique_partition, k_clique_kwargs or {}),
        "kernighan_lin": (kernighan_lin_partition, kernighan_kwargs or {}),
    }
    results = {}
    for name, (func, kwargs) in configs.items():
        try:
            partition = func(G, **kwargs)
            results[name] = partition
            for node, cid in partition.items():
                G.nodes[node][f"partition_{name}"] = cid
        except:
            pass
    return results


def add_vectors_from_dataframe(G, df, node_col, vector_cols):
    """
    Add vector attributes to nodes from a DataFrame.

    Parameters
    ----------
    G : networkx.Graph
    df : pandas.DataFrame
    node_col : str
        Column name for node IDs.
    vector_cols : list of str
        DataFrame columns to use as node attributes.
    """
    for _, row in df.iterrows():
        node = row[node_col]
        if node in G:
            for col in vector_cols:
                G.nodes[node][col] = row[col]
    return G

def nodes_to_dataframe(G):
    """
    Export all node attributes to a pandas DataFrame, including partitions and vector values.

    Parameters
    ----------
    G : networkx.Graph
        The graph whose node attributes are to be exported.

    Returns
    -------
    pandas.DataFrame
        DataFrame with one row per node and one column per node attribute.
    """
    nodes = list(G.nodes(data=True))
    keys = set()
    for _, attrs in nodes:
        keys.update(attrs.keys())
    columns = ["node"] + sorted(keys)
    records = []
    for n, attrs in nodes:
        rec = {"node": n}
        for k in keys:
            rec[k] = attrs.get(k)
        records.append(rec)
    return pd.DataFrame(records, columns=columns)


def save_network(G, filename, formats=None, vector_cols=None, partition_attr="partition"):
    """
    Save a NetworkX graph in various formats, including Pajek (.net, .clu, .vec), GraphML, and GEXF.

    Args:
        G (networkx.Graph): The graph to save.
        filename (str): Path/prefix for saving files (no extension).
        formats (list of str, optional): Formats to save ("pajek", "graphml", "gexf"). Defaults to all.
        vector_cols (list of str, optional): Node attributes to export as .vec files.
        partition_attr (str, optional): Node attribute to use for .clu partition file.
    """

    if formats is None:
        formats = ["pajek", "graphml", "gexf"]

    if "pajek" in formats:
        nx.write_pajek(G, f"{filename}.net")
        # Partition .clu

        if any(any(k.startswith(partition_attr) for k in d) for _, d in G.nodes(data=True)):
            partition_attrs = set()
            for _, d in G.nodes(data=True):
                partition_attrs.update(k for k in d if k.startswith(partition_attr))
            
            # Save each partition to its own .clu file
            for attr in sorted(partition_attrs):  # sorted for consistent order
                with open(f"{filename}_{attr}.clu", "w") as f:
                    f.write(f"*Vertices {G.number_of_nodes()}\n")
                    for n in G.nodes():
                        f.write(f"{G.nodes[n].get(attr, 1)}\n")
        # Vectors .vec
        if vector_cols:
            for vec_col in vector_cols:
                with open(f"{filename}_{vec_col}.vec", "w") as f:
                    f.write(f"*Vertices {G.number_of_nodes()}\n")
                    for n in G.nodes():
                        f.write(f"{G.nodes[n].get(vec_col, 0)}\n")
    if "graphml" in formats:
        nx.write_graphml(G, f"{filename}.graphml")
    if "gexf" in formats:
        nx.write_gexf(G, f"{filename}.gexf")



def save_to_pajek(G, basename, partition_attrs=None, vector_attrs=None):
    """
    Export graph and node attributes to Pajek files.

    Parameters
    ----------
    G : networkx.Graph
    basename : str
        Base filename for .net, .clu, .vec files.
    partition_attrs : str or list of str, optional
    vector_attrs : list of str, optional
    """
    nx.write_pajek(G, f"{basename}.net")
    nodes = list(G.nodes())
    if partition_attrs:
        attrs = [partition_attrs] if isinstance(partition_attrs, str) else partition_attrs
        for attr in attrs:
            fname = f"{basename}.clu" if len(attrs) == 1 else f"{basename}_{attr}.clu"
            with open(fname, "w") as f:
                f.write(f"*Vertices {len(nodes)}\n")
                for n in nodes:
                    f.write(str(G.nodes[n].get(attr, 0)) + "\n")
    if vector_attrs:
        with open(f"{basename}.vec", "w") as f:
            f.write(f"*Vertices {len(nodes)}\n")
            for n in nodes:
                vals = [str(G.nodes[n].get(v, 0)) for v in vector_attrs]
                f.write(" ".join(vals) + "\n")


def save_graph(G, path, file_format="pajek", partition_attrs=None, vector_attrs=None):
    """
    Save graph in various formats (pajek, graphml, gexf, gml, adjlist, edgelist).
    """
    fmt = file_format.lower()
    basename, _ = os.path.splitext(path)
    if fmt == "pajek":
        save_to_pajek(G, basename, partition_attrs, vector_attrs)
    elif fmt == "graphml":
        nx.write_graphml(G, path)
    elif fmt == "gexf":
        nx.write_gexf(G, path)
    elif fmt == "gml":
        nx.write_gml(G, path)
    elif fmt == "adjlist":
        nx.write_adjlist(G, path)
    elif fmt == "edgelist":
        nx.write_edgelist(G, path)
    else:
        raise ValueError(f"Unsupported format: {file_format}")


def compute_basic_stats(G):
    """
    Compute basic statistics of the graph.

    Parameters
    ----------
    G : networkx.Graph

    Returns
    -------
    dict
        Contains num_nodes, num_edges, density, avg_degree, avg_clustering, num_connected_components, largest_cc_size, diameter, avg_shortest_path_length.
    """
    n = G.number_of_nodes()
    m = G.number_of_edges()
    density = nx.density(G)
    degrees = [d for _, d in G.degree()]
    avg_degree = sum(degrees) / n if n else 0
    avg_clustering = nx.average_clustering(G)
    cc = list(nx.connected_components(G))
    num_cc = len(cc)
    largest_cc = max(cc, key=len) if cc else set()
    sub = G.subgraph(largest_cc)
    try:
        diameter = nx.diameter(sub)
        avg_sp = nx.average_shortest_path_length(sub)
    except (nx.NetworkXError, nx.NetworkXNoPath):
        diameter = None
        avg_sp = None
    return {
        "num_nodes": n,
        "num_edges": m,
        "density": density,
        "avg_degree": avg_degree,
        "avg_clustering": avg_clustering,
        "num_connected_components": num_cc,
        "largest_cc_size": len(largest_cc),
        "diameter": diameter,
        "avg_shortest_path_length": avg_sp
    }


def compute_centralities(G):
    """
    Compute centrality measures for the graph.

    Parameters
    ----------
    G : networkx.Graph

    Returns
    -------
    dict
        degree, closeness, betweenness, eigenvector centralities.
    """
    deg = nx.degree_centrality(G)
    clo = nx.closeness_centrality(G)
    bet = nx.betweenness_centrality(G)
    try:
        eig = nx.eigenvector_centrality(G, max_iter=1000)
    except nx.NetworkXError:
        eig = {}
    return {
        "degree_centrality": deg,
        "closeness_centrality": clo,
        "betweenness_centrality": bet,
        "eigenvector_centrality": eig
    }


def compute_cluster_metrics(G, partition_attr):
    """
    Compute density and average degree centrality for each cluster.

    Parameters
    ----------
    G : networkx.Graph
    partition_attr : str
        Node attribute name (with or without "partition_" prefix).

    Returns
    -------
    dict
        Mapping cluster ID to dict with keys:
        - "density": internal edge density of the cluster subgraph
        - "avg_degree_centrality": average degree centrality of nodes in the cluster
        - "size": number of nodes in the cluster
    """
    key = partition_attr if partition_attr.startswith("partition_") else f"partition_{partition_attr}"
    cent = nx.degree_centrality(G)
    clusters = {}
    for n, attrs in G.nodes(data=True):
        cid = attrs.get(key)
        clusters.setdefault(cid, []).append(n)
    metrics = {}
    for cid, nodes in clusters.items():
        sub = G.subgraph(nodes)
        density = nx.density(sub)
        avg_cent = sum(cent.get(n, 0) for n in nodes) / len(nodes) if nodes else 0
        metrics[cid] = {
            "density": density,
            "avg_degree_centrality": avg_cent,
            "size": len(nodes)
        }
    return metrics


def export_cluster_dataframe(G, partition_attr):
    """
    Export clusters to a pandas DataFrame.

    Parameters
    ----------
    G : networkx.Graph
    partition_attr : str
        Node attribute name (with or without "partition_" prefix).

    Returns
    -------
    pandas.DataFrame
        DataFrame with columns: cluster, items (all nodes sorted by descending degree centrality),
        density, centrality
    """
    key = partition_attr if partition_attr.startswith("partition_") else f"partition_{partition_attr}"
    metrics = compute_cluster_metrics(G, key)
    deg_c = nx.degree_centrality(G)
    rows = []
    for cid, m in metrics.items():
        nodes = [n for n, d in G.nodes(data=True) if d.get(key) == cid]
        sorted_nodes = sorted(nodes, key=lambda n: deg_c.get(n, 0), reverse=True)
        items_str = ";".join(str(n) for n in sorted_nodes)
        rows.append({
            "cluster": cid,
            "items": items_str,
            "density": m["density"],
            "centrality": m["avg_degree_centrality"]
        })
    return pd.DataFrame(rows)